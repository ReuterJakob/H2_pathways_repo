# Read techno-economic assumptions and transport distances for transport cost calculation
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from matplotlib import ticker as mtick
import os
from openpyxl import load_workbook
import xlwings as xl

#Directory

path_excel = r'\\dena.de\Daten\Home\Reuter\Desktop\H2_pathways_repo\data\raw\H2_supply_route_assessment_V2.xlsx'
path_csv = r'\\dena.de\Daten\Home\Reuter\Desktop\H2_pathways_repo\data\interim'
path_plt = r'\\dena.de\Daten\Home\Reuter\Desktop\H2_pathways_repo\Plots_v3'

# adjust plotsize and font
params = {'font.size':20,
'font.weight':'normal',
'font.family':'arial',
'lines.linewidth':2
    }
plt.rcParams.update(params)

def df_from_excel(path_excel):
    app = xl.App(visible=False)
    book = app.books.open(path_excel)
    book.save()
    app.kill()
    return pd.read_excel(path_excel)

df = df_from_excel(path_excel)

import policy

##Emission comparison
pipe_em_pol = policy.Pipeline_emissions

lh2_con_pol = policy.LH2_Conversion_emissions
Lh2_et_pol = policy.LH2_Export_terminal_emissions
lh2_ship_pol = policy.LH2_Shipping_emissions
Lh2_it_pol = policy.LH2_Import_terminal_emissions
lh2_recon_pol =policy.LH2_Reconversion_emissions

NH3_con_pol = policy.NH3_Conversion_emissions
NH3_et_pol = policy.NH3_Export_terminal_emissions
NH3_ship_pol = policy.NH3_Shipping_emissions
NH3_it_pol = policy.NH3_Import_terminal_emissions
NH3_recon_pol = policy.NH3_Reconversion_emissions

"""## Read inputs"""

prices = pd.read_excel(path_excel, sheet_name='Commodity Prices', decimal=',', index_col=0)

GHG = pd.read_excel(path_excel, sheet_name='GHG Footprint', decimal=',', index_col=0)

tea_blue = pd.read_excel(path_excel, sheet_name='LCOH_NGR', decimal=',', index_col=0)

lcoh_green_source = pd.read_excel(path_excel, sheet_name='LCOH_RES', decimal=',', index_col=0)

## Plot price inputs.
co2_prices_baseline = prices.loc['CO2 prices [€/t_CO2] Baseline', 2025:2050]
co2_prices_policy = prices.loc['CO2 prices [€/t_CO2] Policy',2025:2050]
el_prices_baseline_nor = prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline', 2025:2050]
el_prices_policy_nor = prices.loc['Electricity prices in Norway [€_2021/MWh] Policy',2025:2050]
el_prices_baseline_ger = prices.loc['Electricity prices in Germany [€_2020/MWh] Baseline', 2025:2050]
el_prices_policy_ger = prices.loc['Electricity prices in Germany [€_2020/MWh] Policy',2025:2050]
gas_prices_baseline_nor = prices.loc['Gas prices in NOR [€_2020/MWh] Baseline', 2025:2050]
gas_prices_policy_nor = prices.loc['Gas prices in NOR [€_2020/MWh] Policy',2025:2050]

fig, ax = plt.subplots(figsize=(10,4), frameon=False, layout = 'constrained')
plt.subplot(1,2,1)
plt.plot(co2_prices_baseline, color= 'red', linestyle= '-', label= ' CO2 Baseline')
plt.plot(co2_prices_policy, color='red', linestyle= '--',label='CO2 Policy')
plt.grid(True, axis = 'y')
plt.xlim(2025,2050)
plt.ylim(0,)
ax.set_axisbelow(True)
plt.ylabel('Prices [€/t CO2]')
plt.legend()

plt.subplot(1,2,2)
plt.plot(el_prices_baseline_ger, color= 'gold', linestyle= '-',label= ' Electricity GER Baseline')
plt.plot(el_prices_policy_ger, color='gold', linestyle= '--',label='Electricity GER Policy')
plt.plot(el_prices_baseline_nor, color= 'orange', linestyle= '-',label= 'Electricity NOR Baseline')
plt.plot(el_prices_policy_nor, color='orange', linestyle= '--',label='Electricity NOR Policy')
plt.plot(gas_prices_baseline_nor, color='blue', linestyle= '-',label='Gas NOR Baseline')
plt.plot(gas_prices_policy_nor, color='blue', linestyle= '--',label='Gas NOR Policy')
plt.ylabel('Prices [€/MWh]')
plt.grid(True, axis = 'y')
#ax.yaxis.set_major_locator(mtick.LinearLocator(18))
plt.ylim(0,)
plt.xlim(2025,2050)
ax.set_axisbelow(True)
plt.legend()

title = '\Inputs_prices'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

# Emission inputs
##read
EF_GER_base = GHG.loc['Grid electricity emission factor GER [g CO2eq/kWh] Baseline', 2025:2050]
EF_GER_pol = GHG.loc['Grid electricity emission factor GER [g CO2eq/kWh] Policy',2025:2050]
EF_NOR_base = GHG.loc['Grid electricity emission factor NOR [g CO2eq/kWh] Baseline', 2025:2050]
EF_NOR_pol = GHG.loc['Grid electricity emission factor NOR [g CO2eq/kWh] Policy',2025:2050]

Blue_em_NOR_base = GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline', 2025:2050]
Blue_em_NOR_pol = GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Policy',2025:2050]

##plot
fig, ax = plt.subplots(figsize=(10,4), frameon=False, layout = 'constrained')
plt.subplot(1,2,1)
plt.plot(EF_GER_base, color= 'gold', linestyle= '-', label= 'GER Baseline')
plt.plot(EF_GER_pol, color='gold', linestyle= '--',label= 'GER Policy')
plt.plot(EF_NOR_base, color= 'orange', linestyle= '-', label= 'NOR Baseline')
plt.plot(EF_NOR_pol, color='orange', linestyle= '--',label= 'NOR Policy')
plt.grid(True, axis = 'y')
plt.xlim(2025,2050)
plt.ylim(0,)
ax.set_axisbelow(True)
plt.ylabel('Grid electricity emission factor [g CO2eq/kWh]')
plt.legend()
#blue hydrogen emission assumptions
plt.subplot(1,2,2)
plt.plot(Blue_em_NOR_base, color= 'blue', linestyle= '-', label= 'Baseline')
plt.plot(Blue_em_NOR_pol, color='blue', linestyle= '--',label='Policy')
plt.grid(True, axis = 'y')
plt.xlim(2025,2050)
plt.ylim(0,)
ax.set_axisbelow(True)
plt.ylabel('Blue hydrogen emissions [kg CO2/kg H2]')
plt.legend()

title = '\Inputs_Emission'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

# Transport inputs

availabilities = pd.read_excel(path_excel, sheet_name='General Assumptions', decimal=',', index_col=0)

this_year = int(availabilities.loc['Technology availability']['Value'])

AV_pipe_new = int(availabilities.loc['New H2 pipeline - time from FID to commissioning [years] Baseline']['Value'])

AV_pipe_retro = int(availabilities.loc['Retrofit H2 pipeline - availability from now [years] Baseline']['Value'])

AV_LH2 = int(availabilities.loc['LH2 shipping - availability']['Value'])

AV_NH3 = int(availabilities.loc['Ammonia shipping - availability']['Value'])

av_new = this_year + AV_pipe_new
av_retro = this_year + AV_pipe_retro
"""# Production emissions

## Definition of variables
"""

# LHV H2 in moles per MJ
LHV_H2_moles_MJ = float(GHG.loc['H2 LHV [mole/MJ]']['Value'])


# CO2 produced during SMR @ 1 mole CO2 per 4 moles H2 [moles CO2/MJ]
CO2_SMR_mole = LHV_H2_moles_MJ * 1/4


# Molecular weight of CO2 [g/Mole]
CO2_g_mole = float(GHG.loc['CO2 [g/Mole]']['Value'])


# Capture rate syngas [%] low
capture_rate_low = float(GHG.loc['Capture rate [%] low']['Value'])


# Capture rate syngas [%] mid
capture_rate_mid = float(GHG.loc['Capture rate [%] mid']['Value'])


# Capture rate syngas [%] high
capture_rate_high = float(GHG.loc['Capture rate [%] high']['Value'])


# Molecular weight of CH4 [g/Mole]
CH4_g_mole = float(GHG.loc['CH4 [g/Mole]']['Value'])


# Heat consumption to drive SMR in [MJ/mole_H2]
Drive_Energy_MJ_mole_h2 = float(GHG.loc['Heat Input [MJ/mole_H2]']['Value'])


# Emission intensity of CH4 [g CO2/MJ]
emission_intensity_CH4 = float(GHG.loc['Combustion emissions CH4 [g CO2/MJ]']['Value'])


# Flue gas capture rate [%] low
capture_rate_flue_gas_low = float(GHG.loc['Capture rate [%] flue gas low']['Value'])


# Flue gas capture rate [%] high
capture_rate_flue_gas_high = float(GHG.loc['Capture rate [%] flue gas high']['Value'])


#GWP20 of methane
GWP20_CH4 = float(GHG.loc['CH4 GWP20 [Years]']['Value'])


#GWP100 of methane
GWP100_CH4 = float(GHG.loc['CH4 GWP100 [Years]']['Value'])


# Methane leakage rate in %
leakage_rate_NOR_Base = float(GHG.loc['Upstream methane leakage rate [%] NOR Baseline']['Value'])
# Methane leakage rate in %
leakage_rate_NOR_Policy = float(GHG.loc['Upstream methane leakage rate [%] NOR Policy']['Value'])

# Methane leakage rate in %
leakage_rate_US_Base = float(GHG.loc['Upstream methane leakage rate [%] US Baseline']['Value'])

# Methane leakage rate in %
leakage_rate_US_Policy = float(GHG.loc['Upstream methane leakage rate [%] US Policy']['Value'])

"""## Direct emissions from SMR"""

# CO2 emissions from SMR process [g CO2/MJ]
Direct_emissions_grey = CO2_g_mole * CO2_SMR_mole


# CO2 emissions from SMR @ 55% capture rate [g CO2/MJ]
Direct_emissions_low = Direct_emissions_grey * (1-capture_rate_low)


# CO2 emissions from SMR @ 80% capture rate [g CO2/MJ]
Direct_emissions_mid = Direct_emissions_grey * (1-capture_rate_mid)


# CO2 emissions from SMR @ 95% capture rate [g CO2/MJ]
Direct_emissions_high = Direct_emissions_grey * (1-capture_rate_high)


#CH4 consumed during SMR @ 1 mole CH4 per 4 moles H2 [g CH4/MJ]
CH4_SMR_grams = 1.03 * CH4_g_mole


"""Emissions from energy to drive SMR

## w/o flue gas capture (Grey)
"""

# When burning natural gas for heat production CO2 is emitted [CO2/mole_H2]
Drive_CO2_emissions_mole = Drive_Energy_MJ_mole_h2 * emission_intensity_CH4


# CO2 emissions per MJ hydrogen in order to create heat and pressure - w/o flue gas capture [g CO2/MJ_H2]
Drive_CO2_emissions_grams_grey = Drive_CO2_emissions_mole * LHV_H2_moles_MJ


"""## w/o flue gas capture (blue)"""

# CO2 emissions drive with flue gas capture - low [g CO2/MJ]
Drive_CO2_emissions_grams_blue_flue_low = Drive_CO2_emissions_grams_grey * (1 - capture_rate_flue_gas_low)


# CO2 emissions drive with flue gas capture - high [g CO2/MJ]
Drive_CO2_emissions_grams_blue_flue_high = Drive_CO2_emissions_grams_grey * (1 - capture_rate_flue_gas_high)


"""## Check this cell!!"""

# CH4 consumed to drive the process [g CH4/MJ_H2]
Drive_CH4_consumed = Drive_Energy_MJ_mole_h2 * 1/CO2_g_mole * CH4_g_mole


"""## Total direct CO2 emissions

### Without flue gas capture
"""

# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_grey = Direct_emissions_grey + Drive_CO2_emissions_grams_grey


# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_blue_low_no_fluegas = Direct_emissions_low + Drive_CO2_emissions_grams_grey


# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_blue_mid_no_fluegas = Direct_emissions_mid + Drive_CO2_emissions_grams_grey


# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_blue_high_no_fluegas = Direct_emissions_high + Drive_CO2_emissions_grams_grey


"""## With flue gas capture

### low (65%)
"""

# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_blue_low_with_fluegas_low = Direct_emissions_low + Drive_CO2_emissions_grams_blue_flue_low


# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_blue_mid_with_fluegas_low = Direct_emissions_mid + Drive_CO2_emissions_grams_blue_flue_low


# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_blue_high_with_fluegas_low = Direct_emissions_high + Drive_CO2_emissions_grams_blue_flue_low


"""### high (90%)"""

# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_blue_low_with_fluegas_high = Direct_emissions_low + Drive_CO2_emissions_grams_blue_flue_high


# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_blue_mid_with_fluegas_high = Direct_emissions_mid + Drive_CO2_emissions_grams_blue_flue_high


# Total CO2 emissions to produce grey hydrogen [g CO2/MJ_H2]
Total_CO2_emissions_blue_high_with_fluegas_high = Direct_emissions_high + Drive_CO2_emissions_grams_blue_flue_high


"""## Indirect emissions from natural gas transport and storage"""

# Indirect emissions from natural gas transport and storage [g CO2/MJ_H2]
Indirect_emissions = Total_CO2_emissions_grey * 0.075




"""## CH4 upstream emissions"""

# CH4 consumption as feedstock and energy to drive SMR [g CH4/MJ_H2]
Total_methane_SMR = CH4_SMR_grams + Drive_CH4_consumed


# Quantity of methane consumed to produce grey hydrogen [g CH4/MJ_H2]
upstream_emissions_CH4_low = leakage_rate_NOR_Policy * Total_methane_SMR


# Quantity of methane consumed to produce grey hydrogen [g CH4/MJ_H2]
upstream_emissions_CH4_mid = leakage_rate_US_Base * Total_methane_SMR


# Quantity of methane consumed to produce grey hydrogen [g CH4/MJ_H2]
upstream_emissions_CH4_high = leakage_rate_US_Policy * Total_methane_SMR


# Upstream emissions at GWP20 of CH4 [g CO2e/MJ_H2]
upstream_emissions_GWP20_NOR_Policy = upstream_emissions_CH4_low * GWP20_CH4
upstream_emissions_GWP20_NOR_Base = leakage_rate_NOR_Base * Total_methane_SMR * GWP20_CH4

# Upstream emissions at GWP20 of CH4 [g CO2e/MJ_H2]
upstream_emissions_GWP20_NOR_Policy = upstream_emissions_CH4_low * GWP20_CH4
upstream_emissions_GWP20_NOR_Policy = leakage_rate_NOR_Policy * Total_methane_SMR * GWP20_CH4

# Upstream emissions at GWP100 of CH4 [g CO2e/MJ_H2]
upstream_emissions_GWP100_NOR_Policy = upstream_emissions_CH4_low * GWP100_CH4
upstream_emissions_GWP100_NOR_Base = leakage_rate_NOR_Base * Total_methane_SMR * GWP100_CH4

# Upstream emissions at GWP100 of CH4 [g CO2e/MJ_H2]
upstream_emissions_GWP100_NOR_Policy = upstream_emissions_CH4_low * GWP100_CH4
upstream_emissions_GWP100_NOR_Policy = leakage_rate_NOR_Policy * Total_methane_SMR * GWP100_CH4

# Upstream emissions at GWP20 of CH4 [g CO2e/MJ_H2]
upstream_emissions_GWP20_mid = upstream_emissions_CH4_mid * GWP20_CH4


# Upstream emissions at GWP20 of CH4 [g CO2e/MJ_H2]
upstream_emissions_GWP20_high = upstream_emissions_CH4_high * GWP20_CH4


# Upstream emissions at GWP100 of CH4 [g CO2e/MJ_H2]
upstream_emissions_GWP100_low = upstream_emissions_CH4_low * GWP100_CH4


# Upstream emissions at GWP100 of CH4 [g CO2e/MJ_H2]
upstream_emissions_GWP100_mid = upstream_emissions_CH4_mid * GWP100_CH4


# Upstream emissions at GWP100 of CH4 [g CO2e/MJ_H2]
upstream_emissions_GWP100_high = upstream_emissions_CH4_high * GWP100_CH4


"""## Total emissions"""

# Total CO2 emissions including indirect emissions from natural gas transport and storage
Total_CO2_emissions = Total_CO2_emissions_grey + Indirect_emissions


"""## w/o flue gas capture

#### @ GWP20
"""

# Total Emissions for grey hydrogen: CO2 + fugitive CH4 [g CO2e/MJ]
Total_emissions_grey_GWP20_low = Total_CO2_emissions_grey + Indirect_emissions + upstream_emissions_GWP20_NOR_Policy
Total_emissions_grey_GWP20_low

# Total Emissions for grey hydrogen: CO2 + fugitive CH4 [g CO2e/MJ]
Total_emissions_grey_GWP20_mid = Total_CO2_emissions_grey + Indirect_emissions + upstream_emissions_GWP20_mid
Total_emissions_grey_GWP20_mid

# Total Emissions for grey hydrogen: CO2 + fugitive CH4 [g CO2e/MJ]
Total_emissions_grey_GWP20_high = Total_CO2_emissions_grey + Indirect_emissions + upstream_emissions_GWP20_high
Total_emissions_grey_GWP20_high

"""### @ GWP100"""

# Total Emissions for grey hydrogen: CO2 + fugitive CH4 [g CO2e/MJ]
Total_emissions_grey_GWP100_low = Total_CO2_emissions + upstream_emissions_GWP100_low
Total_emissions_grey_GWP100_low

# Total Emissions for grey hydrogen: CO2 + fugitive CH4 [g CO2e/MJ]
Total_emissions_grey_GWP100_mid = Total_CO2_emissions + upstream_emissions_GWP100_mid
Total_emissions_grey_GWP100_mid

# Total Emissions for grey hydrogen: CO2 + fugitive CH4 [g CO2e/MJ]
Total_emissions_grey_GWP100_high = Total_CO2_emissions + upstream_emissions_GWP100_high
Total_emissions_grey_GWP100_high

"""## w/ flue gas capture

### @GWP20
"""

# Total Emissions for grey hydrogen: CO2 + fugitive CH4 [g CO2e/MJ]
Total_emissions_grey_GWP20_low = Total_CO2_emissions + upstream_emissions_GWP20_NOR_Policy
Total_emissions_grey_GWP20_low

# Total Emissions for grey hydrogen: CO2 + fugitive CH4 [g CO2e/MJ]
Total_emissions_grey_GWP20_mid = Total_CO2_emissions + upstream_emissions_GWP20_mid
Total_emissions_grey_GWP20_mid

# Total Emissions for grey hydrogen: CO2 + fugitive CH4 [g CO2e/MJ]
Total_emissions_grey_GWP20_high = Total_CO2_emissions + upstream_emissions_GWP20_high
Total_emissions_grey_GWP20_high

"""## Energy to power carbon capture: xx none according to Bauer et al."""



"""flue gas capture (FGC)"""



Direct_emissions = [Direct_emissions_grey, Direct_emissions_grey, Direct_emissions_grey, Direct_emissions_grey, Direct_emissions_grey, Direct_emissions_grey, Direct_emissions_grey, Direct_emissions_grey,
                    Direct_emissions_high, Direct_emissions_high, Direct_emissions_high, Direct_emissions_high,
                    Direct_emissions_high, Direct_emissions_high, Direct_emissions_high, Direct_emissions_high]

Drive_emissions_plt = [Drive_CO2_emissions_grams_grey, Drive_CO2_emissions_grams_grey, Drive_CO2_emissions_grams_grey, Drive_CO2_emissions_grams_grey,
                       Drive_CO2_emissions_grams_grey, Drive_CO2_emissions_grams_grey, Drive_CO2_emissions_grams_grey, Drive_CO2_emissions_grams_grey,
                       Drive_CO2_emissions_grams_grey, Drive_CO2_emissions_grams_grey, Drive_CO2_emissions_grams_grey, Drive_CO2_emissions_grams_grey,
                       Drive_CO2_emissions_grams_blue_flue_high, Drive_CO2_emissions_grams_blue_flue_high, Drive_CO2_emissions_grams_blue_flue_high, Drive_CO2_emissions_grams_blue_flue_high]

Indirect_emissions_plt = [Indirect_emissions] * 16

Upstream_emissions = [upstream_emissions_GWP100_NOR_Base, upstream_emissions_GWP20_NOR_Base, upstream_emissions_GWP100_NOR_Policy, upstream_emissions_GWP20_NOR_Policy,
                      upstream_emissions_GWP100_mid, upstream_emissions_GWP20_mid, upstream_emissions_GWP100_high, upstream_emissions_GWP20_high,
                      upstream_emissions_GWP20_NOR_Base, upstream_emissions_GWP20_NOR_Policy, upstream_emissions_GWP20_mid, upstream_emissions_GWP20_high,
                      upstream_emissions_GWP20_NOR_Base, upstream_emissions_GWP20_NOR_Policy, upstream_emissions_GWP20_mid, upstream_emissions_GWP20_high]

"""## Calc. total emissions for each case"""

Grey_GWP100 = Direct_emissions_grey + Drive_CO2_emissions_grams_grey + Indirect_emissions + upstream_emissions_GWP100_low
Grey_GWP100

Grey_GWP20 = Direct_emissions_grey + Drive_CO2_emissions_grams_grey + Indirect_emissions + upstream_emissions_GWP20_NOR_Policy
Grey_GWP20

pess = Direct_emissions_low + Drive_CO2_emissions_grams_grey + Indirect_emissions + upstream_emissions_GWP20_NOR_Policy
pess

opt = Direct_emissions_high + Drive_CO2_emissions_grams_blue_flue_high + Indirect_emissions + upstream_emissions_GWP20_NOR_Policy
opt

"""## Plot total emissions"""

x =['Grey, GWP100\n0.02%', 'Grey, GWP20\n0.02%', 'Grey, GWP100\n1.3%', 'Grey, GWP20\n1.3%', 'Grey, GWP100\n3.7%', 'Grey, GWP20\n3.7%', 'Grey, GWP100\n9%', 'Grey, GWP20\n9%',
    'Blue,  SGC 90%\n0.02%', 'Blue, SCG 90%\n1.3%','Blue, SGC 90%\n3.7%', 'Blue, SGC 90%\n9%',
    'Blue\nSGC 90%, FGC 95%, 0.02%','Blue\nSGC 95%, FGC 90%, 1.3%','Blue\nSGC 95%, FGC 90%, 3.7%', 'Blue\nSGC 95%, FGC 90%, 9%']

# Creating a stacked bar chart to display emissions. Adding lists for the bottom method.
fig, ax = plt.subplots(figsize=(10,4), frameon=False, layout = 'constrained')
width = 0.5
#y_axis = np.arange(101,step=10)
direct = plt.bar(x, Direct_emissions, width, color = 'royalblue', label='CO2 from SMR', bottom=list(map(lambda x, y, z: x + y + z, Upstream_emissions, Indirect_emissions_plt, Drive_emissions_plt)))
drive = plt.bar(x, Drive_emissions_plt, width, color = 'dodgerblue',label='CO2 from NG combustion to drive SMR', bottom=list(map(lambda x, y: x + y, Upstream_emissions, Indirect_emissions_plt)))
indirect = plt.bar(x, Indirect_emissions, width, color = 'grey', label = 'CO2 from NG infrastructure', bottom=Upstream_emissions)
upstream = plt.bar(x, Upstream_emissions, width, color = 'lightcoral',label = 'CH4 from fugitive upstream methane')
plt.grid(True, axis = 'y')
ax.set_ylabel(ylabel= ('[g CO2eq/MJ H2]'))
ax.yaxis.set_major_locator(mtick.LinearLocator(18))
ax.set_ylim(0,250)
ax.set_axisbelow(True)
ax2 = ax.secondary_yaxis('right', functions=(lambda MJ: MJ*120/1000, lambda kg: kg/120))
ax2.yaxis.set_major_locator(mtick.LinearLocator(18))
ax2.set_ylim(0,30)
ax2.set_ylabel(ylabel= '[kg CO2eq/kg H2]')
plt.xticks(rotation = 90)
plt.legend(loc='upper right')
plt.title('GHG emissions from blue hydrogen production', fontweight = 'bold')


title = '\Blue_emission_breakdown'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()
#ax.set_ylabel('kg CO2/kg H2')

"""Interpolated blue hydrogen production emissions starting from 55% total system capture rate -> up to 81%

## Sensitivity analysis

### Leakage rate
"""

leakage_rate = np.arange(0,0.105,0.005)
sensitivity = []
# calculate emissions from blue hydrogen production in [g CO2eq/MJ H2]
def blue_h2_emissions(Direct_emissions_grey, capture_rate_high, Drive_CO2_emissions_grams_grey, capture_rate_flue_gas_high, Indirect_emissions, leakage_rate_NOR_Base, Total_methane_SMR, GWP20_CH4):

    for i in leakage_rate:
        result = (Direct_emissions_grey * (1 - capture_rate_high)) + (Drive_CO2_emissions_grams_grey * (1 - capture_rate_flue_gas_high)) + Indirect_emissions + i * Total_methane_SMR * GWP20_CH4

        sensitivity.append(result)


    return sensitivity

blue_h2_emissions(Direct_emissions_grey, capture_rate_high, Drive_CO2_emissions_grams_grey, capture_rate_flue_gas_high, Indirect_emissions, leakage_rate_NOR_Policy, Total_methane_SMR, GWP20_CH4)

leakage_sensi_GWP20 = pd.DataFrame(sensitivity, index=leakage_rate, columns=['Total_Emissions [g CO2eq/MJ H2]'])


output_file = os.path.join(path_csv, 'leakage_sensi_GWP20' + '.csv')
leakage_sensi_GWP20.to_csv(output_file, sep = ';')

# calculate emissions from blue hydrogen production in [g CO2eq/MJ H2]
sensitivity = []
def blue_h2_emissions(Direct_emissions_grey, capture_rate_high, Drive_CO2_emissions_grams_grey, capture_rate_flue_gas_high, Indirect_emissions, leakage_rate_NOR_Base, Total_methane_SMR, GWP20_CH4):

    for i in leakage_rate:
        result = (Direct_emissions_grey * (1 - capture_rate_high)) + (Drive_CO2_emissions_grams_grey * (1 - capture_rate_flue_gas_high)) + Indirect_emissions + i * Total_methane_SMR * GWP100_CH4

        sensitivity.append(result)


    return sensitivity

blue_h2_emissions(Direct_emissions_grey, capture_rate_high, Drive_CO2_emissions_grams_grey, capture_rate_flue_gas_high, Indirect_emissions, leakage_rate_NOR_Policy, Total_methane_SMR, GWP20_CH4)

leakage_sensi_GWP100 = pd.DataFrame(sensitivity, index=leakage_rate, columns=['Total_Emissions [g CO2eq/MJ H2]'])

output_file = os.path.join(path_csv, 'leakage_sensi_GWP100' + '.csv')
leakage_sensi_GWP100.to_csv(output_file, sep = ';')

"""### Capture rate"""

capture_rates = np.arange(0.5, 1.01, 0.1)
sensitivity = []
# calculate emissions from blue hydrogen production in [g CO2eq/MJ H2]
def blue_h2_emissions(Direct_emissions_grey, capture_rate_high, Drive_CO2_emissions_grams_grey, capture_rate_flue_gas_high, Indirect_emissions, leakage_rate_NOR_Base, Total_methane_SMR, GWP20_CH4):

    for i in capture_rates:

        result = (Direct_emissions_grey * (1-i)) + (Drive_CO2_emissions_grams_grey * (1 - i)) + Indirect_emissions + 0.0002 * Total_methane_SMR * GWP20_CH4

        sensitivity.append(result)

    return sensitivity

blue_h2_emissions(Direct_emissions_grey, capture_rate_high, Drive_CO2_emissions_grams_grey, capture_rate_flue_gas_high, Indirect_emissions, leakage_rate_NOR_Policy, Total_methane_SMR, GWP20_CH4)

capture_sensi = pd.DataFrame(sensitivity, index=capture_rates, columns=['Total_Emissions [g CO2eq/MJ H2]'])
#to_csv
output_file = os.path.join(path_csv, 'capture_sensi' + '.csv')
capture_sensi.to_csv(output_file, sep = ';')

# Plot leakage and capture sensi
fig = plt.figure(figsize=(10,4))

#leakage rates
ax = fig.add_subplot(1, 2, 1)
ax2 = ax.secondary_yaxis('right', functions=(lambda MJ: MJ*120/1000, lambda kg: kg/120))
ax2.set_yticks(np.arange(0,19,3))
ax.locator_params(axis='y', nbins=7)
plt.plot(leakage_sensi_GWP20, color='blue', linestyle='solid', label = 'GWP20' )
plt.plot(leakage_sensi_GWP100, color='dodgerblue',linestyle='-', label = 'GWP100')

plt.grid(True, axis='y')
ax.set_axisbelow(True)


plt.xticks(np.arange(0,0.105, 0.02), ['0%', '2%', '4%', '6%', '8%', '10%'] )
plt.xlim(0,0.1)
plt.ylim(0,150)
plt.ylabel('[g CO2eq/MJ H2]')
plt.xlabel('Leakage rate')
plt.legend()

#Capture rates
ax3 = fig.add_subplot(1, 2, 2)
ax4 = ax3.secondary_yaxis('right', functions=(lambda MJ: MJ*120/1000, lambda kg: kg/120))

ax4.locator_params(axis='y', nbins=7)
ax4.set_ylabel(ylabel= '[kg CO2eq/kg H2]')

plt.plot(capture_sensi, color='blue', linestyle='solid')
ax3.set_yticks(np.arange(0,51, (50/6)))
plt.grid(True, axis='y')
ax.set_axisbelow(True)
plt.xlabel('System capture rate')
plt.xticks(capture_rates, ['50%',  '60%',  '70%',  '80%',  '90%',  '100%'] )
plt.xlim(0.5,1)
plt.ylim(0,50)
fig.tight_layout(pad=1)

title = '\Leakage_and_capture_sensi'
plt.savefig(path_plt+title+'.png', transparent = True)
plt.show()


# single Plot capture sensi
fig, ax = plt.subplots(figsize = (5,4), layout = 'constrained')
plt.plot(capture_sensi, color='blue', linestyle='solid')
plt.grid(True, axis='y')
ax.set_axisbelow(True)

ax5 = ax.secondary_yaxis('right', functions=(lambda MJ: MJ*120/1000, lambda kg: kg/120))
ax5.set_ylabel(ylabel= '[kg CO2eq/kg H2]')
plt.ylabel('[g CO2eq/MJ H2]')
plt.xlabel('System capture rate')

plt.xticks(capture_rates, ['50%',  '60%',  '70%',  '80%',  '90%',  '100%'] )

title = '\Capture_sensi'
plt.savefig(path_plt+title+'.png', transparent = True)
plt.show()

"""# Production cost"""

# Interest rate (WACC) in %
i = float(tea_blue.loc['Discount rate [%]']['NGR with CCS'])
i

# Economic lifetime of the plant in years
l_ngr = float(tea_blue.loc['Lifetime [Years]']['NGR with CCS'])


opex_share = float(tea_blue.loc['Opex [% of Capex]']['NGR with CCS'])
opex_share

# Calculate the amortisation factor alpha
alpha_ngr = (i * (1 + i) ** l_ngr) / (((1 + i) ** l_ngr) - 1)

#round(alpha, 2)

CF = float(tea_blue.loc['Availability [%]']['NGR with CCS'])
CF

# Plant efficiency in %
n = float(tea_blue.loc['Efficiency [%]']['NGR with CCS'])
n

# LHV of hydrogen is 33.33 kWh/kg
LHV_h2 = float(tea_blue.loc['LHV H2 [kWh/kg]']['NGR with CCS'])
LHV_h2

"""## Calc. blue LCOH

Definition of the cost calculation function for LCOH from NGR. Time relevant variables = capex, opex, P_ng, P_co2
LHV H2 [kWh/kg]
capex_y  [€/kW]
opex_y [€/kW/a]
CF [%]
P_ng_y [€/MWh]
Q_ce [kgCO2/kgH2]
Q_ue [kgCO2/kgH2]
P_ccs [€/t CO2]
P_co2_y [€/t CO2]
"""
#LCOH Baseline NOR
def calculate_lcoh_ngr():

    result = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760) + P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y + Q_ue_y * P_co2_y) / 1000)

    return result

# Calculation of LCOH from NGR for every year from 2025 to 2050.
years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['LCOH_blue'])
result.index.name = 'Years'


for year in years:

    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][year])
    P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][year])
    P_co2_y = float(prices.loc['CO2 prices [€/t_CO2] Baseline'][year])
    P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][year])

    # calculate lcoe of specific year
    result.LCOH_blue.loc[year] = calculate_lcoh_ngr()

result
# Create csv file from results dataframe
output_file = os.path.join(path_csv,'LCOH_blue_NOR_Base.csv')
result.to_csv(output_file, sep = ';')
LCOH_blue_NOR_Base = result

'''
# Cost components blue hydrogen
'''
def calculate_lcoh_ngr_components():

    result = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y + Q_ue_y * P_co2_y ) / 1000)
    return  result

#capex
components = []
for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = 0#float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][year])
    P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][year])
    P_co2_y = float(prices.loc['CO2 prices [€/t_CO2] Baseline'][year])
    P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][year])
    
    result = calculate_lcoh_ngr_components()
    components.append(result)

LCOH_blue_capex = np.subtract(LCOH_blue_NOR_Base['LCOH_blue'], components)
#opex
components = []
for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = 0  #capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][year])
    P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][year])
    P_co2_y = float(prices.loc['CO2 prices [€/t_CO2] Baseline'][year])
    P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][year])

    result = calculate_lcoh_ngr_components()
    components.append(result)

lcoh_blue_opex = np.subtract(LCOH_blue_NOR_Base['LCOH_blue'], components)
# captured emissions
components = []
for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = 0  #float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][year])
    P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][year])
    P_co2_y = float(prices.loc['CO2 prices [€/t_CO2] Baseline'][year])
    P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][year])

    result = calculate_lcoh_ngr_components()
    components.append(result)

lcoh_blue_Q_ce_y = np.subtract(LCOH_blue_NOR_Base['LCOH_blue'], components)
# uncaptured emissions
components = []
for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][year])
    Q_ue_y = 0  #float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][year])
    P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][year])
    P_co2_y = float(prices.loc['CO2 prices [€/t_CO2] Baseline'][year])
    P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][year])

    result = calculate_lcoh_ngr_components()
    components.append(result)

lcoh_blue_Q_ue_y = np.subtract(LCOH_blue_NOR_Base['LCOH_blue'], components)
# NG price
components = []
for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][year])
    P_ng_y = 0  #float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][year])
    P_co2_y = float(prices.loc['CO2 prices [€/t_CO2] Baseline'][year])
    P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][year])

    result = calculate_lcoh_ngr_components()
    components.append(result)

lcoh_blue_P_ng_y = np.subtract(LCOH_blue_NOR_Base['LCOH_blue'], components)
# CO2 price
components = []
for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][year])
    P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][year])
    P_co2_y = 0  #float(prices.loc['CO2 prices [€/t_CO2] Baseline'][year])
    P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][year])

    result = calculate_lcoh_ngr_components()
    components.append(result)

lcoh_blue_P_co2_y = np.subtract(LCOH_blue_NOR_Base['LCOH_blue'], components)
# CCS cost
components = []
for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][year])
    P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][year])
    P_co2_y = float(prices.loc['CO2 prices [€/t_CO2] Baseline'][year])
    P_ccs_y = 0  #float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][year])

    result = calculate_lcoh_ngr_components()
    components.append(result)

lcoh_blue_P_ccs_y = np.subtract(LCOH_blue_NOR_Base['LCOH_blue'], components)

## Plot blue cost components
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
x = np.arange(2025,2051)


labels = ['Capex', 'Opex', 'Q_UE', 'Q_CE', 'P_NG', 'P_CO2', 'P_CCS']
#colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x,LCOH_blue_capex, lcoh_blue_opex, lcoh_blue_Q_ce_y, lcoh_blue_Q_ue_y, lcoh_blue_P_ng_y, lcoh_blue_P_co2_y, lcoh_blue_P_ccs_y, labels = labels )
plt.grid(True, axis='y')
plt.grid(True, axis='x')
ax.set_axisbelow(True)
plt.xlim(2025,2050)

plt.ylabel('LCOH [€/kg H2]')

title = 'LCOH_blue_components'
output_file = os.path.join(path_plt,title)
plt.savefig(output_file+'.png', transparent = True)

plt.show()



#LCOH NOR Policy
alpha_ngr = 0.08
def calculate_lcoh_ngr():

    result = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760) + P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y + Q_ue_y * P_co2_y) / 1000)

    return result

years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['LCOH_blue'])
result.index.name = 'Years'


for year in years:

    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Policy'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Policy'][year])
    P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Policy'][year])
    P_co2_y = float(prices.loc['CO2 prices [€/t_CO2] Policy'][year])
    P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Policy'][year])

    # calculate lcoe of specific year
    result.LCOH_blue.loc[year] = calculate_lcoh_ngr()

result
# Create csv file from results dataframe
output_file = os.path.join(path_csv,'LCOH_blue_NOR_Policy.csv')
result.to_csv(output_file, sep = ';')
LCOH_blue_NOR_Policy = result

# Plot cost curve of hydrogen production from NGR with CCS
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'blue', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
#plt.title('Cost curve for blue hydrogen production', fontweight='bold')
plt.ylabel('[€/kg H2]')

title = '\LLCOH_blue_NOR_Policy'
plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()


"""## Green LCOH"""
LCOH_green_NOR_Policy= ((lcoh_green_source.loc['Norway_Onshore_2_low_temp_optimistic', 2025:2050]).mul(0.89))


LCOH_green_NOR_Base = ((lcoh_green_source.loc['Norway_Onshore_2_low_temp_baseline', 2025:2050]).mul(0.89))
LCOH_green_Base = pd.DataFrame(LCOH_green_NOR_Base)
LCOH_green_Base.index.name = 'Years'
# Create csv file from results dataframe
output_file = os.path.join(path_csv,'LCOH_green_Base.csv')
LCOH_green_Base.to_csv(output_file, sep = ';')

# Plot cost curve of hydrogen production from RES
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(LCOH_green_Base, color = 'green', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
#plt.title('Cost curve for green hydrogen production', fontweight='bold')
#plt.xlabel('Year')
plt.ylabel('Cost [€/kg H2]')

title = '\LCOH_green_Base'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

# Green and blue LCOH
# Plot cost curves of hydrogen production from NGR with CCS and RES
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
plt.subplot(1,2,1)
#baseline
plt.plot(LCOH_green_NOR_Base, color ='green', linestyle ='-', label='Green H2')
plt.plot(LCOH_blue_NOR_Base, color ='blue', linestyle ='-', label='Blue H2')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
ax.legend()

plt.ylabel('LCOH [€/kg H2]')
plt.ylim(1.5,)
plt.xlim(2025,2050)
#policy
plt.subplot(1,2,2)
plt.plot(LCOH_green_NOR_Policy, color ='green', linestyle= '-', label='Green H2')
plt.plot(LCOH_blue_NOR_Policy, color ='blue', linestyle= '-', label='Blue H2')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
#plt.title('LCOH green and blue', fontweight='bold')
ax.legend()
#plt.xlabel('Year')
plt.ylabel('LCOH [€/kg H2]')
plt.ylim(1.5,)
plt.xlim(2025,2050)
title = '\LCOH_green_blue_base_pol'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

"""## LCOH blue Sensi Baseline"""
sensi_year = 2030

capex_y = float(tea_blue.loc['Capex [€/kW]'][sensi_year])
opex_y = capex_y * opex_share
Q_grey = float(GHG.loc['Grey hydrogen emissions [kg CO2/kg H2] - NOR Baseline']['Value'])
Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][sensi_year])
Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][sensi_year])
P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][sensi_year])
P_co2_y = float(prices.loc['CO2 prices [€/t_CO2] Baseline'][sensi_year])
P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][sensi_year])
'''
# Cost components blue hydrogen
'''
lcoh_blue = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y  + Q_ue_y * P_co2_y ) / 1000)

lcoh_blue_capex = float(LHV_h2 * ((alpha_ngr * capex_y* 0 + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y  + Q_ue_y * P_co2_y ) / 1000)

lcoh_blue_opex = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y * 0 ) / (CF * 8760) + P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y + Q_ue_y * P_co2_y) / 1000)

lcoh_blue_P_NG = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y ) / (CF * 8760) + P_ng_y * 0 / 1000 * n) + (Q_ce_y * P_ccs_y + Q_ue_y * P_co2_y) / 1000)

lcoh_blue_P_CCS = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y ) / (CF * 8760) + P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y * 0 + Q_ue_y * P_co2_y) / 1000)

lcoh_blue_P_CO2 = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y ) / (CF * 8760) + P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y  + Q_ue_y * P_co2_y * 0) / 1000)


## Green vs blue

# Green costs from EWI adjusted to EUR_2022
lcoh_green = pd.read_excel(path_excel, sheet_name='LCOH_RES', decimal=',', index_col=0)
LCOH_green_NOR_opt = (lcoh_green.filter(regex ='Norway.*', axis = 'index').filter(regex ='optimistic$', axis = 'index').filter(regex='low_temp*', axis= 'index').transpose().loc[2025:2050]).mul(0.89)
#Blue costs for 0-300€ Co2 price
# Calculation of LCOH from NGR for every year from 2025 to 2050.
years = np.arange(2025,2051)
n_years = len(years)
P_CO2_range = np.arange(0,501,100)
n_prices =len(P_CO2_range)
# get capex, opex, natural gas price and CO2 price of specific year


years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['LCOH_BLUE'])
result.index.name = 'Years'

def calculate_lcoh_ngr_sensi_P_CO2():

    result = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y + Q_ue_y * P_co2_y ) / 1000)
    return  result

sensitivity = []
for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2] - NOR Baseline'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline'][year])
    P_ng_y = float(prices.loc['Gas prices in NOR [€_2020/MWh] Baseline'][year])
    P_ccs_y = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2] Baseline'][year])
    for P_co2_y in P_CO2_range:
        result = calculate_lcoh_ngr_sensi_P_CO2()
        sensitivity.append(result)

#sensitivity = np.array(sensitivity)
sensitivity = np.array(sensitivity)
result = np.reshape(sensitivity, (n_years, n_prices))

# Create Dataframe
cols = P_CO2_range
rows = years
data = result

price_sensitivities = pd.DataFrame(
    data=data,
    index=rows,
    columns=cols,
)
price_sensitivities

# Plot green vs blue sensi
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
years = np.arange(2025,2051,5)
#plt.subplot(1,2,1)
plt.plot(LCOH_green_NOR_opt, color='green', linestyle='solid', label = 'H2 from RES')
plt.plot(price_sensitivities, color='blue', linestyle='solid', label = 'H2 from NGR with CCS')
#plt.plot(lcoh_ngr_sensi_P_Co2, color='dodgerblue',linestyle='-', label = 'CO2 price')
#plt.plot(lcoh_ngr_sensi_P_CCS, color='royalblue',linestyle='-', label = 'CCS cost')
plt.grid(True, axis='y')
plt.grid(True, axis='x')
ax.set_axisbelow(True)
plt.xlim(2025,2050)
#ax.xaxis.set_major_formatter(mtick.PercentFormatter(10, decimals=None))
#plt.locator_params(axis='x', nbins=10)
plt.ylabel('[€/kg H2]')
plt.xticks(rotation = 90)
#plt.xlabel('Change')
#plt.legend()

title = 'LCOH_green_blue_sensi_P_co2'
output_file = os.path.join(path_plt,title)
plt.savefig(output_file+'.png', transparent = True)

plt.show()

# Lifetime sensi

lifetime = np.arange(1,30,1)
sensitivity = []
def calculate_lcoh_ngr_sensi_Lifetime(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2):

    for l_ngr in lifetime:


        result = float(LHV_h2 * ((((i * (1 + i) ** l_ngr) / (((1 + i) ** l_ngr) - 1)) * capex_y + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y + Q_ue_y * P_co2_y ) / 1000)

        sensitivity.append(result)


    return sensitivity

calculate_lcoh_ngr_sensi_Lifetime(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2)

# save as df
lcoh_ngr_sensi_Lifetime= pd.DataFrame(sensitivity, index= lifetime, columns=['LCOH_NGR [€/kg H2]'])
lcoh_ngr_sensi_Lifetime.index.name = 'Lifetime of plant in years'

# Capture rate sensi on costs
capture_rates = np.arange(0, 1.01, 0.1)
sensitivity = []
def calculate_lcoh_ngr_sensi_Lifetime(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2):

    for rate in capture_rates:


        result = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + ((Q_grey *  rate) * P_ccs_y + (Q_grey  * (1- rate)) * P_co2_y ) / 1000)

        sensitivity.append(result)


    return sensitivity

calculate_lcoh_ngr_sensi_Lifetime(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2)

lcoh_ngr_sensi_capture_rate= pd.DataFrame(sensitivity, index= capture_rates, columns=['LCOH_NGR [€/kg H2]'])
lcoh_ngr_sensi_capture_rate.index.name = 'capture rate in %'
lcoh_ngr_sensi_capture_rate

#Plot lifetime and capture rate sensi LCOH blue
fig, ax = plt.subplots(figsize=(5,4), layout = 'constrained')
plt.plot(lcoh_ngr_sensi_Lifetime, color='blue', linestyle='solid')
plt.grid(True, axis='y')
plt.grid(True, axis='x')
plt.xlim(0,30)
ax.set_axisbelow(True)
#plt.locator_params(axis='x', nbins=5)
plt.ylabel('LCOH [€/kg H2]')
plt.xlabel('Plant lifetime in years')

title = 'LCOH_ngr_lifetime'
output_file = os.path.join(path_plt,title)
plt.savefig(output_file+'.png', transparent = True)

fig, ax = plt.subplots(figsize=(5,4), layout = 'constrained')
plt.plot(lcoh_ngr_sensi_capture_rate, color='blue', linestyle='solid')
plt.grid(True, axis='y')
plt.grid(True, axis='x')
ax.set_axisbelow(True)
plt.xticks([0,0.25,0.5,0.75,1], ['0%','25%', '50%', '75%', '100%'])
plt.locator_params(axis='x', nbins=5)
plt.xlim(0,1)
plt.ylabel('LCOH [€/kg H2]')
plt.xlabel('Capture rate')
#plt.legend()

title = 'LCOH_ngr_capture'
output_file = os.path.join(path_plt,title)
plt.savefig(output_file+'.png', transparent = True)

# P_NG sensi
sensitivity = []
def calculate_lcoh_ngr_sensi_P_NG(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2):

    for x in range(21):
        factor = x/10

        result = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760)+ P_ng_y * factor / 1000 * n) + (Q_ce_y * P_ccs_y + Q_ue_y * P_co2_y ) / 1000)

        sensitivity.append(result)


    return sensitivity

calculate_lcoh_ngr_sensi_P_NG(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2)
# save as df
lcoh_ngr_sensi_P_NG= pd.DataFrame(sensitivity, index= range(21), columns=['LCOH_NGR [€/kg H2]'])
lcoh_ngr_sensi_P_NG.index.name = 'P_NG_price_change in %'

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lcoh_ngr_sensi_P_NG_.csv')
lcoh_ngr_sensi_P_NG.to_csv(output_file, sep = ';')

# P_CO2 sensi
sensitivity = []
def calculate_lcoh_ngr_sensi_P_CO2():

    for x in range(21):
        factor = x/10

        result = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y + Q_ue_y * P_co2_y * factor) / 1000)

        sensitivity.append(result)


    return sensitivity

calculate_lcoh_ngr_sensi_P_CO2()

lcoh_ngr_sensi_P_CO2= pd.DataFrame(sensitivity, index= range(21), columns=['LCOH_NGR [€/kg H2]'])
lcoh_ngr_sensi_P_CO2.index.name = 'P_CO2_price_change in %'

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lcoh_ngr_sensi_P_CO2_.csv')
lcoh_ngr_sensi_P_CO2.to_csv(output_file, sep = ';')


# P_CCS sensi
sensitivity = []
def calculate_lcoh_ngr_sensi_P_CCS(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2):

    for x in range(21):
        factor = x/10

        result = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y * factor + Q_ue_y * P_co2_y ) / 1000)

        sensitivity.append(result)


    return sensitivity

calculate_lcoh_ngr_sensi_P_CCS(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2)

lcoh_ngr_sensi_P_CCS= pd.DataFrame(sensitivity, index= range(21), columns=['LCOH_NGR [€/kg H2]'])
lcoh_ngr_sensi_P_CCS.index.name = 'P_CCS_price_change in %'

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lcoh_ngr_sensi_P_CCS_.csv')
lcoh_ngr_sensi_P_CCS.to_csv(output_file, sep = ';')

# WACC sensi
sensitivity = []
def calculate_lcoh_ngr_sensi_WACC(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2):

    for x in np.arange(0.000000001,21,1):
        factor = x/10

        result = float(LHV_h2 * (((i* factor * (1 + i* factor) ** l_ngr) / (((1 + i* factor) ** l_ngr) - 1) * capex_y + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y  + Q_ue_y * P_co2_y ) / 1000)

        sensitivity.append(result)


    return sensitivity

calculate_lcoh_ngr_sensi_WACC(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2)

lcoh_ngr_sensi_WACC= pd.DataFrame(sensitivity, index= range(21), columns=['LCOH_NGR [€/kg H2]'])
lcoh_ngr_sensi_WACC.index.name = 'WACC_change in %'

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lcoh_ngr_sensi_WACC_.csv')
lcoh_ngr_sensi_WACC.to_csv(output_file, sep = ';')

# Capex sensi
sensitivity = []
def calculate_lcoh_ngr_sensi_capex(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2):

    for x in range(21):
        factor = x/10

        result = float(LHV_h2 * ((alpha_ngr * capex_y * factor + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs_y  + Q_ue_y * P_co2_y ) / 1000)

        sensitivity.append(result)


    return sensitivity

calculate_lcoh_ngr_sensi_capex(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs_y, Q_ue_y, P_co2_y, LHV_h2)

lcoh_ngr_sensi_capex= pd.DataFrame(sensitivity, index= range(21), columns=['LCOH_NGR [€/kg H2]'])
lcoh_ngr_sensi_capex.index.name = 'capex_change in %'

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lcoh_ngr_sensi_capex_.csv')
lcoh_ngr_sensi_capex.to_csv(output_file, sep = ';')

# Plot LCOH sensi
import matplotlib.ticker as mtick

fig, ax = plt.subplots(figsize=(5,4), layout = 'constrained')
#plt.subplot(1,2,1)


plt.plot(lcoh_ngr_sensi_P_NG, color='blue', linestyle='-', label = 'Gas Price [€/MWh]')
plt.plot(lcoh_ngr_sensi_P_CO2, color='dodgerblue',linestyle='-', label = 'CO2 Price [€/t CO2]')
plt.plot(lcoh_ngr_sensi_P_CCS, color='royalblue',linestyle='-', label = 'CCS Cost [€/t CO2]')
plt.plot(lcoh_ngr_sensi_WACC, color='aqua',linestyle='-', label = 'WACC [%]')
plt.plot(lcoh_ngr_sensi_capex, color='lightsteelblue',linestyle='-', label = 'CAPEX [€/kW]')
plt.grid(True, axis='y')
plt.grid(True, axis='x')
ax.set_axisbelow(True)
ax.xaxis.set_major_formatter(mtick.PercentFormatter(10, decimals=None))
plt.locator_params(axis='x', nbins=5)
plt.ylabel('LCOH [€/kg H2]')
plt.xlabel('Change')
plt.xlim(0,20)

plt.legend()

title = '\LCOH_blue_sensi'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

#Capture rates
plt.subplot(1,2,2)
plt.plot(capture_sensi, color='blue', linestyle='solid')
plt.grid(True, axis='y')
ax.set_axisbelow(True)
plt.ylabel('[g CO2eq/MJ H2]')
plt.xlabel('System capture rate')
#plt.xticks(np.arange(0.55,0.95, 0.1), ['55%', '60%', '65%', '70%', '75%', '80%', '85%', '90%', '95%'] )
plt.xticks(np.arange(0.55,1, 0.1), ['55%',  '65%',  '75%',  '85%',  '95%'] )


## Min. production cost & emissions

def choose_minimal_production_costs():
    result = min(LCOH_green, LCOH_blue)

    return result

def choose_minimal_production_cost_technology():
    if min(LCOH_green, LCOH_blue) == LCOH_green:
        result = 'Green'
    else:
        result = 'Blue'

    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Minimal_production_costs', 'Production_Technology'])
result.index.name = 'Years'
for year in years:
    # get all costs

    LCOH_green = float(LCOH_green_Base.loc[year]['Norway_Onshore_2_low_temp_baseline'])
    LCOH_blue = float(LCOH_blue_NOR_Base.loc[year]['LCOH_blue'])

    # calculate costs of specific year
    result.Minimal_production_costs.loc[year] = choose_minimal_production_costs()
    result.Production_Technology.loc[year] = choose_minimal_production_cost_technology()

result

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'LCOH_min_tech.csv')
result.to_csv(output_file, sep=',')
LCOH_min_tech = result

"""### Plot production costs"""



# Plot cost curves of hydrogen production from NGR with CCS and RES
fig, ax = plt.subplots(figsize=(10, 6), layout = 'constrained')

plt.plot(LCOH_green_Base, color='green', linestyle='solid', label='Green hydrogen')
plt.plot(LCOH_blue_NOR_Base, color='blue', linestyle='solid', label='Blue hydrogen')
plt.grid(True, axis='y')
ax.set_axisbelow(True)
#plt.title('Cost curves for green and blue hydrogen production', fontweight='bold')
ax.legend()
plt.ylabel('[€/kg H2]', fontweight='bold')
#plt.xticks(ticks=None, labels=np.arange(2025,2051,5))

title = '\Green_and_blue_costs'
plt.savefig(path_plt+title+'.png', transparent = True)


plt.show()

# Plot cost curve for production cost
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
#plt.plot(result, color='black', linestyle='solid')
plt.title('Minimal production costs in €/kg_H2')
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

"""##Baseline - Respective production emissions"""
blue_emissions = pd.DataFrame(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - NOR Baseline', 2025:2050])
blue_emissions.index.name = 'Years'
blue_emissions.columns= ['Emissions']
blue_emissions

"""
Emissions from blue hydrogen including h2 leakage
result = x.add(1*blue_prod_leakage*GWP20_H2)
result"""

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Blue_production_emissions_H2_leakage_Base.csv')
result.to_csv(output_file, sep=',')


years = np.arange(2025, 2051)
emissions = []
def append_emissions():
    for year in years:

        if LCOH_min_tech.loc[year]['Production_Technology'] == 'Green':
            result = 0
        else:
            result = blue_emissions.loc[year]['Emissions']

        emissions.append(result)

    return emissions

append_emissions()

LCOH_min_tech_em = LCOH_min_tech.assign(Emissions=emissions)

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'LCOH_min_tech_em.csv')
LCOH_min_tech_em.to_csv(output_file, sep=',')

"""# Transport cost"""

#Definition of variables for pipeline transport costs
tea_pipe = pd.read_excel(path_excel, sheet_name='Pipeline Transport', decimal=',', index_col=0)
tea_pipe

tra_d = pd.read_excel(path_excel, sheet_name='Transport Distances', decimal=',')
tra_d

prices = pd.read_excel(path_excel, sheet_name='Commodity Prices', index_col=0, decimal=',')
prices


# WACC for all transport investments (pipelines, terminals, liquefaction plants)
i_tra = float(tea_pipe.loc['Discount rate [%]']['Parameter'])
i_tra

i = i_tra

"""## Pipeline"""

# Economic lifetime for pipelines
l_pipe = float(tea_pipe.loc['Lifetime [Years]']['Parameter'])
l_pipe
# Amortisation factor for onshore pipelines
alpha_pipe = (i_tra * (1 + i_tra) ** l_pipe) / (((1 + i_tra) ** l_pipe) - 1)

# Utilisation of the pipeline in %
pipe_use = float(tea_pipe.loc['Pipeline load factor [%]']['Parameter'])
pipe_use
# Pipeline Opex [€/a as % of Capex]
pipe_opex_share = float(tea_pipe.loc['Pipeline Opex [€/a as % of Capex]']['Parameter'])
pipe_opex_share
# Compression Opex [€/a as % of Capex]
comp_opex_share = float(tea_pipe.loc['Compressor Opex [€/a as % of Capex]']['Parameter'])
comp_opex_share
# New pipeline

"""### New offshore pipeline"""

# Pipeline capex offshore new [€/kg/1000km]
capex_pipe_new_off_EHB = float(
    tea_pipe.loc['Medium - New Offshore (EHB 2022) Capex Pipeline [€/kg/1000km]']['Parameter'])
capex_pipe_new_off_EHB
# Compression capex offshore new [€/kg/1000km]
capex_comp_new_off_EHB = float(
    tea_pipe.loc['Medium - New Offshore (EHB 2022) Capex Compression [€/kg/1000km]']['Parameter'])
capex_comp_new_off_EHB
capex_new_off = capex_pipe_new_off_EHB + capex_comp_new_off_EHB
capex_new_off
opex_new_off = capex_pipe_new_off_EHB * pipe_opex_share + capex_comp_new_off_EHB * comp_opex_share
opex_new_off
# Onshore pipeline distance between two countries [km]
d_on = float(tra_d[(tra_d['Origin_Country'] == 'Norway') & (tra_d['Destination_Country'] == 'Germany')][
                 'onshore distance (km)'].values)
d_on
#not used here
off_factor = float(tea_pipe.loc['Offshore Capex cost factor ']['Parameter'])
off_factor
# Offshore pipeline distance between two countries [km]
d_sea = float(tra_d[(tra_d['Origin_Country'] == 'Norway') & (tra_d['Destination_Country'] == 'Germany')]['offshore distance (km)'].values)


"""### Electricity cost of compression"""

# Compression capacity in [MWel/1000km]
capa_comp = float(tea_pipe.loc['Medium - Compression capacity @ 48inch [MW_el/1000km]']['Parameter'])
capa_comp
# Load factor in hours/day
lf_comp = float(tea_pipe.loc['Assumed load factor for compressor electricity consumption [h/a]']['Parameter'])
lf_comp
#Capacity at full load factor [kg H2/a]
capa_pipe = float(tea_pipe.loc['Capacity at full load factor [kg H2/a]']['Parameter'])
capa_pipe
# Electricity cost for compression [€/kg/1000km]
#comp_el_cost = capa_comp * pipe_use * 8760 * p_el_y / capa_pipe


"""## Calc. offshore new pipeline cost"""

# Costs for new pipeline transport [€/kg]
#alpha, capex_pipe_new_on_EHB, capex_comp_new_on_EHB, d_on, d_off, capex_pipe_new_off_EHB, capex_comp_new_off_EHB, capa_comp, pipe_use, capa_pipe
def calculate_off_pipe_new():

    result = ((alpha_pipe * capex_new_off / pipe_use + opex_new_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea / 1000

    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['New_Pipeline_costs_off'])
result.index.name = 'Years'

for year in years:
    p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][year])

    result.New_Pipeline_costs_off.loc[year] = calculate_off_pipe_new()

result

output_file = os.path.join(path_csv,'New_Pipeline_costs_off.csv')
result.to_csv(output_file, sep = ';')
New_Pipeline_costs_off = result

"""## Retrofit offshore pipeline cost"""

#Retrofit  capex in [€/kg/100km]
capex_pipe_retrofit_off_EHB = float(tea_pipe.loc['Medium - retrofit Offshore (EHB 2022) Capex Pipeline [€/kg/1000km]']['Parameter'])
capex_pipe_retrofit_off_EHB

#Compression capex in [€/kg/100km]
capex_comp_retrofit_off_EHB = float(tea_pipe.loc['Medium - retrofit Offshore (EHB 2022) Capex Compression [€/kg/1000km]']['Parameter'])
capex_comp_retrofit_off_EHB

capex_retrofit_off = capex_pipe_retrofit_off_EHB + capex_comp_retrofit_off_EHB
capex_retrofit_off

opex_retrofit_off = capex_pipe_retrofit_off_EHB * pipe_opex_share + capex_comp_retrofit_off_EHB * comp_opex_share
opex_retrofit_off

"""### Calc. offshore retrofit pipeline costs"""

# Costs for new pipeline transport [€/kg]
#alpha, capex_pipe_new_on_EHB, capex_comp_new_on_EHB, d_on, d_off, capex_pipe_new_off_EHB, capex_comp_new_off_EHB, capa_comp, pipe_use, capa_pipe
def calculate_off_pipe_retrofit():
    result = (((alpha_pipe * capex_retrofit_off / pipe_use + opex_retrofit_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea) / 1000
    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Retrofit_pipeline_costs_off'])
result.index.name = 'Years'

for year in years:
    p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][year])

    result.Retrofit_pipeline_costs_off.loc[year] = calculate_off_pipe_retrofit()

result

Retrofit_pipeline_costs_off = result
output_file = os.path.join(path_csv,'Retrofit_pipeline_costs_off.csv')
Retrofit_pipeline_costs_off.to_csv(output_file, sep = ';')


"""## Sensitivity

### Electricity price sensi
"""

sensi_year = 2030
d_sea_sensi = 1000
p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][sensi_year])

P_el_sensi = np.arange(0,121,10)
sensitivity = []

def calculate_off_pipe_retrofit_sensi():
    for p_el_y in P_el_sensi:

        result = (((alpha_pipe * capex_retrofit_off / pipe_use + opex_retrofit_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea_sensi) / 1000

        sensitivity.append(result)

    return sensitivity

calculate_off_pipe_retrofit_sensi()

retro_pipe_sensi_P_el = pd.DataFrame(sensitivity, P_el_sensi, columns=['Retrofit pipeline transport costs [€/kg H2]'])
retro_pipe_sensi_P_el.index.name = 'Electricity Price [€/MWh]'
retro_pipe_sensi_P_el

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'retro_pipe_sensi_P_el.csv')
retro_pipe_sensi_P_el.to_csv(output_file, sep=',')

"""### Transport distance sensi Costs"""

transport_distance = np.arange(0,10001, 500)
sensitivity = []
def calculate_off_pipe_retrofit_sensi_distance():
    for d_sea in transport_distance:

        result = (((alpha_pipe * capex_retrofit_off / pipe_use + opex_retrofit_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea) / 1000

        sensitivity.append(result)

    return sensitivity

calculate_off_pipe_retrofit_sensi_distance()

retro_pipe_sensi_distance_LCOT = pd.DataFrame(sensitivity, transport_distance, columns=['Retrofit pipeline transport costs [€/kg H2]'])
retro_pipe_sensi_distance_LCOT.index.name = 'Transport distance in km'


# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'retro_pipe_sensi_distance_LCOT.csv')
retro_pipe_sensi_distance_LCOT.to_csv(output_file, sep=',')

"""### Sensitivity
#### Electricity price sensi
"""

sensi_year = 2030
#p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'b][year])

P_el_sensi = np.arange(0,121,10)
sensitivity = []

def calculate_off_pipe_new_sensi_P_el():
    for p_el_y in P_el_sensi:


        result = ((alpha_pipe * capex_new_off / pipe_use + opex_new_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea_sensi / 1000

        sensitivity.append(result)

    return sensitivity

calculate_off_pipe_new_sensi_P_el()

new_pipe_sensi_P_el = pd.DataFrame(sensitivity, P_el_sensi, columns=['New pipeline transport costs [€/kg H2]'])
new_pipe_sensi_P_el.index.name = 'Electricity price [€/MWh]'
new_pipe_sensi_P_el

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'new_pipe_sensi_P_el.csv')
new_pipe_sensi_P_el.to_csv(output_file, sep=',')

"""#### Transport distance sensi"""

sensi_year = 2030
p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][sensi_year])

transport_distance = np.arange(0,10001, 500)
sensitivity = []
def calculate_off_pipe_retrofit_sensi_distance():
    for d_sea in transport_distance:

        result = ((alpha_pipe * capex_new_off / pipe_use + opex_new_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea / 1000

        sensitivity.append(result)

    return sensitivity

calculate_off_pipe_retrofit_sensi_distance()

new_pipe_sensi_distance_LCOT = pd.DataFrame(sensitivity, transport_distance, columns=['New pipeline transport costs [€/kg H2]'])
new_pipe_sensi_distance_LCOT.index.name = 'Transport distance in km'
new_pipe_sensi_distance_LCOT

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'new_pipe_sensi_distance_LCOT.csv')
new_pipe_sensi_distance_LCOT.to_csv(output_file, sep=',')

"""## LH2 transport"""

tea_lh2 = pd.read_excel(path_excel, sheet_name='LH2', decimal=',', index_col=0)
tea_lh2

"""### Liquefaction"""

# WACC for liquefaction plant in %
i_liq = float(tea_lh2.loc['Liquefaction - Discount rate [%]']['LH2'])
i_liq

# Economic lifetime for the liquefaction plant [years]
l_liq = float(tea_lh2.loc['Liquefaction - Lifetime [Years]']['LH2'])
l_liq

# Amortisation factor alpha for the liquefaction plant
alpha_liq = (i_liq * (1 + i_liq) ** l_liq) /(((1 + i_liq) ** l_liq) - 1)
alpha_liq

opex_liq_share = float(tea_lh2.loc['Liquefaction - Opex opt. [% of Capex]']['LH2'])
opex_liq_share

"""Calc. liquefaction cost"""

# Definition of the cost calculation function for liquefaction costs. Time relevant variables = capex [€/tpa], opex [€/tpa], electricity use [kwh/kg H2], electricity price [$/MWh]
def calculate_liq_costs_LH2(alpha_liq, capex_liq_y, opex_liq_y, el_liq_y, p_el_y):

    result = float((alpha_liq * capex_liq_y/1000 + opex_liq_y/1000) + el_liq_y * p_el_y*0.89/1000)

    return result

# Calculation of liquefaction costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['LH2_Liquefaction_costs'])
result.index.name = 'Years'

for year in years:

    # get capex, opex, natural gas price and CO2 price of specific year
    capex_liq_y = float(tea_lh2.loc['Liquefaction - Capex opt. [€/t/a]'][year])
    opex_liq_y = capex_liq_y * opex_liq_share
    el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
    p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][year])

    # calculate costs of specific year
    result.LH2_Liquefaction_costs.loc[year] = calculate_liq_costs_LH2(alpha_liq=alpha_liq, capex_liq_y=capex_liq_y, opex_liq_y=opex_liq_y, el_liq_y=el_liq_y, p_el_y=p_el_y)

result

LH2_Liquefaction_costs = result

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'LH2_Liquefaction_costs.csv')
result.to_csv(output_file, sep = ';')

"""Plot liquefaction costs"""

# Plot cost curve for liquefaction
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'cyan', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Cost curve for liquefaction', fontweight = 'bold')
plt.legend(['Liquefaction costs in €/kg_H2'])
plt.xlabel('Year')
plt.ylabel('Cost [€/kg_H2]')
plt.show()

"""### Export terminal"""

# Lifetime of import and export terminals
l_ter = float(tea_lh2.loc['Export Terminal - Technical lifetime [Years]']['LH2'])

# Amortisation factor for the export terminal
alpha_et = (i_tra * (1 + i_tra) ** l_ter) / (((1 + i_tra) ** l_ter) - 1)

# Average storage time in the export terminal tanks in days
t_et = float(tea_lh2.loc['Export Terminal - Storage length per load [Days]']['LH2'])

# Electricity consumption for the export terminal [kWh/kg_h2]
el_et = float(tea_lh2.loc['Export Terminal - Electricity use [kWh/kgH2]']['LH2'])
el_et

# Boil-off hydrogen in [%/day]
bog_et = float(tea_lh2.loc['Export Terminal - Boil off rate [%/day]']['LH2'])
bog_et

# Average storage time in the export terminal tanks in days
t_et = float(tea_lh2.loc['Export Terminal - Storage length per load [Days]']['LH2'])
t_et

el_reliq = float(tea_lh2.loc['Export Terminal - Energy use to reliquefy opt. [kWh/kg H2/day]']['LH2'])
el_reliq

opex_et_share =  float(tea_lh2.loc['Export Terminal - Annual OPEX [% of CAPEX]']['LH2'])
opex_et_share

"""Calc. export terminal cost

Boil-off and transfer gas are re-liquefied @

Time relevant variables = capex [$/tpa], opex [$/tpa], electricity use [kwh/kg H2], electricity price [$/MWh] storage time in days
"""

# Definition of the cost calculation function for the Export terminal.
def calculate_export_terminal_costs_LH2(alpha_et, capex_et_y, opex_et_y, el_et, p_el_y, el_reliq):
    result = (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * t_et) * p_el_y* 0.89/1000

    return result

#Calculation of export terminal costs [$/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Export_terminal_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, xx
    capex_et_y = float(tea_lh2.loc['Export Terminal - CAPEX/tank [€/t/a]'][year])
    opex_et_y = capex_et_y * opex_et_share
    p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][year])


    # calculate costs of specific year
    result.LH2_Export_terminal_costs.loc[year] = calculate_export_terminal_costs_LH2(alpha_et=alpha_et, capex_et_y=capex_et_y,
                                                              opex_et_y=opex_et_y, el_et=el_et, p_el_y=p_el_y, el_reliq=el_reliq)

result

LH2_Export_terminal_costs =result

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'LH2_Export_terminal_costs.csv')
result.to_csv(output_file, sep=',')

#"Plot export terminal costs"

# Plot cost curve for export terminal costs
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'red', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='red', linestyle='dashed')
plt.title('Export terminal costs over time', fontweight='bold', fontsize='large')
plt.legend(['Export terminal costs in €/kg_H2'])
plt.xlabel('Year', fontsize='large')
plt.ylabel('Cost [€/kg_H2]', fontsize='large')
plt.show()

"""### Shipping"""

# Technical lifetime of ships in years
l_ship = float(tea_lh2.loc['Shipping - Technical Lifetime [Years]']['LH2'])
l_ship

# Amortisation factor for shipping
alpha_ship = (i_tra * (1 + i_tra) ** l_ship) / (((1 + i_tra) ** l_ship) - 1)
alpha_ship

# Ship speed in [km/h]
v_ship = float(tea_lh2.loc['Shipping - Ship speed [km/h]']['LH2'])
v_ship

# Berthing time (time for unloading and loading in a harbour) in [h]
h_ship = float(tea_lh2.loc['Shipping - Berthing time [hours]']['LH2'])
h_ship

# Boil-off when shipping in [%/day] to [%/hour]
bog_ship = float(tea_lh2.loc['Shipping - Boil off opt. [%/day]']['LH2'])/24
bog_ship

# Fuel consumption of a ship in [kg_h2/t/km]
f_ship = float(tea_lh2.loc['Shipping - Fuel use [kg H2/t/km]']['LH2'])/1000
f_ship

opex_ship_share = float(tea_lh2.loc['Shipping - Annual OPEX [% of CAPEX]']['LH2'])
opex_ship_share

"""Fuel cost calculation"""

# Cost of transported hydrogen carrier in [€/kg_h2] in year y
H2_costs = (LCOH_min_tech_em.loc[::]['Minimal_production_costs']).apply(pd.to_numeric)
Con_costs = (LH2_Liquefaction_costs.loc[::]['LH2_Liquefaction_costs']).apply(pd.to_numeric)
ET_costs =  (LH2_Export_terminal_costs.loc[::]['LH2_Export_terminal_costs']).apply(pd.to_numeric)

# €/kg_H2
LH2_cargo_cost = (H2_costs + Con_costs + ET_costs).to_frame('LH2_cargo_cost')
LH2_cargo_cost

"""Calc. shipping costs @ 10.000km"""

"""Definition of the cost calculation function for the maritime shipping.
 [%]
d_sea [km]
v_ship [km/h]
h_ship [h] berthing time
bog_ship [%/day]
f_ship [kg/kg/km]] Fuel use
Time relevant variables:
capex [€/kg/a]
opex [€/kg/a]
lcoh in [€/kg_h2]

Fuel assumptions: outward journey - boil-off is used as fuel.
Return journey: Residual H2 is used as fuel.
"""

def calculate_ship_costs_LH2():

    result = (alpha_ship * capex_ship_y + opex_ship_y) / (8760/(2*(d_sea/v_ship + h_ship)))\
             /(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea))\
             + (bog_ship * d_sea/v_ship + f_ship * d_sea) * lcoh

    return result

# Calculation of shipping costs [€/kg_h2] from 2025 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Shipping_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, xx
    capex_ship_y = float(tea_lh2.loc['Shipping - Capex/Ship opt. [€/t/a]'][year])/1000
    opex_ship_y = capex_ship_y * opex_ship_share
    lcoh = float(LH2_cargo_cost.loc[year]['LH2_cargo_cost'])

    # calculate costs of specific year
    result.LH2_Shipping_costs.loc[year] = calculate_ship_costs_LH2()

result


# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'LH2_Shipping_costs_baseline.csv')
result.to_csv(output_file, sep=',')

LH2_Shipping_costs = result

"""Plot shipping costs"""

# Plot cost curve of hydrogen production from NGR with CCS
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'green', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Shipping costs over time in €/kg H2', fontweight='bold')
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

"""### Import terminal"""

# Definition of variables for the import terminal costs

# Amortisation factor for the import terminal
alpha_it = alpha_et

# Electricity consumption in kWh/kg H2
el_it = float(tea_lh2.loc['Import Terminal - Electricity use [kWh/kg H2]']['LH2'])
el_it

# Import terminal boil-off in [%/day]
bog_it = float(tea_lh2.loc['Import Terminal - Boil-off [%/day]']['LH2'])
bog_it

# Average storage time in the export terminal tanks in days
t_it = float(tea_lh2.loc['Import Terminal - Storage length per load [days]']['LH2'])
t_it

opex_it_share =  float(tea_lh2.loc['Import Terminal - Annual OPEX [% of Opex]']['LH2'])
opex_it_share

# Import terminal costs in [$/kg_h2]:
# Capex for the import terminal in [$/tpa]
# Opex for the import terminal in [$/tpa]
# Electricity consumption for the import terminal in [kWh/kg_h2]
# Price for electricity in importing country [$/MWh]
# Import terminal boil-off in [%/h]
# Berthing time in import terminal in [h]
# Cost of transported hydrogen in [$/kg_h2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Import_terminal_costs'])
result.index.name = 'Years'

def calculate_import_terminal_costs_LH2():
    result = (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * t_it) * p_el_y* 0.89/1000

    return result

for year in years:
    # get capex, opex, xx
    capex_it_y = float(tea_lh2.loc['Import Terminal - CAPEX [€/t/a]'][year])
    opex_it_y = capex_it_y * opex_it_share
    p_el_y = float(prices.loc['Electricity prices in Germany [€_2020/MWh] Baseline'][year])


    # calculate costs of specific year
    result.LH2_Import_terminal_costs.loc[year] = calculate_import_terminal_costs_LH2()

result

LH2_Import_terminal_costs = result



# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'LH2_Import_terminal_costs.csv')
result.to_csv(output_file, sep=',')

"""Plot import terminal costs"""

fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'red', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='red', linestyle='dashed')
plt.title('Import terminal costs over time', fontweight='bold', fontsize='large')
plt.legend(['Import terminal costs in €/kg_H2'])
plt.xlabel('Year', fontsize='large')
plt.ylabel('Cost [€/kg_H2]', fontsize='large')
plt.show()

"""### Reconversion"""

# Economic lifetime for the reconversion plant [years]
l_recon = float(tea_lh2.loc['Reconversion - Lifetime [Years]']['LH2'])
l_recon

# Amortisation factor alpha for the reconversion plant
alpha_recon =  (i_liq * (1 + i_liq) ** l_recon) /(((1 + i_liq) ** l_recon) - 1)

# Electricty consumption for the conversion in [kWh/kg_h2]
el_recon = float(tea_lh2.loc['Import Terminal - Electricity use [kWh/kg H2]']['LH2'])
el_recon

opex_recon_share = float(tea_lh2.loc['Reconversion - Opex opt. [% of Capex]']['LH2'])
opex_recon_share

def calculate_recon_costs_LH2(alpha_recon, capex_recon_y, opex_recon_y, el_recon_y, p_el_y):
    result = float(
        (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + el_recon_y * p_el_y / 1000 * 0.89)
    return result

# Calculation of conversion costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Reconversion_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_recon_y = float(tea_lh2.loc['Reconversion - Capex opt. [€/t/a]'][year])
    opex_recon_y = capex_recon_y * opex_recon_share
    el_recon_y = float(tea_lh2.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
    p_el_y = float(prices.loc['Electricity prices in Germany [€_2020/MWh] Baseline'][year])

    # calculate costs of specific year
    result.LH2_Reconversion_costs.loc[year] = calculate_recon_costs_LH2(alpha_recon=alpha_recon, capex_recon_y=capex_recon_y,
                                                                opex_recon_y=opex_recon_y, el_recon_y=el_recon_y,
                                                                p_el_y=p_el_y)
result

LH2_Reconversion_costs = result

output_file = os.path.join(path_csv, 'LH2_Reconversion_costs.csv')
result.to_csv(output_file, sep=',')

"""## Total LH2 transport costs"""

liq_costs = LH2_Liquefaction_costs
et_costs = LH2_Export_terminal_costs
ship_costs = LH2_Shipping_costs
it_costs = LH2_Import_terminal_costs
recon_costs = LH2_Reconversion_costs

it_costs

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_transport_costs'])
result.index.name = 'Years'

# Calculate total transport costs from liquefaction costs (LC), export terminal costs (EC), Shipping costs (SC), import terminal costs (IC)
def calculate_seaborne_transport_costs_LH2():
    result = LC + EC + SC + IC + RC

    return result

for year in years:
    # get all costs
    LC = float(liq_costs.loc[year]['LH2_Liquefaction_costs'])
    EC = float(et_costs.loc[ year]['LH2_Export_terminal_costs'])
    SC = float(ship_costs.loc[year]['LH2_Shipping_costs'])
    IC = float(it_costs.loc[year]['LH2_Import_terminal_costs'])
    RC = float(recon_costs.loc[year]['LH2_Reconversion_costs'])
    # calculate costs of specific year
    result.LH2_transport_costs.loc[year] = calculate_seaborne_transport_costs_LH2()

result

LH2_transport_costs =result

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'LH2_transport_costs.csv')
result.to_csv(output_file, sep=',')

# Plot cost curve for LH2 transport
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'green', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Cost curve for LH2 shipping [€/kg H2]', fontweight='bold')
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

# Cost breakdown for LH2 shipping
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)
LC = (liq_costs.loc[::5]['LH2_Liquefaction_costs']).apply(pd.to_numeric)
EC = (et_costs.loc[::5]['LH2_Export_terminal_costs']).apply(pd.to_numeric)
SC = (ship_costs.loc[::5]['LH2_Shipping_costs']).apply(pd.to_numeric)
IC = (it_costs.loc[::5]['LH2_Import_terminal_costs']).apply(pd.to_numeric)
RC = (recon_costs.loc[::5]['LH2_Reconversion_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

RC_plt = plt.bar(x,RC, width, color= 'orange', label='Regasification costs', bottom= LC + EC + SC + IC)
IC_plt = plt.bar(x,IC, width, color= 'wheat',label='Import terminal costs', bottom= LC + EC + SC)
SC_plt = plt.bar(x,SC, width, color= 'crimson',label='Shipping costs', bottom= LC + EC)
EC_plt = plt.bar(x,EC, width, color= 'cornflowerblue',label='Export terminal costs', bottom= LC)
LC_plt = plt.bar(x,LC, width, color= 'darkblue',label='Liquefaction costs')

plt.title('Cost breakdown for LH2 shipping', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('Transport cost [€/kg H2]')
plt.xlabel('Years')

title= '\LH2_cost_breakdown'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()
#stackplot base pol Cost breakdown for LH2 shipping
#read policy inputs
liq_costs_pol = policy.LH2_Liquefaction_costs
et_costs_pol = policy.LH2_Export_terminal_costs
ship_costs_pol = policy.LH2_Shipping_costs
it_costs_pol = policy.LH2_Import_terminal_costs
recon_costs_pol = policy.LH2_Reconversion_costs
#plot
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
plt.subplot(1,2,1)
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
LC = (liq_costs.loc[::]['LH2_Liquefaction_costs']).apply(pd.to_numeric)
EC = (et_costs.loc[::]['LH2_Export_terminal_costs']).apply(pd.to_numeric)
SC = (ship_costs.loc[::]['LH2_Shipping_costs']).apply(pd.to_numeric)
IC = (it_costs.loc[::]['LH2_Import_terminal_costs']).apply(pd.to_numeric)
RC = (recon_costs.loc[::]['LH2_Reconversion_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Liquefaction', 'Export Terminal', 'Shipping', 'Import Terminal', 'Regasification']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, LC, EC, SC, IC, RC, labels = labels, colors = colors )

plt.ylim(0,2.3)
plt.xlim(2025,2050)
plt.legend(loc='upper right')
plt.ylabel('Transport costs [€/kg H2]')
plt.xlabel('Baseline')
#policy
plt.subplot(1,2,2)
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
LC = (liq_costs_pol.loc[::]['LH2_Liquefaction_costs']).apply(pd.to_numeric)
EC = (et_costs_pol.loc[::]['LH2_Export_terminal_costs']).apply(pd.to_numeric)
SC = (ship_costs_pol.loc[::]['LH2_Shipping_costs']).apply(pd.to_numeric)
IC = (it_costs_pol.loc[::]['LH2_Import_terminal_costs']).apply(pd.to_numeric)
RC = (recon_costs_pol.loc[::]['LH2_Reconversion_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Liquefaction', 'Export Terminal', 'Shipping', 'Import Terminal', 'Regasification']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, LC, EC, SC, IC, RC, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.ylim(0,2.3)
plt.legend(loc='upper right')
plt.ylabel('Transport costs [€/kg H2]')
plt.xlabel('Policy')
title = '\LH2_cost_breakdown_base_pol'
plt.savefig(path_plt + title + '.png', transparent = True)
plt.show()
"""### Sensitivity"""

year = 2030
el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
capex_liq_y = float(tea_lh2.loc['Liquefaction - Capex opt. [€/t/a]'][year])
opex_liq_y = capex_liq_y * opex_liq_share
el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
capex_et_y = float(tea_lh2.loc['Export Terminal - CAPEX/tank [€/t/a]'][year])
opex_et_y = capex_et_y * opex_et_share
capex_it_y = float(tea_lh2.loc['Import Terminal - CAPEX [€/t/a]'][year])
opex_it_y = capex_it_y * opex_it_share
capex_ship_y = float(tea_lh2.loc['Shipping - Capex/Ship opt. [€/t/a]'][year])/1000
opex_ship_y = capex_ship_y * opex_ship_share

capex_recon_y = float(tea_lh2.loc['Reconversion - Capex opt. [€/t/a]'][year])
opex_recon_y = capex_recon_y * opex_recon_share
el_recon_y = float(tea_lh2.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])

p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][year])

"""#### Electricity price sensi"""
lcoh = 3
LH2_cargo_costs_sensi = (lcoh+((alpha_liq * capex_liq_y/1000 + opex_liq_y/1000) + el_liq_y * p_el_y/1000) + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * t_et) * p_el_y/1000)

P_el_sensi = np.arange(0,121,10)
sensitivity = []

def lh2_transport_sensi_P_el(p_el_y, alpha_liq, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_liq_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea_seni ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y):

    for p_el_y in P_el_sensi:

        result = \
        ((alpha_liq * capex_liq_y/1000 + opex_liq_y/1000) + el_liq_y * p_el_y/1000)\
        + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * t_et) * p_el_y/1000 \
        + (alpha_ship * capex_ship_y + opex_ship_y) / (8760/(2*(d_sea_seni/v_ship + h_ship)))/(1-(bog_ship * d_sea_seni/v_ship) - (f_ship  * d_sea_seni)) + (bog_ship * d_sea_seni/v_ship + f_ship * d_sea_seni) * LH2_cargo_costs_sensi\
        + (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * t_it) * p_el_y/1000 \
        + (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + el_recon_y * p_el_y / 1000                 # Reconversion

        sensitivity.append(result)

    return sensitivity

lh2_transport_sensi_P_el(p_el_y, alpha_liq, alpha_et, alpha_ship, alpha_it, alpha_recon, capex_liq_y, capex_et_y, capex_ship_y, capex_it_y, bog_ship, capex_recon_y, d_sea_sensi, v_ship, f_ship, el_et, el_it, t_it, el_recon_y)

lh2_transport_sensi_P_el= pd.DataFrame(sensitivity, P_el_sensi, columns=['LH2 transport costs [€/kg H2]'])
lh2_transport_sensi_P_el.index.name = 'Electricity price [€/MWh]'
lh2_transport_sensi_P_el

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lh2_transport_sensi_P_el.csv')
lh2_transport_sensi_P_el.to_csv(output_file, sep = ';')

"""#### Transport distance sensi"""

transport_distance = np.arange(0,10001, 500)
sensitivity = []

def lh2_transport_distance(p_el_y, alpha_liq, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_liq_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y):

    for d_sea in transport_distance:

        result = \
        ((alpha_liq * capex_liq_y/1000 + opex_liq_y/1000) + el_liq_y * p_el_y/1000)\
        + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * t_et) * p_el_y/1000\
        + (alpha_ship * capex_ship_y + opex_ship_y) / (8760/(2*(d_sea/v_ship + h_ship)))/(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea)) + (bog_ship * d_sea/v_ship + f_ship * d_sea) * LH2_cargo_costs_sensi\
        + (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * t_it) * p_el_y/1000\
        + (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + el_recon_y * p_el_y / 1000\

        sensitivity.append(result)

    return sensitivity


lh2_transport_distance(p_el_y, alpha_liq, alpha_et, alpha_ship, alpha_it, alpha_recon, capex_liq_y, capex_et_y,
                       capex_ship_y, capex_it_y, bog_ship, capex_recon_y, d_sea, v_ship, f_ship, el_et, el_it, t_it,
                       el_recon_y)

lh2_transport_sensi_distance_LCOT= pd.DataFrame(sensitivity, transport_distance, columns=['LH2 transport costs [€/kg H2]'])
lh2_transport_sensi_distance_LCOT.index.name = 'Transport distance in km'
lh2_transport_sensi_distance_LCOT

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lh2_transport_sensi_distance_LCOT.csv')
lh2_transport_sensi_distance_LCOT.to_csv(output_file, sep = ';')

"""## NH3 Transport"""

tea_NH3 = pd.read_excel(path_excel, sheet_name='NH3', decimal=',', index_col=0)
tea_NH3

"""### Conversion"""

# Economic lifetime for the conversion plant [years]
l_con = float(tea_NH3.loc['Conversion - Lifetime [Years]']['NH3'])
l_con

# Amortisation factor alpha for the conversion plant
alpha_con = (i * (1 + i) ** l_con) /(((1 + i) ** l_con) - 1)
alpha_con

'not used'# Conversion efficiency [% of LHV]
#eff_con = float(tea_NH3.loc['Conversion - Efficiency opt. [% of LHV]'][2025])

opex_con_share = (tea_NH3.loc['Conversion - Opex opt. [% of Capex]']['NH3'])
opex_con_share

"""Calc. conversion cost

Definition of the cost calculation function for conversion costs. Time relevant variables =
capex [€/tpa]
opex [€/tpa]
electricity use [kwh/kg H2]
electricity price [$/MWh]
"""

def calculate_con_costs_NH3():

    result = float((alpha_con * capex_con_y/1000 + opex_con_y/1000) + el_con_y * p_el_y/1000)

    return result

# Calculation of conversion costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['NH3_Conversion_costs'])
result.index.name = 'Years'

for year in years:

    # get capex, opex, natural gas price and CO2 price of specific year
    capex_con_y = float(tea_NH3.loc['Conversion - Capex opt. [€/t/a]'][year])
    opex_con_y = capex_con_y * opex_con_share
    el_con_y = float(tea_NH3.loc['Conversion - Electricity consumption opt. [kWh/kgH2]'][year])
    p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][year])

    # calculate costs of specific year
    result.NH3_Conversion_costs.loc[year] = calculate_con_costs_NH3()

result

NH3_Conversion_costs = result


# Create csv file from results dataframe
output_file = os.path.join(path_csv,'NH3_Conversion_costs.csv')
result.to_csv(output_file, sep = ';')

"""Plot conversion costs"""

# Plot cost curve for conversion of H2 to NH3
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'cyan', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Cost curve for conversion of H2 to NH3 [€/kg]', fontweight = 'bold')
plt.legend(['Conversion costs'])
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

"""### Export terminal"""

# Lifetime of import and export terminals
l_ter = float(tea_NH3.loc['Export Terminal - Technical lifetime [Years]']['NH3'])
# Amortisation factor for the export terminal
alpha_et = (i * (1 + i) ** l_ter) / (((1 + i) ** l_ter) - 1)

# Electricity consumption for the export terminal [kWh/kg_h2]
el_et = float(tea_NH3.loc['Export Terminal - Electricity use [kWh/kgH2]']['NH3'])
el_et

# Boil-off hydrogen in [%/day]
bog_et = float(tea_NH3.loc['Export Terminal - Boil off rate [%/day]']['NH3'])
bog_et

# Average storage time in the export terminal tanks in days
t_et = float(tea_NH3.loc['Export Terminal - Storage length per load [Days]']['NH3'])
t_et

#Lower heating value of Nh3 [kWh/kg]
NH3_lhv =  float(tea_NH3.loc['Properties - LHV [kWh/kg]']['NH3'])
NH3_lhv

#Lower heating value of H2 [kWh/kg]
H2_lhv = 33.33

# transformed to kWh/kg h2/day
el_reliq = float(tea_NH3.loc['Export Terminal - Energy use to reliquefy opt. [kWh/kg NH3/day]']['NH3'])
el_reliq

opex_et_share

"""Calc. export terminal cost

Time relevant variables =
capex [€/tpa]
opex [€/tpa]
el_et electricity use [kwh/kg H2]
electricity price [$/MWh]
constant =
boil off in [%/d]
storage time in days
Energy use to reliquefy opt. [kWh/kg NH3/day]
"""

# Definition of the cost calculation function for the Export terminal.

def calculate_export_terminal_costs_NH3():
    result = (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * (NH3_lhv/H2_lhv) * t_et) * p_el_y/1000

    return result

#Calculation of export terminal costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Export_terminal_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, xx
    capex_et_y = float(tea_NH3.loc['Export Terminal - CAPEX for storage tanks [€/t/a]'][year])
    opex_et_y = float(tea_NH3.loc['Export Terminal - Annual OPEX [€/t/a]'][year])
    p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][year])

    # calculate costs of specific year
    result.NH3_Export_terminal_costs.loc[year] = calculate_export_terminal_costs_NH3()

result

NH3_Export_terminal_costs = result


# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'NH3_Export_terminal_costs.csv')
result.to_csv(output_file, sep=',')

"""Plot export terminal costs"""

# Plot cost curve for export terminal costs
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'red', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='red', linestyle='dashed')
plt.title('Export terminal costs over time', fontweight='bold', fontsize='large')
plt.legend(['Export terminal costs in €/kg_H2'])
plt.xlabel('Year', fontsize='large')
plt.ylabel('Cost [€/kg_H2]', fontsize='large')
plt.show()

"""### Shipping"""

# Technical lifetime of ships in years
l_ship = float(tea_NH3.loc['Shipping - Technical Lifetime [Years]']['NH3'])
l_ship

# Amortisation factor for shipping
alpha_ship = (i * (1 + i) ** l_ship) / (((1 + i) ** l_ship) - 1)
alpha_ship

# Ship speed in [km/h]
v_ship = float(tea_NH3.loc['Shipping - Ship speed [km/h]']['NH3'])
v_ship

# Berthing time (time for unloading and loading in a harbour) in [h]
h_ship = float(tea_NH3.loc['Shipping - Berthing time [hours]']['NH3'])
h_ship

#Lower heating value of Nh3[ kWh / kg]
NH3_lhv = float(tea_NH3.loc['Properties - LHV [kWh/kg]']['NH3'])
NH3_lhv
#Lower heating value of H2 [kWh/kg]
H2_lhv = 33.33

# Boil-off when shipping in [%_NH3/day] to [%/hour]
bog_ship_nh3 = float(tea_NH3.loc['Shipping - Boil off opt. [%/day]']['NH3'])/24 * (NH3_lhv/H2_lhv)


# Fuel consumption of a ship in [kg_NH3/t_NH3/km]
#f_ship = float(tea_NH3.loc['Shipping - Fuel use [kg NH3/t/km]']['NH3'])/1000 * (NH3_lhv/H2_lhv)

opex_ship_share = (tea_NH3.loc['Shipping - Annual OPEX [% of CAPEX]']['NH3'])
opex_ship_share

# Cost of transported hydrogen in [$/kg_h2] in year y

# Cost of transported hydrogen carrier in [€/kg_h2] in year y
H2_costs = (LCOH_min_tech_em.loc[::]['Minimal_production_costs']).apply(pd.to_numeric)
Con_costs = (NH3_Conversion_costs.loc[::]['NH3_Conversion_costs']).apply(pd.to_numeric)
ET_costs = (NH3_Export_terminal_costs.loc[::]['NH3_Export_terminal_costs']).apply(pd.to_numeric)

# €/kg_H2 transported in the form of NH3
NH3_cargo_cost = (H2_costs + Con_costs + ET_costs).to_frame('NH3_cargo_cost')
NH3_cargo_cost

"""Definition of the cost calculation function for the maritime shipping.
Time relevant variables:
capex [€/t/a]
opex [€/t/a]
lcoh in [€/kg_h2]
static:
alpha [%]
d_sea [km]
v_ship [km/h]
h_ship [h] berthing time
bog_ship_nh3 [%_H2/day]
f_ship_nh3 [kg_H2/kg_NH3/km]] Fuel use

"""

def calculate_ship_costs_NH3(alpha_ship, capex_ship_y, opex_ship_y, d_sea, v_ship, h_ship, bog_ship_nh3, f_ship_nh3, NH3_costs_y):
    result = (alpha_ship * capex_ship_y + opex_ship_y)/(8760/(2*(d_sea/v_ship + h_ship)))\
             /(1-(bog_ship_nh3 * d_sea/v_ship) - (f_ship_nh3  * d_sea))\
             + (bog_ship_nh3 * d_sea/v_ship + f_ship_nh3 * d_sea) * NH3_costs_y

    return result

# Calculation of shipping costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Shipping_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, xx
    capex_ship_y = float(tea_NH3.loc['Shipping - Capex/Ship opt. [€/t/a]'][year])/1000
    opex_ship_y = capex_ship_y * opex_ship_share
    f_ship_nh3 = float(tea_NH3.loc['Shipping - Fuel use [kg NH3/t/km]'][year])/1000 * (NH3_lhv/H2_lhv)
    NH3_costs_y = float(NH3_cargo_cost.loc[year])

    # calculate costs of specific year
    result.NH3_Shipping_costs.loc[year] = calculate_ship_costs_NH3(alpha_ship=alpha_ship, capex_ship_y=capex_ship_y, opex_ship_y=opex_ship_y, d_sea=d_sea, v_ship=v_ship, h_ship=h_ship, bog_ship_nh3=bog_ship_nh3, f_ship_nh3=f_ship_nh3, NH3_costs_y=NH3_costs_y)

result

NH3_Shipping_costs = result


output_file = os.path.join(path_csv, 'NH3_Shipping_costs.csv')
result.to_csv(output_file, sep=',')

"""Plot shipping costs"""

# Plot cost curve of hydrogen production from NGR with CCS
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'green', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Shipping costs over time in €/kg H2', fontweight='bold')
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

"""### Import terminal"""

# Definition of variables for the import terminal costs

# Amortisation factor for the import terminal
alpha_it = alpha_et

# Electricity consumption in kWh/kg H2
el_it = float(tea_NH3.loc['Import Terminal - Electricity use [kWh/kg H2]']['NH3'])
el_it

# Import terminal boil-off in [%/day]
bog_it = float(tea_NH3.loc['Import Terminal - Boil-off [%/day]']['NH3'])
bog_it

# Average storage time in the export terminal tanks in days
t_it = float(tea_NH3.loc['Import Terminal - Storage length per load [days]']['NH3'])
t_it

opex_it_share = opex_et_share

"""Time relevant variables =
capex [€/tpa]
opex [€/tpa]
electricity use [kwh/kg H2]
electricity price [$/MWh] in Germany
lcoh in [€/kg_h2]
constant =
boil off in [%/d]
storage time in [days]


"""

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Import_terminal_costs'])
result.index.name = 'Years'

def calculate_import_terminal_costs_NH3(alpha_it, capex_it_y, opex_it_y, el_it, p_el_y, t_it):
    result = (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * p_el_y /1000
    return result

for year in years:
    # get capex, opex, xx
    capex_it_y = float(tea_NH3.loc['Import Terminal - CAPEX for storage tanks [€/t/a]'][year])
    opex_it_y = float(tea_NH3.loc['Import Terminal - Annual OPEX [€/t/a]'][year])
    p_el_y = float(prices.loc['Electricity prices in Germany [€_2020/MWh] Baseline'][year])


    # calculate costs of specific year
    result.NH3_Import_terminal_costs.loc[year] = calculate_import_terminal_costs_NH3(alpha_it=alpha_it, capex_it_y=capex_it_y,
                                                              opex_it_y=opex_it_y, el_it=el_it, p_el_y=p_el_y, t_it=t_it)

result

NH3_Import_terminal_costs = result

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'NH3_Import_terminal_costs.csv')
result.to_csv(output_file, sep=',')

"""
Plot import terminal costs"""

fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.plot(result, color = 'red', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='red', linestyle='dashed')
plt.title('Import terminal costs over time', fontweight='bold', fontsize='large')
plt.legend(['Import terminal costs in €/kg_H2'])
plt.xlabel('Year', fontsize='large')
plt.ylabel('Cost [€/kg_H2]', fontsize='large')
plt.show()

"""


### Reconversion"""

# Economic lifetime for the reconversion plant [years]
l_recon = float(tea_NH3.loc['Reconversion - Lifetime [Years]']['NH3'])
l_recon

# Amortisation factor alpha for the conversion plant
alpha_recon = (i * (1 + i) ** l_recon) / (((1 + i) ** l_recon) - 1)
alpha_recon

opex_recon_share = float(tea_NH3.loc['Reconversion - Annual OPEX opt. [% of CAPEX]']['NH3'])

"""Calc. reconversion cost
Definition of the cost calculation function for conversion costs.Time relevant variables =
capex[€/tpa]
opex[€/tpa]
electricity use[kwh/kg H2]
electricity price[$/ MWh]

Energy demand for heat in cracking is provided locally
"""

def calculate_recon_costs_NH3(alpha_recon, capex_recon_y, opex_recon_y, el_recon_y, p_el_y):
    result = float((alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + (heat_recon_y + el_recon_y) * p_el_y / 1000)
    return result

# Calculation of conversion costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Reconversion_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_recon_y = float(tea_NH3.loc['Reconversion - Capex opt. [€/t/a]'][year])
    opex_recon_y = capex_recon_y * opex_recon_share
    el_recon_y = float(tea_NH3.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
    heat_recon_y = float(tea_NH3.loc['Reconversion - Heat consumption opt. [kWh/kg H2]'][year])
    p_el_y = float(prices.loc['Electricity prices in Germany [€_2020/MWh] Baseline'][year])

    # calculate costs of specific year
    result.NH3_Reconversion_costs.loc[year] = calculate_recon_costs_NH3(alpha_recon=alpha_recon, capex_recon_y=capex_recon_y,
                                                            opex_recon_y=opex_recon_y, el_recon_y=el_recon_y, p_el_y=p_el_y)
result

NH3_Reconversion_costs = result

"""## Total NH3 transport costs"""

con_costs = NH3_Conversion_costs
et_costs = NH3_Export_terminal_costs
ship_costs = NH3_Shipping_costs
it_costs = NH3_Import_terminal_costs
recon_costs = NH3_Reconversion_costs

it_costs

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_transport_costs'])
result.index.name = 'Years'

# Calculate total transport costs from liquefaction costs (LC), export terminal costs (EC), Shipping costs (SC), import terminal costs (IC)
def calculate_seaborne_transport_costs_NH3():
    result = CC + EC + SC + IC + RC

    return result

for year in years:
    # get all costs
    CC = float(con_costs.loc[year]['NH3_Conversion_costs'])
    EC = float(et_costs.loc[year]['NH3_Export_terminal_costs'])
    SC = float(ship_costs.loc[year]['NH3_Shipping_costs'])
    IC = float(it_costs.loc[year]['NH3_Import_terminal_costs'])
    RC = float(recon_costs.loc[year]['NH3_Reconversion_costs'])
    # calculate costs of specific year
    result.NH3_transport_costs.loc[year] = calculate_seaborne_transport_costs_NH3()

result

NH3_transport_costs = result

# Nh3 transport costs w/o cracking
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_transport_costs_wo_cracking'])
result.index.name = 'Years'

for year in years:
    # get all costs
    CC = float(con_costs.loc[year]['NH3_Conversion_costs'])
    EC = float(et_costs.loc[year]['NH3_Export_terminal_costs'])
    SC = float(ship_costs.loc[year]['NH3_Shipping_costs'])
    IC = float(it_costs.loc[year]['NH3_Import_terminal_costs'])
    #RC = float(recon_costs.loc[year]['NH3_Reconversion_costs'])
    # calculate costs of specific year
    result.NH3_transport_costs_wo_cracking.loc[year] = calculate_seaborne_transport_costs_NH3()

result

NH3_transport_costs_wo_recon =result

    # Create csv file from results dataframe
output_file = os.path.join(path_csv, 'NH3_transport_costs_wo_recon.csv')
NH3_transport_costs_wo_recon.to_csv(output_file, sep=',')

# Cost breakdown for NH3 shipping w/ recon
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)
CC = (con_costs.loc[::5]['NH3_Conversion_costs']).apply(pd.to_numeric)
EC = (et_costs.loc[::5]['NH3_Export_terminal_costs']).apply(pd.to_numeric)
SC = (ship_costs.loc[::5]['NH3_Shipping_costs']).apply(pd.to_numeric)
IC = (it_costs.loc[::5]['NH3_Import_terminal_costs']).apply(pd.to_numeric)
RC = (recon_costs.loc[::5]['NH3_Reconversion_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

RC_plt = plt.bar(x,RC, width, color = 'orange', label='Reconversion costs', bottom= CC + EC + SC + IC)
IC_plt = plt.bar(x,IC, width, color = 'wheat',label='Import terminal costs', bottom= CC + EC + SC)
SC_plt = plt.bar(x,SC, width, color = 'crimson',label='Shipping costs', bottom= CC + EC)
EC_plt = plt.bar(x,EC, width, color = 'cornflowerblue',label='Export terminal costs', bottom= CC)
CC_plt = plt.bar(x,CC, width, color = 'darkblue',label='Conversion costs')

plt.title('Cost breakdown for NH3 shipping', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('Transport cost [€/kg H2]')
plt.xlabel('Years')


title= '\Ammonia_cost_breakdown_w_recon'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()
#nh3 cost breakdown base pol
#read pol
con_costs_pol = policy.NH3_Conversion_costs
et_costs_pol = policy.NH3_Export_terminal_costs
ship_costs_pol = policy.NH3_Shipping_costs
it_costs_pol = policy.NH3_Import_terminal_costs
recon_costs_pol = policy.NH3_Reconversion_costs
##plot
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
plt.subplot(1,2,1)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
CC = (con_costs.loc[::]['NH3_Conversion_costs']).apply(pd.to_numeric)
EC = (et_costs.loc[::]['NH3_Export_terminal_costs']).apply(pd.to_numeric)
SC = (ship_costs.loc[::]['NH3_Shipping_costs']).apply(pd.to_numeric)
IC = (it_costs.loc[::]['NH3_Import_terminal_costs']).apply(pd.to_numeric)
RC = (recon_costs.loc[::]['NH3_Reconversion_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Conversion', 'Export Terminal', 'Shipping', 'Import Terminal', 'Reconversion']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, CC, EC, SC, IC, RC, labels = labels, colors = colors )

plt.ylim(0,3)
plt.xlim(2025,2050)
plt.legend(loc='upper right')
plt.ylabel('Transport costs [€/kg H2]')
plt.xlabel('Baseline')

# policy
plt.subplot(1,2,2)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
CC = (con_costs_pol.loc[::]['NH3_Conversion_costs']).apply(pd.to_numeric)
EC = (et_costs_pol.loc[::]['NH3_Export_terminal_costs']).apply(pd.to_numeric)
SC = (ship_costs_pol.loc[::]['NH3_Shipping_costs']).apply(pd.to_numeric)
IC = (it_costs_pol.loc[::]['NH3_Import_terminal_costs']).apply(pd.to_numeric)
RC = (recon_costs_pol.loc[::]['NH3_Reconversion_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Conversion', 'Export Terminal', 'Shipping', 'Import Terminal', 'Reconversion']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, CC, EC, SC, IC, RC, labels = labels, colors = colors )

plt.ylim(0,3)
plt.xlim(2025,2050)
plt.legend(loc='upper right')
plt.ylabel('Transport costs [€/kg H2]')
plt.xlabel('Policy')
title = '\Ammonia_cost_breakdown_stackplot_base_pol'
plt.savefig(path_plt + title + '.png', transparent = True)
plt.show()

# Cost breakdown for NH3 shipping w/0 recon
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)
CC = (con_costs.loc[::5]['NH3_Conversion_costs']).apply(pd.to_numeric)
EC = (et_costs.loc[::5]['NH3_Export_terminal_costs']).apply(pd.to_numeric)
SC = (ship_costs.loc[::5]['NH3_Shipping_costs']).apply(pd.to_numeric)
IC = (it_costs.loc[::5]['NH3_Import_terminal_costs']).apply(pd.to_numeric)
#RC = (recon_costs.loc[::5]['NH3_Reconversion_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

#RC_plt = plt.bar(x,RC, width, label='Reconversion costs', bottom= CC + EC + SC + IC)
IC_plt = plt.bar(x,IC, width, color = 'wheat', label='Import terminal costs', bottom= CC + EC + SC)
SC_plt = plt.bar(x,SC, width,color = 'crimson', label='Shipping costs', bottom= CC + EC)
EC_plt = plt.bar(x,EC, width, color = 'cornflowerblue', label='Export terminal costs', bottom= CC)
CC_plt = plt.bar(x,CC, width, color = 'darkblue', label='Conversion costs')

plt.title('Cost breakdown for NH3 shipping w/o cracking', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('Transport cost [€/kg H2]')
plt.xlabel('Years')

title= '\Ammonia_cost_breakdown_wo_recon'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

#nh3 cost breakdown w/o recon
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
CC = (con_costs.loc[::]['NH3_Conversion_costs']).apply(pd.to_numeric)
EC = (et_costs.loc[::]['NH3_Export_terminal_costs']).apply(pd.to_numeric)
SC = (ship_costs.loc[::]['NH3_Shipping_costs']).apply(pd.to_numeric)
IC = (it_costs.loc[::]['NH3_Import_terminal_costs']).apply(pd.to_numeric)
#RC = (recon_costs.loc[::]['NH3_Reconversion_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Conversion', 'Export Terminal', 'Shipping', 'Import Terminal']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat']
plt.stackplot(x, CC, EC, SC, IC, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.legend(loc='upper right')
plt.ylabel('[g CO2eq/kg H2]')

title = '\Ammonia_cost_breakdown_wo_recon_stackplot'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()

#nh3 cost breakdwon for baseline and policy

"""### Sensitivity"""

year = 2030

capex_con_y = float(tea_NH3.loc['Conversion - Capex opt. [€/t/a]'][year])
opex_con_y = capex_con_y * opex_con_share
el_con_y = float(tea_NH3.loc['Conversion - Electricity consumption opt. [kWh/kgH2]'][year])

capex_et_y = float(tea_NH3.loc['Export Terminal - CAPEX for storage tanks [€/t/a]'][year])
opex_et_y = capex_et_y * opex_et_share

capex_it_y = float(tea_NH3.loc['Import Terminal - CAPEX for storage tanks [€/t/a]'][year])
opex_it_y = capex_it_y * opex_it_share

capex_ship_y = float(tea_NH3.loc['Shipping - Capex/Ship opt. [€/t/a]'][year])/1000
opex_ship_y = capex_ship_y * opex_ship_share
f_ship_nh3 = float(tea_NH3.loc['Shipping - Fuel use [kg NH3/t/km]'][year])/1000 * (NH3_lhv/H2_lhv)


capex_recon_y = float(tea_NH3.loc['Reconversion - Capex opt. [€/t/a]'][year])
opex_recon_y = capex_recon_y * opex_recon_share
el_recon_y = float(tea_NH3.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
heat_recon_y = float(tea_NH3.loc['Reconversion - Heat consumption opt. [kWh/kg H2]'][year])

p_el_y = float(prices.loc['Electricity prices in Norway [€_2021/MWh] Baseline'][year])

"""#### Electricity price sensi"""
NH3_cargo_cost_sensi = (lcoh + (alpha_con * capex_con_y/1000 + opex_con_y/1000) + el_con_y * p_el_y / 1000+ (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * (NH3_lhv/H2_lhv) * t_et) * p_el_y / 1000)
P_el_sensi = np.arange(0,121,10)
sensitivity = []

def nh3_transport_sensi_P_el(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship_nh3, capex_recon_y, d_sea ,v_ship, f_ship_nh3, el_et ,el_it ,t_it ,el_recon_y):

    for p_el_y in P_el_sensi:

        result = \
            (alpha_con * capex_con_y/1000 + opex_con_y/1000) + el_con_y * p_el_y / 1000 \
            + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * (NH3_lhv/H2_lhv) * t_et) * p_el_y / 1000 \
            + (alpha_ship * capex_ship_y + opex_ship_y) / (8760 / (2 * (d_sea_sensi / v_ship + h_ship))) / (1 - (bog_ship_nh3 * d_sea_sensi / v_ship) - (f_ship_nh3 * d_sea_sensi)) + (bog_ship_nh3 * d_sea_sensi / v_ship + f_ship_nh3 * d_sea_sensi) * NH3_cargo_cost_sensi \
            + (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * p_el_y / 1000 \
            + (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + (heat_recon_y + el_recon_y) * p_el_y / 1000

        sensitivity.append(result)

    return sensitivity

nh3_transport_sensi_P_el(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship_nh3, capex_recon_y, d_sea ,v_ship, f_ship_nh3, el_et ,el_it ,t_it ,el_recon_y)

NH3_transport_sensi_P_el= pd.DataFrame(sensitivity, P_el_sensi, columns=['NH3 transport costs [€/kg H2]'])
NH3_transport_sensi_P_el.index.name = 'Electricity price [€/MWh]'
NH3_transport_sensi_P_el

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'NH3_transport_sensi_P_el.csv')
NH3_transport_sensi_P_el.to_csv(output_file, sep = ';')

# Pel sensi w/o recon
sensitivity = []

def nh3_transport_sensi_P_el_wo_recon(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship_nh3, capex_recon_y, d_sea ,v_ship, f_ship_nh3, el_et ,el_it ,t_it ,el_recon_y):

    for p_el_y in P_el_sensi:

        result = \
            (alpha_con * capex_con_y/1000 + opex_con_y/1000) + el_con_y * p_el_y / 1000 \
            + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * (NH3_lhv/H2_lhv) * t_et) * p_el_y / 1000 \
            + (alpha_ship * capex_ship_y + opex_ship_y) / (8760 / (2 * (d_sea_sensi / v_ship + h_ship))) / (1 - (bog_ship_nh3 * d_sea_sensi / v_ship) - (f_ship_nh3 * d_sea_sensi)) + (bog_ship_nh3 * d_sea_sensi / v_ship + f_ship_nh3 * d_sea_sensi) * NH3_cargo_cost_sensi \
            + (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * p_el_y / 1000 \
        #+ (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + (heat_recon_y + el_recon_y) * p_el_y / 1000

        sensitivity.append(result)

    return sensitivity

nh3_transport_sensi_P_el_wo_recon(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship_nh3, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y)

nh3_transport_sensi_P_el_wo_recon= pd.DataFrame(sensitivity, P_el_sensi, columns=['NH3 transport costs [€/kg H2]'])
nh3_transport_sensi_P_el_wo_recon.index.name = 'Electricity price [€/MWh]'
nh3_transport_sensi_P_el_wo_recon

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'nh3_transport_sensi_P_el_wo_recon.csv')
nh3_transport_sensi_P_el_wo_recon.to_csv(output_file, sep = ';')

"""#### Transport distance sensi"""

transport_distance = np.arange(0,10001, 500)
sensitivity = []
def nh3_transport_sensi_distance(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship_nh3, capex_recon_y, d_sea ,v_ship, f_ship_nh3, el_et ,el_it ,t_it ,el_recon_y):

    for d_sea in transport_distance:

        result = \
        (alpha_con * capex_con_y/1000 + opex_con_y/1000) + el_con_y * p_el_y/1000 \
        + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * (NH3_lhv/H2_lhv) * t_et) * p_el_y /1000 \
        + (alpha_ship * capex_ship_y + opex_ship_y)/(8760/(2*(d_sea/v_ship + h_ship)))/(1-(bog_ship_nh3 * d_sea/v_ship) - (f_ship_nh3  * d_sea))+ (bog_ship_nh3 * d_sea/v_ship + f_ship_nh3 * d_sea) * NH3_cargo_cost_sensi\
        + (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * p_el_y /1000\
        + (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + (heat_recon_y + el_recon_y) * p_el_y / 1000

        sensitivity.append(result)

    return sensitivity

nh3_transport_sensi_distance(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship_nh3, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y)

NH3_transport_sensi_distance_LCOT= pd.DataFrame(sensitivity, transport_distance, columns=['NH3 transport costs [€/kg H2]'])
NH3_transport_sensi_distance_LCOT.index.name = 'Transport distance in km'
NH3_transport_sensi_distance_LCOT

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'NH3_transport_sensi_distance_LCOT.csv')
NH3_transport_sensi_distance_LCOT.to_csv(output_file, sep =';')

"""## Sensitivity Plots

### Electricity price sensi
"""

lh2_transport_sensi_P_el

fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#plt.subplot(1,2,1)

plt.plot(lh2_transport_sensi_P_el, color='blue', linestyle='solid', label = 'LH2')
plt.plot(NH3_transport_sensi_P_el, color='darkorange', linestyle='solid', label = 'NH3')
plt.plot(new_pipe_sensi_P_el, color='dodgerblue',linestyle='-', label = 'New pipeline')
plt.plot(retro_pipe_sensi_P_el, color='royalblue',linestyle='-', label = 'Retrofit pipeline')
plt.grid(True, axis='y')
#plt.grid(True, axis='x')
ax.set_axisbelow(True)
#ax.xaxis.set_major_formatter(mtick.PercentFormatter(10, decimals=None))
plt.locator_params(axis='x', nbins=8)
plt.locator_params(axis='y', nbins=7)

plt.axvline(x=31.8, color='grey', linestyle = '--')
plt.axvline(x=57, color='grey', linestyle = '--')
plt.axvline(x=97, color='grey', linestyle = '--')
plt.axvline(x=84, color='grey', linestyle = '--')
plt.text(31.8,2.9, 'NOR 2021', horizontalalignment='center', verticalalignment='center')
plt.text(57,2.9, 'NOR 2030', horizontalalignment='center', verticalalignment='center')
plt.text(97,2.9, 'GER 2021', horizontalalignment='center', verticalalignment='center')
plt.text(84,2.9, 'GER 2030', horizontalalignment='center', verticalalignment='center')


plt.ylabel('[€/kg H2]')
plt.xlim(0,120)
plt.xlabel('Electricity price [€/MWh]')
plt.legend(loc = 'upper left')



title = '\Transport_P_el_sensi'
plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()

"""### Plot Transport distance sensi LCOT"""

fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#plt.subplot(1,2,1)


plt.plot(lh2_transport_sensi_distance_LCOT, color='blue', linestyle='-', label = 'LH2')
plt.plot(NH3_transport_sensi_distance_LCOT, color='darkorange', linestyle='-', label ='NH3')
plt.plot(new_pipe_sensi_distance_LCOT, color='dodgerblue', linestyle='-', label ='New pipeline')
plt.plot(retro_pipe_sensi_distance_LCOT, color='royalblue', linestyle='-', label ='Retrofit pipeline')
#orientation
d_sea_hammerfest = 2500
d_sea_US = 9500

plt.axvline(x=d_sea, color='grey', linestyle = '--')
plt.axvline(x=d_sea_hammerfest, color='grey', linestyle = '--')
plt.axvline(x=d_sea_US, color='grey', linestyle = '--')

plt.text(av_new, 6, 'Norway Kårstø', horizontalalignment='center', verticalalignment='center')
plt.text(av_retro,6,  'Norway Hammerfest', horizontalalignment='center', verticalalignment='center')
plt.text(av_retro,6,  'USA Texas', horizontalalignment='center', verticalalignment='center')
plt.grid(True, axis='y')
#plt.grid(True, axis='x')
ax.set_axisbelow(True)
plt.locator_params(axis='x', nbins=12)
plt.ylabel('Transport costs [€/kg H2]')
plt.xlim(0,10000)
plt.ylim(0,6)
plt.xlabel('Transport Distance in km')
plt.legend()



title = '\Transport_distance_sensi_LCOT'
plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()

"""# Transport emissions

## Pipeline transport

Compression capacity in [MWel/1000km]
Utilisation of the pipeline in %
8760 h/a
Capacity at full load factor [kg H2/a]
Transport distance in [km]
"""

# Calc. energy use for compression [kWh/kg_H2]
#en_comp = capa_comp * 1000 * pipe_use * 8760 / capa_pipe / 1000 * (d_on + d_sea)
#en_comp

"""Energy use per kg H2 delivered times emission factor for energy use.
en_comp = Total energy use for compression [kWh/kg H2]
EF_y_n = Emission factor of used energy in year y and country n [g CO2eq/kWh]
"""

def Pipeline_emissions():

    result = capa_comp * 1000 * pipe_use * 8760 / capa_pipe / 1000 * (d_sea) * EF_y_n
    return result

"""Calc. pipeline emissions over time

Using grid electricity for compression.
Compressor station for subsea pipelines in exporting country.
Emission factor for grid electricity in year y and country n [g CO2eq/kWh]
Energy use for compression in [kWh/kg H2]
"""

years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['Pipeline_emissions'])
result.index.name = 'Years'
#Baseline
for year in years:

    EF_y_n = float(GHG.loc['Grid electricity emission factor NOR [g CO2eq/kWh] Baseline'][year])

    result.Pipeline_emissions.loc[year] = Pipeline_emissions()

result

Pipeline_emissions = result

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Pipeline_emissions.csv')
Pipeline_emissions.to_csv(output_file, sep=',')



"""### Sensitivity"""



EF_sensi_range = np.arange(0, 401, 20)
sensitivity = []
def Pipeline_emissions_sensi():

    for EF_y_n in EF_sensi_range:

        result = capa_comp * 1000 * pipe_use * 8760 / capa_pipe / 1000 * (d_sea_sensi) * EF_y_n

        sensitivity.append(result)

    return result


Pipeline_emissions_sensi()

Pipeline_emissions_sensi_EF = pd.DataFrame(sensitivity, EF_sensi_range, columns=['Pipeline transport emission [g CO2eq/kg H2]'])
Pipeline_emissions_sensi_EF.index.name = 'Electricity emission factor [g CO2eq/kWh]'
Pipeline_emissions_sensi_EF

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Pipeline_emissions_sensi_EF.csv')
Pipeline_emissions_sensi_EF.to_csv(output_file, sep=',')

EF_sensi_fixed = 118
# Transport sensi emissions
sensitivity = []
def Pipeline_emissions_sensi():

    for d_sea_sensi in transport_distance:

        result = capa_comp * 1000 * pipe_use * 8760 / capa_pipe / 1000 * (d_sea_sensi) * EF_sensi_fixed

        sensitivity.append(result)

    return result


Pipeline_emissions_sensi()

Pipeline_sensi_distance_LEOT = pd.DataFrame(sensitivity, transport_distance, columns=['Pipeline transport emission [g CO2eq/kg H2]'])
Pipeline_sensi_distance_LEOT.index.name = 'Transport distance in km'
Pipeline_sensi_distance_LEOT

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Pipeline_sensi_distance_LEOT.csv')
Pipeline_sensi_distance_LEOT.to_csv(output_file, sep=',')


"""Plot emission breakdown for pipeline transport

## LH2 transport

Liquefaction energy use

Electricity use for liquefaction in [kWh/kg H2]
Emission factor for grid electricity in year y and country n [g CO2eq/kWh]

### Liquefaction
"""

def Conversion_emissions():
    result = el_liq_y * EF_y_n
    return result

# Calc. liquefaction emissions over time [g CO2eq/kg H2] Baseline
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Conversion_emissions'])
for year in years:
    el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
    EF_y_n = float(GHG.loc['Grid electricity emission factor NOR [g CO2eq/kWh] Baseline'][year])

    result.LH2_Conversion_emissions.loc[year] = Conversion_emissions()

result

LH2_Conversion_emissions = result

output_file = os.path.join(path_csv,'LH2_Conversion_emissions.csv')
result.to_csv(output_file, sep = ';')
"""### Export terminal"""

# Electricity consumption for the export terminal [kWh/kg_h2]
el_et = float(tea_lh2.loc['Export Terminal - Electricity use [kWh/kgH2]']['LH2'])
el_et

el_reliq = float(tea_lh2.loc['Export Terminal - Energy use to reliquefy opt. [kWh/kg H2/day]']['LH2'])
el_reliq

# Average storage time in the export terminal tanks in days
t_et = float(tea_lh2.loc['Export Terminal - Storage length per load [Days]']['LH2'])
t_et

def ET_emissions():
    result = (el_et + el_reliq * t_et) * EF_y_n
    return result

# Calc. liquefaction emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Export_terminal_emissions'])
result.index.name = 'Years'

for year in years:
    EF_y_n = float(GHG.loc['Grid electricity emission factor NOR [g CO2eq/kWh] Baseline'][year])

    result.LH2_Export_terminal_emissions.loc[year] = ET_emissions()

result

LH2_Export_terminal_emissions = result


output_file = os.path.join(path_csv,'LH2_Export_terminal_emissions.csv')
result.to_csv(output_file, sep = ';')

"""### Shipping

Calc. emission factor for cargo/fuel

Cost of transported hydrogen in [$/kg_h2] in year y
"""


# Emissions of transported hydrogen carrier in [g CO2eq/kg_H2] in year y
H2_emissions = (LCOH_min_tech_em.loc[::]['Emissions']).apply(pd.to_numeric) * 1000
Con_emissions = (LH2_Conversion_emissions.loc[::]['LH2_Conversion_emissions']).apply(pd.to_numeric)
ET_emissions = (LH2_Export_terminal_emissions.loc[::]['LH2_Export_terminal_emissions']).apply(pd.to_numeric)

# Emission of cargo in [g CO2eq/kg_H2]
LH2_cargo_emissions = (H2_emissions + Con_emissions + ET_emissions).to_frame('LH2_cargo_emissions')
LH2_cargo_emissions

output_file = os.path.join(path_csv,'LH2_cargo_emissions.csv')
result.to_csv(output_file, sep = ';')

# Boil-off when shipping in [%/day] to [%/hour]
bog_ship = float(tea_lh2.loc['Shipping - Boil off opt. [%/day]']['LH2']) / 24
bog_ship

# Fuel consumption of a ship in [kg_h2/t/km]
f_ship = float(tea_lh2.loc['Shipping - Fuel use [kg H2/t/km]']['LH2']) / 1000
f_ship


# Ship speed in [km/h]
v_ship = float(tea_lh2.loc['Shipping - Ship speed [km/h]']['LH2'])
v_ship


"""Calc. LH2 shipping emissions"""

def calculate_ship_emissions():
    result = 1/(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea)) \
             + (bog_ship * d_sea/v_ship + f_ship * d_sea) * LH2_cargo_ghg

    return result

# Calculation of shipping emissions [g CO2eq/kg_h2] from 2025 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Shipping_emissions'])
result.index.name = 'Years'

for year in years:

    LH2_cargo_ghg = float(LH2_cargo_emissions.loc[year]['LH2_cargo_emissions'])

    # calculate emissions of specific year
    result.LH2_Shipping_emissions.loc[year] = calculate_ship_emissions()

result

LH2_Shipping_emissions = result

output_file = os.path.join(path_csv,'LH2_Shipping_emissions.csv')
result.to_csv(output_file, sep = ';')

"""### Import terminal

"""

# Electricity consumption for the import terminal [kWh/kg_h2]
el_it = float(tea_lh2.loc['Import Terminal - Electricity use [kWh/kg H2]']['LH2'])
el_it

# Electricity consumption to reliquefy BOG [kWh/kg_h2]
el_reliq = float(tea_lh2.loc['Import Terminal - Energy use to reliquefy opt. [kWh/kg H2/day]']['LH2'])
el_reliq

# Average storage time in the import terminal tanks in days
t_it = float(tea_lh2.loc['Import Terminal - Storage length per load [days]']['LH2'])
t_it

def IT_emissions():
    result = (el_it + el_reliq * t_it) * EF_y_n
    return result


# Calc. liquefaction emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Import_terminal_emissions'])
result.index.name = 'Years'

for year in years:
    EF_y_n = float(GHG.loc['Grid electricity emission factor GER [g CO2eq/kWh] Baseline'][year])

    result.LH2_Import_terminal_emissions.loc[year] = IT_emissions()

result

LH2_Import_terminal_emissions = result

output_file = os.path.join(path_csv, 'LH2_Import_terminal_emissions.csv')
result.to_csv(output_file, sep=',')



"""### Reconversion

"""

def calculate_recon_emissions():
    result = el_recon_y * EF_y_n
    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Reconversion_emissions'])
result.index.name = 'Years'

for year in years:

    el_recon_y = float(tea_lh2.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
    EF_y_n = float(GHG.loc['Grid electricity emission factor GER [g CO2eq/kWh] Baseline'][year])

    # calculate emissions of specific year
    result.LH2_Reconversion_emissions.loc[year] = calculate_recon_emissions()
result

LH2_Reconversion_emissions = result

output_file = os.path.join(path_csv, 'LH2_Reconversion_emissions.csv')
result.to_csv(output_file, sep=',')



"""### Total LH2 transport emissions

con_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Conversion_emissions.csv', delimiter=';', decimal=',', index_col= 0)
et_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Export_terminal_emissions.csv', delimiter=';', decimal=',', index_col= 0)
ship_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Shipping_emissions.csv', delimiter=';', decimal=',', index_col= 0)
it_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Import_terminal_emissions.csv', delimiter=';', decimal=',', index_col= 0)
recon_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Reconversion_emissions.csv', delimiter=';', decimal=',', index_col= 0)
"""

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_transport_emissions'])
result.index.name = 'Years'

# [g CO2eq/kg H2]
def calculate_LH2_transport_emissions():
    result = CE + EE + SE + IE + RE

    return result

for year in years:
    # get all emissions
    CE = float(LH2_Conversion_emissions.loc[year]['LH2_Conversion_emissions'])
    EE = float(LH2_Export_terminal_emissions.loc[year]['LH2_Export_terminal_emissions'])
    SE = float(LH2_Shipping_emissions.loc[year]['LH2_Shipping_emissions'])
    IE = float(LH2_Import_terminal_emissions.loc[year]['LH2_Import_terminal_emissions'])
    RE = float(LH2_Reconversion_emissions.loc[year]['LH2_Reconversion_emissions'])
    # calculate costs of specific year
    result.LH2_transport_emissions.loc[year] = calculate_LH2_transport_emissions()

result
output_file = os.path.join(path_csv, 'LH2_transport_emissions.csv')
result.to_csv(output_file, sep=',')
LH2_transport_emissions = result


# Emission breakdown for LH2 seaborne transport - Bar chart
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)
CE = (LH2_Conversion_emissions.loc[::5]['LH2_Conversion_emissions']).apply(pd.to_numeric)
EE = (LH2_Export_terminal_emissions.loc[::5]['LH2_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (LH2_Shipping_emissions.loc[::5]['LH2_Shipping_emissions']).apply(pd.to_numeric)
IE = (LH2_Import_terminal_emissions.loc[::5]['LH2_Import_terminal_emissions']).apply(pd.to_numeric)
RE = (LH2_Reconversion_emissions.loc[::5]['LH2_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

RE_plt = plt.bar(x,RE, width, color = 'orange', label='Regasification emissions', bottom= CE + EE + SE + IE)
IE_plt = plt.bar(x,IE, width, color = 'wheat',label='Import terminal emissions', bottom= CE + EE + SE)
SE_plt = plt.bar(x,SE, width, color = 'crimson',label='Shipping emissions', bottom= CE + EE)
EE_plt = plt.bar(x,EE, width, color = 'cornflowerblue',label='Export terminal emissions', bottom= CE)
CE_plt = plt.bar(x,CE, width, color = 'darkblue',label='Liquefaction emissions')

#plt.title('Emission breakdown for LH2 shipping ', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('[g CO2eq/kg H2]')
plt.xlabel('Years')

title = '\LH2_emissions_2'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()
##



# Emission breakdown for LH2 seaborne transport - stackplot
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')

plt.subplot(1,2,1)
x = np.arange(2025, 2051)
CE = (LH2_Conversion_emissions.loc[::]['LH2_Conversion_emissions']).apply(pd.to_numeric)
EE = (LH2_Export_terminal_emissions.loc[::]['LH2_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (LH2_Shipping_emissions.loc[::]['LH2_Shipping_emissions']).apply(pd.to_numeric)
IE = (LH2_Import_terminal_emissions.loc[::]['LH2_Import_terminal_emissions']).apply(pd.to_numeric)
RE = (LH2_Reconversion_emissions.loc[::]['LH2_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Liquefaction', 'Export Terminal', 'Shipping', 'Import Terminal', 'Regasification']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, CE, EE, SE, IE, RE, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.ylim(0,730)
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
#plt.legend(loc='upper right')
plt.ylabel('Transport emissions [g CO2eq/kg H2]')
plt.xlabel('Baseline')

plt.subplot(1,2,2)

fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)

CE = (lh2_con_pol.loc[::]['LH2_Conversion_emissions']).apply(pd.to_numeric)
EE = (Lh2_et_pol.loc[::]['LH2_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (lh2_ship_pol.loc[::]['LH2_Shipping_emissions']).apply(pd.to_numeric)
IE = (Lh2_it_pol.loc[::]['LH2_Import_terminal_emissions']).apply(pd.to_numeric)
RE = (lh2_recon_pol.loc[::]['LH2_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Liquefaction', 'Export Terminal', 'Shipping', 'Import Terminal', 'Regasification']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, CE, EE, SE, IE, RE, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.ylim(0,730)
plt.legend(loc='upper right')
plt.ylabel('Transport emissions [g CO2eq/kg H2]')
plt.xlabel('Policy')
title = '\LH2_emissions_stackplot_base_pol'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()

"""@ 10.000 km shipping distance. Large contributers to total emissions are emissions factors for grid electricity in import/exporting countries (assumed zero from 2045).

### Sensitivity of LH2 to transport distance
"""

year = 2030
el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
EF_y_n = 118# float(GHG.loc['Grid electricity emission factor NOR [g CO2eq/kWh] Baseline'][year])

el_recon_y = float(tea_lh2.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
#EF_y_G = float(GHG.loc['Grid electricity emission factor GER [g CO2eq/kWh] Baseline'][year])
H2_emissions_sensi = 2500
LH2_cargo_ghg_sensi = (H2_emissions_sensi + (el_liq_y * EF_y_n ) + ((el_et + el_reliq * t_et) * EF_y_n ))
sensitivity = []
def lh2_transport_sensi(el_liq_y, EF_y_n, el_et, el_reliq, t_et, bog_ship, d_sea ,v_ship, f_ship, LH2_cargo_ghg ,el_it ,t_it ,el_recon_y):

    for EF_y_n in EF_sensi_range:

        result = \
        (el_liq_y * EF_y_n )\
        + ((el_et + el_reliq * t_et) * EF_y_n )\
        + (1 / (1 - (bog_ship * d_sea_sensi / v_ship) - (f_ship * d_sea_sensi)) + (bog_ship * d_sea_sensi / v_ship + f_ship * d_sea_sensi) * LH2_cargo_ghg_sensi)\
        + ((el_it + el_reliq * t_it) * EF_y_n )\
        + (el_recon_y * EF_y_n )

        sensitivity.append(result)

    return sensitivity


lh2_transport_sensi(el_liq_y, EF_y_n, el_et, el_reliq, t_et, bog_ship, d_sea, v_ship, f_ship, LH2_cargo_ghg, el_it,
                    t_it, el_recon_y)

lh2_transport_sensi_EF= pd.DataFrame(sensitivity, EF_sensi_range, columns=['LH2 transport emission [g CO2eq/kg H2]'])
lh2_transport_sensi_EF.index.name = 'Electricity emission [g CO2eq/kWh]'
lh2_transport_sensi_EF

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'LH2_EF_sensi.csv')
lh2_transport_sensi_EF.to_csv(output_file, sep =';')

import matplotlib.ticker as mtick

# Transport distance sensi emissions
sensitivity = []
def lh2_transport_sensi(el_liq_y, EF_y_n, el_et, el_reliq, t_et, bog_ship, d_sea ,v_ship, f_ship, LH2_cargo_ghg ,el_it ,t_it ,el_recon_y):

    for d_sea_sensi in transport_distance:

        result = \
        (el_liq_y * EF_sensi_fixed)\
        + ((el_et + el_reliq * t_et) * EF_sensi_fixed)\
        + (1 / (1 - (bog_ship * d_sea_sensi / v_ship) - (f_ship * d_sea_sensi)) + (bog_ship * d_sea_sensi / v_ship + f_ship * d_sea_sensi) * LH2_cargo_ghg_sensi)\
        + ((el_it + el_reliq * t_it) * EF_sensi_fixed)\
        + (el_recon_y * EF_sensi_fixed)

        sensitivity.append(result)

    return sensitivity


lh2_transport_sensi(el_liq_y, EF_y_n, el_et, el_reliq, t_et, bog_ship, d_sea, v_ship, f_ship, LH2_cargo_ghg, el_it,
                    t_it, el_recon_y)

lh2_transport_sensi_distance_LEOT= pd.DataFrame(sensitivity, transport_distance, columns=['LH2 transport emission [g CO2eq/kg H2]'])
lh2_transport_sensi_distance_LEOT.index.name = 'Transport distance in km'
lh2_transport_sensi_distance_LEOT

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lh2_transport_sensi_distance_LEOT.csv')
lh2_transport_sensi_distance_LEOT.to_csv(output_file, sep =';')

# LH2 transport distance sensi cargo emissions
cargo_emissions_range = np.arange(0, 10001, 500)
d_sea_sensi_cargo = 10000
sensitivity = []
def lh2_transport_sensi(el_liq_y, EF_y_n, el_et, el_reliq, t_et, bog_ship, d_sea ,v_ship, f_ship, LH2_cargo_ghg ,el_it ,t_it ,el_recon_y):

    for H2_sensi_emissions in cargo_emissions_range:

        result = \
        (el_liq_y * EF_sensi_fixed)\
        + ((el_et + el_reliq * t_et) * EF_sensi_fixed)\
        + (1 / (1 - (bog_ship * d_sea_sensi_cargo / v_ship) - (f_ship * d_sea_sensi_cargo)) + (bog_ship * d_sea_sensi_cargo / v_ship + f_ship * d_sea_sensi_cargo) * ((el_liq_y * EF_sensi_fixed ) + ((el_et + el_reliq * t_et) * EF_sensi_fixed ) + H2_sensi_emissions))\
        + ((el_it + el_reliq * t_it) * EF_sensi_fixed)\
        + (el_recon_y * EF_sensi_fixed)

        sensitivity.append(result)

    return sensitivity


lh2_transport_sensi(el_liq_y, EF_y_n, el_et, el_reliq, t_et, bog_ship, d_sea, v_ship, f_ship, LH2_cargo_ghg, el_it,
                    t_it, el_recon_y)

lh2_transport_sensi_cargo_emissions= pd.DataFrame(sensitivity, cargo_emissions_range, columns=['LH2 transport emission [g CO2eq/kg H2]'])
lh2_transport_sensi_cargo_emissions.index.name = 'H2 Emissions [g CO2eq/kg H2]'
lh2_transport_sensi_cargo_emissions

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lh2_transport_sensi_cargo_emissions.csv')
lh2_transport_sensi_cargo_emissions.to_csv(output_file, sep =';')

"""## NH3 transport

### Conversion emissions
"""
def Conversion_emissions():
    result = el_con_y * EF_y_n
    return result

# Calc. liquefaction emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Conversion_emissions'])
for year in years:
    el_con_y = float(tea_NH3.loc['Conversion - Electricity consumption opt. [kWh/kgH2]'][year])
    EF_y_n = float(GHG.loc['Grid electricity emission factor NOR [g CO2eq/kWh] Baseline'][year])

    result.NH3_Conversion_emissions.loc[year] = Conversion_emissions()

result

output_file = os.path.join(path_csv, 'NH3_Conversion_emissions.csv')
result.to_csv(output_file, sep=',')

NH3_Conversion_emissions = result

"""### Export terminal"""

# Electricity consumption for the export terminal [kWh/kg_h2]
el_et_nh3 = float(tea_NH3.loc['Export Terminal - Electricity use [kWh/kgH2]']['NH3'])
el_et_nh3

el_reliq_nh3 = float(tea_NH3.loc['Export Terminal - Energy use to reliquefy opt. [kWh/kg NH3/day]']['NH3'])
el_reliq_nh3

#Lower heating value of Nh3[ kWh / kg]
NH3_lhv = float(tea_NH3.loc['Properties - LHV [kWh/kg]']['NH3'])
NH3_lhv
#Lower heating value of H2 [kWh/kg]
H2_lhv = 33.33

"""same storage time as for LH2"""

# Average storage time in the export terminal tanks in days
t_et

def ET_emissions():
    result = (el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_y_n
    return result

# Calc. liquefaction emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Export_terminal_emissions'])
result.index.name = 'Years'

for year in years:
    EF_y_n = float(GHG.loc['Grid electricity emission factor NOR [g CO2eq/kWh] Baseline'][year])

    result.NH3_Export_terminal_emissions.loc[year] = ET_emissions()

result


output_file = os.path.join(path_csv, 'NH3_Export_terminal_emissions.csv')
result.to_csv(output_file, sep=',')

NH3_Export_terminal_emissions = result

"""### Shipping

Emissions of transported hydrogen carrier in [g CO2/kg_h2] in year y
NH3_Conversion_emissions = pd.read_csv(
    "/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Conversion_emissions.csv", delimiter=';',
    decimal=',', index_col=0)
NH3_Export_terminal_emissions = pd.read_csv(
    "/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Export_terminal_emissions.csv", delimiter=';',
    decimal=',', index_col=0)
"""

# Boil-off when shipping in [%/day] to [%/hour]
bog_ship_nh3 = float(tea_NH3.loc['Shipping - Boil off opt. [%/day]']['NH3']) / 24 * (NH3_lhv/H2_lhv)
bog_ship_nh3

# Fuel consumption of a ship in [kg_h2/t/km]
f_ship_nh3 = float(tea_NH3.loc['Shipping - Fuel use [kg NH3/t/km]']['NH3']) / 1000 * (NH3_lhv/H2_lhv)


"""d_sea and v_ship are equivalent to LH2 transport

Calc. emission factor for cargo/fuel
"""

# Emissions of transported hydrogen carrier in [g CO2eq/kg_H2] in year y
H2_emissions = (LCOH_min_tech_em.loc[::]['Emissions']).apply(pd.to_numeric) * 1000
Con_emissions = (NH3_Conversion_emissions.loc[::]['NH3_Conversion_emissions']).apply(pd.to_numeric)
ET_emissions = (NH3_Export_terminal_emissions.loc[::]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
# Emission of cargo in [g CO2eq/kg_H2]
NH3_cargo_emissions = (H2_emissions + Con_emissions + ET_emissions).to_frame('NH3_cargo_emissions')
NH3_cargo_emissions

output_file = os.path.join(path_csv, 'NH3_cargo_emissions.csv')
result.to_csv(output_file, sep=',')


# Calc.NH3 shipping emissions
def calculate_ship_emissions():
    result = 1 / (1 - (bog_ship_nh3 * d_sea / v_ship) - (f_ship_nh3 * d_sea))\
             + (bog_ship_nh3 * d_sea / v_ship + f_ship_nh3 * d_sea) * NH3_cargo_ghg
    return result

# Calculation of shipping emissions [g CO2eq/kg_h2] from 2025 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Shipping_emissions'])
result.index.name = 'Years'

for year in years:
    NH3_cargo_ghg = float(LH2_cargo_emissions.loc[year]['LH2_cargo_emissions'])
    f_ship_nh3 = float(tea_NH3.loc['Shipping - Fuel use [kg NH3/t/km]'][year])/1000 * (NH3_lhv/H2_lhv)
    # calculate emissions of specific year
    result.NH3_Shipping_emissions.loc[year] = calculate_ship_emissions()

result

output_file = os.path.join(path_csv, 'NH3_Shipping_emissions.csv')
result.to_csv(output_file, sep=',')

NH3_Shipping_emissions = result

"""### Import terminal"""

# Electricity consumption for the import terminal [kWh/kg_h2]
el_it = float(tea_NH3.loc['Import Terminal - Electricity use [kWh/kg H2]']['NH3'])
el_it

# Electricity consumption to reliquefy BOG [kWh/kg_h2]
el_reliq = float(tea_NH3.loc['Import Terminal - Energy use to reliquefy opt. [kWh/kg NH3/day]']['NH3'])
el_reliq

# Average storage time in the import terminal tanks in days
t_it = float(tea_NH3.loc['Import Terminal - Storage length per load [days]']['NH3'])
t_it

def IT_emissions():
    result = (el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * EF_y_n
    return result

# Calc. import terimal emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Import_terminal_emissions'])
result.index.name = 'Years'

for year in years:
    EF_y_n = float(GHG.loc['Grid electricity emission factor GER [g CO2eq/kWh] Baseline'][year])

    result.NH3_Import_terminal_emissions.loc[year] = IT_emissions()

result

output_file = os.path.join(path_csv, 'NH3_Import_terminal_emissions.csv')
result.to_csv(output_file, sep=',')

NH3_Import_terminal_emissions = result

"""### Reconversion"""

def calculate_recon_emissions():
    result = (heat_recon_y + el_recon_y) * EF_y_n
    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Reconversion_emissions'])
result.index.name = 'Years'

for year in years:
    el_recon_y = float(tea_NH3.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
    heat_recon_y = float(tea_NH3.loc['Reconversion - Heat consumption opt. [kWh/kg H2]'][year])
    EF_y_n = float(GHG.loc['Grid electricity emission factor GER [g CO2eq/kWh] Baseline'][year])

     # calculate emissions of specific year
    result.NH3_Reconversion_emissions.loc[year] = calculate_recon_emissions()
result


output_file = os.path.join(path_csv, 'NH3_Reconversion_emissions.csv')
result.to_csv(output_file, sep=',')

NH3_Reconversion_emissions = result

"""### Total NH3 transport emissions

con_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Conversion_emissions.csv',
                     delimiter=';', decimal=',', index_col=0)
et_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Export_terminal_emissions.csv',
                    delimiter=';', decimal=',', index_col=0)
ship_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Shipping_emissions.csv',
                      delimiter=';', decimal=',', index_col=0)
it_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Import_terminal_emissions.csv',
                    delimiter=';', decimal=',', index_col=0)
recon_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Reconversion_emissions.csv',
                       delimiter=';', decimal=',', index_col=0)
"""

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_transport_emissions'])
result.index.name = 'Years'

# [g CO2eq/kg H2]
def calculate_NH3_transport_emissions():
    result = CE + EE + SE + IE + RE

    return result

for year in years:
    # get all emissions
    CE = float(NH3_Conversion_emissions.loc[year]['NH3_Conversion_emissions'])
    EE = float(NH3_Export_terminal_emissions.loc[year]['NH3_Export_terminal_emissions'])
    SE = float(NH3_Shipping_emissions.loc[year]['NH3_Shipping_emissions'])
    IE = float(NH3_Import_terminal_emissions.loc[year]['NH3_Import_terminal_emissions'])
    RE = float(NH3_Reconversion_emissions.loc[year]['NH3_Reconversion_emissions'])
    # calculate costs of specific year
    result.NH3_transport_emissions.loc[year] = calculate_NH3_transport_emissions()

result

output_file = os.path.join(path_csv, 'NH3_transport_emissions.csv')
result.to_csv(output_file, sep=',')
NH3_transport_emissions = result

def calculate_NH3_transport_emissions_wo_recon():
    result = CE + EE + SE + IE# + RE

    return result

for year in years:
    # get all emissions
    CE = float(NH3_Conversion_emissions.loc[year]['NH3_Conversion_emissions'])
    EE = float(NH3_Export_terminal_emissions.loc[year]['NH3_Export_terminal_emissions'])
    SE = float(NH3_Shipping_emissions.loc[year]['NH3_Shipping_emissions'])
    IE = float(NH3_Import_terminal_emissions.loc[year]['NH3_Import_terminal_emissions'])
    #RE = float(NH3_Reconversion_emissions.loc[year]['NH3_Reconversion_emissions'])
    # calculate costs of specific year
    result.NH3_transport_emissions.loc[year] = calculate_NH3_transport_emissions_wo_recon()

result

output_file = os.path.join(path_csv, 'NH3_transport_emissions_wo_recon.csv')
result.to_csv(output_file, sep=',')
NH3_transport_emissions_wo_recon = result

# Emission breakdown for NH3 seaborne transport with cracking
fig, ax = plt.subplots(figsize=(10, 6), layout = 'constrained')
plt.grid(True, axis='y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)
CE = (NH3_Conversion_emissions.loc[::5]['NH3_Conversion_emissions']).apply(pd.to_numeric)
EE = (NH3_Export_terminal_emissions.loc[::5]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (NH3_Shipping_emissions.loc[::5]['NH3_Shipping_emissions']).apply(pd.to_numeric)
IE = (NH3_Import_terminal_emissions.loc[::5]['NH3_Import_terminal_emissions']).apply(pd.to_numeric)
RE = (NH3_Reconversion_emissions.loc[::5]['NH3_Reconversion_emissions']).apply(pd.to_numeric)
width = 2  # the width of the bars: can also be len(x) sequence

RE_plt = plt.bar(x, RE, width, color = 'orange', label='Reconversion (Ammonia cracking) emissions', bottom=CE + EE + SE + IE)
IE_plt = plt.bar(x, IE, width, color = 'wheat',label='Import terminal emissions', bottom=CE + EE + SE)
SE_plt = plt.bar(x, SE, width, color = 'crimson',label='Shipping emissions', bottom=CE + EE)
EE_plt = plt.bar(x, EE, width, color = 'cornflowerblue', label='Export terminal emissions', bottom=CE)
CE_plt = plt.bar(x, CE, width, color = 'darkblue',label='Conversion (Ammonia synthesis) emissions')

plt.title('Emission breakdown for NH3 shipping [g CO2eq/kg H2]', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('[g CO2eq/kg H2]')

title = '\Ammonia_emissions_w_reconversion'
plt.savefig(path_plt+title+'.png', transparent = True)

plt.show()
#nh3 emission breakdown base pol
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#baseline
plt.subplot(1,2,1)
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
CE = (NH3_Conversion_emissions.loc[::]['NH3_Conversion_emissions']).apply(pd.to_numeric)
EE = (NH3_Export_terminal_emissions.loc[::]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (NH3_Shipping_emissions.loc[::]['NH3_Shipping_emissions']).apply(pd.to_numeric)
IE = (NH3_Import_terminal_emissions.loc[::]['NH3_Import_terminal_emissions']).apply(pd.to_numeric)
RE = (NH3_Reconversion_emissions.loc[::]['NH3_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Conversion', 'Export Terminal', 'Shipping', 'Import Terminal', 'Reconversion']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, CE, EE, SE, IE, RE, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.ylim(0,5000)
plt.legend(loc='upper right')
plt.ylabel('Transport Emissions [g CO2eq/kg H2]')
plt.xlabel('Baseline')
#policy
plt.subplot(1,2,2)
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
CE = (policy.NH3_Conversion_emissions.loc[::]['NH3_Conversion_emissions']).apply(pd.to_numeric)
EE = (policy.NH3_Export_terminal_emissions.loc[::]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (policy.NH3_Shipping_emissions.loc[::]['NH3_Shipping_emissions']).apply(pd.to_numeric)
IE = (policy.NH3_Import_terminal_emissions.loc[::]['NH3_Import_terminal_emissions']).apply(pd.to_numeric)
RE = (policy.NH3_Reconversion_emissions.loc[::]['NH3_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Conversion', 'Export Terminal', 'Shipping', 'Import Terminal', 'Reconversion']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, CE, EE, SE, IE, RE, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.ylim(0,5000)
plt.legend(loc='upper right')
plt.ylabel('Transport Emissions [g CO2eq/kg H2]')
plt.xlabel('Policy')
#save
title = '\Ammonia_emissions_stackplot_base_pol'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()
## nh3 emission breakdown without cracking base pol
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#baseline
plt.subplot(1,2,1)
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
CE = (NH3_Conversion_emissions.loc[::]['NH3_Conversion_emissions']).apply(pd.to_numeric)
EE = (NH3_Export_terminal_emissions.loc[::]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (NH3_Shipping_emissions.loc[::]['NH3_Shipping_emissions']).apply(pd.to_numeric)
IE = (NH3_Import_terminal_emissions.loc[::]['NH3_Import_terminal_emissions']).apply(pd.to_numeric)
#RE = (NH3_Reconversion_emissions.loc[::]['NH3_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Conversion', 'Export Terminal', 'Shipping', 'Import Terminal']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, CE, EE, SE, IE, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.ylim(0,140)
plt.legend(loc='upper right')
plt.ylabel('Transport Emissions [g CO2eq/kg H2]')
plt.xlabel('Baseline')
#policy
plt.subplot(1,2,2)
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
CE = (policy.NH3_Conversion_emissions.loc[::]['NH3_Conversion_emissions']).apply(pd.to_numeric)
EE = (policy.NH3_Export_terminal_emissions.loc[::]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (policy.NH3_Shipping_emissions.loc[::]['NH3_Shipping_emissions']).apply(pd.to_numeric)
IE = (policy.NH3_Import_terminal_emissions.loc[::]['NH3_Import_terminal_emissions']).apply(pd.to_numeric)
#RE = (policy.NH3_Reconversion_emissions.loc[::]['NH3_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Conversion', 'Export Terminal', 'Shipping', 'Import Terminal']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat', 'orange']
plt.stackplot(x, CE, EE, SE, IE, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.ylim(0,140)
plt.legend(loc='upper right')
plt.ylabel('Transport Emissions [g CO2eq/kg H2]')
plt.xlabel('Policy')
title = '\Ammonia_emissions_stackplot_wo_cracking_base_pol'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()
"""Largest contributer is reconversion at around 15 kWh/kg H2 are required to crack ammonia and it is assumed that electricity and heat are sourced from grid electricity, which has a high emission factor of around 300 g CO2eq/kWh in 2025 (in Germany).
"""
# Emission breakdown for NH3 seaborne transport without cracking
fig, ax = plt.subplots(figsize=(10, 6), layout = 'constrained')
plt.grid(True, axis='y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)
CE = (NH3_Conversion_emissions.loc[::5]['NH3_Conversion_emissions']).apply(pd.to_numeric)
EE = (NH3_Export_terminal_emissions.loc[::5]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (NH3_Shipping_emissions.loc[::5]['NH3_Shipping_emissions']).apply(pd.to_numeric)
IE = (NH3_Import_terminal_emissions.loc[::5]['NH3_Import_terminal_emissions']).apply(pd.to_numeric)
#RE = (NH3_Reconversion_emissions.loc[::5]['NH3_Reconversion_emissions']).apply(pd.to_numeric)
width = 2  # the width of the bars: can also be len(x) sequence

#RE_plt = plt.bar(x, RE, width, label='Reconversion (Ammonia cracking) emissions', bottom=CE + EE + SE + IE)
IE_plt = plt.bar(x, IE, width, color = 'wheat', label='Import terminal emissions', bottom=CE + EE + SE)
SE_plt = plt.bar(x, SE, width, color = 'crimson',label='Shipping emissions', bottom=CE + EE)
EE_plt = plt.bar(x, EE, width, color = 'cornflowerblue',label='Export terminal emissions', bottom=CE)
CE_plt = plt.bar(x, CE, width, color = 'darkblue',label='Conversion (Ammonia synthesis) emissions')

plt.title('Emission breakdown for NH3 shipping w/o cracking [g CO2eq/kg H2]', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('[g CO2eq/kg H2]')

title = '\Ammonia_emissions_wo_reconversion'
plt.savefig(path_plt+title+'.png', transparent = True)

plt.show()
#stackplot
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
CE = (NH3_Conversion_emissions.loc[::]['NH3_Conversion_emissions']).apply(pd.to_numeric)
EE = (NH3_Export_terminal_emissions.loc[::]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (NH3_Shipping_emissions.loc[::]['NH3_Shipping_emissions']).apply(pd.to_numeric)
IE = (NH3_Import_terminal_emissions.loc[::]['NH3_Import_terminal_emissions']).apply(pd.to_numeric)
#RE = (NH3_Reconversion_emissions.loc[::]['NH3_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Conversion', 'Export Terminal', 'Shipping', 'Import Terminal']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat']
plt.stackplot(x, CE, EE, SE, IE, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.legend(loc='upper right')
plt.ylabel('[g CO2eq/kg H2]')

title = '\Ammonia_emissions_wo_reconversion_stackplot'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()

#Ammonia emissions stackplot
fig, ax = plt.subplots(figsize=(10,10), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
plt.grid(True, axis = 'x')
ax.set_axisbelow(True)
x = np.arange(2025, 2051)
CE = (NH3_Conversion_emissions.loc[::]['NH3_Conversion_emissions']).apply(pd.to_numeric)
EE = (NH3_Export_terminal_emissions.loc[::]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (NH3_Shipping_emissions.loc[::]['NH3_Shipping_emissions']).apply(pd.to_numeric)
IE = (NH3_Import_terminal_emissions.loc[::]['NH3_Import_terminal_emissions']).apply(pd.to_numeric)
#RE = (NH3_Reconversion_emissions.loc[::]['NH3_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

labels = ['Conversion', 'Export Terminal', 'Shipping', 'Import Terminal']
colors = ['darkblue', 'cornflowerblue', 'crimson', 'wheat']
plt.stackplot(x, CE, EE, SE, IE, labels = labels, colors = colors )

plt.xlim(2025,2050)
plt.legend(loc='upper right')
plt.ylabel('[g CO2eq/kg H2]')

title = '\Ammonia_emissions_wo_reconversion_stackplot'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()
### Sensitivity


year = 2030

el_con_y = float(tea_NH3.loc['Conversion - Electricity consumption opt. [kWh/kgH2]'][year])
NH3_cargo_ghg = float(LH2_cargo_emissions.loc[year]['LH2_cargo_emissions'])
f_ship_nh3 = float(tea_NH3.loc['Shipping - Fuel use [kg NH3/t/km]'][year])/1000 * (NH3_lhv/H2_lhv)
el_recon_y = float(tea_NH3.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
heat_recon_y = float(tea_NH3.loc['Reconversion - Heat consumption opt. [kWh/kg H2]'][year])
EF_y_n = 118  # float(GHG.loc['Grid electricity emission factor NOR [g CO2eq/kWh] Baseline'][year])

NH3_cargo_ghg_sensi = (H2_emissions_sensi + (el_con_y * EF_y_n ) + ((el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_y_n ))
sensitivity = []

def NH3_sensi_w_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg, el_it, heat_recon_y, t_it, el_recon_y):

    for EF_y_n in EF_sensi_range:

        result =\
            (el_con_y * EF_y_n )\
        + ((el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_y_n )\
        + (1 / (1 - (bog_ship_nh3 * d_sea_sensi / v_ship) - (f_ship_nh3 * d_sea_sensi)) + (bog_ship_nh3 * d_sea_sensi / v_ship + f_ship_nh3 * d_sea_sensi) * NH3_cargo_ghg_sensi)\
        + ((el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * EF_y_n )\
        + ((heat_recon_y + el_recon_y) * EF_y_n )

        sensitivity.append(result)


    return sensitivity


NH3_sensi_w_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg,
                  el_it, heat_recon_y, t_it, el_recon_y)

NH3_sensi_w_recon_EF = pd.DataFrame(sensitivity, EF_sensi_range, columns=['NH3 transport emission [g CO2eq/kg H2]'])
NH3_sensi_w_recon_EF.index.name = 'Electricity emission [g CO2eq/KWh]]'
# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'NH3_sensi_w_recon_EF.csv')
NH3_sensi_w_recon_EF.to_csv(output_file, sep=',')

sensitivity = []
def NH3_sensi_wo_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg, el_it, heat_recon_y, t_it, el_recon_y):

    for EF_y_n in EF_sensi_range:

        result =\
            (el_con_y * EF_y_n )\
        + ((el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_y_n )\
        + (1 / (1 - (bog_ship_nh3 * d_sea_sensi / v_ship) - (f_ship_nh3 * d_sea_sensi)) + (bog_ship_nh3 * d_sea_sensi / v_ship + f_ship_nh3 * d_sea_sensi) * NH3_cargo_ghg_sensi)\
        + ((el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * EF_y_n )
        #+ ((heat_recon_y + el_recon_y) * EF_y_n )

        sensitivity.append(result)


    return sensitivity

# recall function
NH3_sensi_wo_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg,
                   el_it, heat_recon_y, t_it, el_recon_y)

NH3_sensi_wo_recon_EF = pd.DataFrame(sensitivity, EF_sensi_range, columns=['NH3 transport emission w/o recon [g CO2eq/kg H2]'])
NH3_sensi_wo_recon_EF.index.name = 'Electricity emission factor [g CO2eq/kWh]'
# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'NH3_sensi_wo_recon_EF.csv')
NH3_sensi_wo_recon_EF.to_csv(output_file, sep=',')

# Transport distance sensi NH3 emissions
sensitivity = []
def NH3_sensi_w_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg, el_it, heat_recon_y, t_it, el_recon_y):

    for d_sea_sensi in transport_distance:

        result =\
            (el_con_y * EF_sensi_fixed)\
        + ((el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_sensi_fixed)\
        + (1 / (1 - (bog_ship_nh3 * d_sea_sensi / v_ship) - (f_ship_nh3 * d_sea_sensi)) + (bog_ship_nh3 * d_sea_sensi / v_ship + f_ship_nh3 * d_sea_sensi) * NH3_cargo_ghg_sensi)\
        + ((el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * EF_sensi_fixed)\
        + ((heat_recon_y + el_recon_y) * EF_sensi_fixed)

        sensitivity.append(result)


    return sensitivity


NH3_sensi_w_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg,
                  el_it, heat_recon_y, t_it, el_recon_y)

NH3_sensi_w_recon_distance_LEOT = pd.DataFrame(sensitivity, transport_distance, columns=['NH3 transport emission [g CO2eq/kg H2]'])
NH3_sensi_w_recon_distance_LEOT.index.name = 'Transport distance in km'
# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'NH3_sensi_w_recon_distance_LEOT.csv')
NH3_sensi_w_recon_distance_LEOT.to_csv(output_file, sep=',')

sensitivity = []
def NH3_sensi_wo_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg, el_it, heat_recon_y, t_it, el_recon_y):

    for d_sea_sensi in transport_distance:

        result =\
            (el_con_y * EF_sensi_fixed)\
        + ((el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_sensi_fixed)\
        + (1 / (1 - (bog_ship_nh3 * d_sea_sensi / v_ship) - (f_ship_nh3 * d_sea_sensi)) + (bog_ship_nh3 * d_sea_sensi / v_ship + f_ship_nh3 * d_sea_sensi) * NH3_cargo_ghg_sensi)\
        + ((el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * EF_sensi_fixed)
        #+ ((heat_recon_y + el_recon_y) * EF_y_n )

        sensitivity.append(result)


    return sensitivity

# recall function
NH3_sensi_wo_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg,
                   el_it, heat_recon_y, t_it, el_recon_y)

NH3_sensi_wo_recon_distance_LEOT = pd.DataFrame(sensitivity, transport_distance, columns=['NH3 transport emission w/o recon [g CO2eq/kg H2]'])
NH3_sensi_wo_recon_distance_LEOT.index.name = 'Electricity emission factor [g CO2eq/kWh]'
# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'NH3_sensi_wo_recon_distance_LEOT.csv')
NH3_sensi_wo_recon_distance_LEOT.to_csv(output_file, sep=',')

# NH3 transport sensi cargo emissions


sensitivity = []
def NH3_sensi_w_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg, el_it, heat_recon_y, t_it, el_recon_y):

    for H2_sensi_emissions in cargo_emissions_range:

        result =\
            (el_con_y * EF_sensi_fixed)\
        + ((el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_sensi_fixed)\
        + (1 / (1 - (bog_ship_nh3 * d_sea_sensi_cargo / v_ship) - (f_ship_nh3 * d_sea_sensi_cargo)) + (bog_ship_nh3 * d_sea_sensi_cargo / v_ship + f_ship_nh3 * d_sea_sensi_cargo) * ((el_con_y * EF_sensi_fixed) + ((el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_sensi_fixed) + H2_sensi_emissions))\
        + ((el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * EF_sensi_fixed)\
        + ((heat_recon_y + el_recon_y) * EF_sensi_fixed)

        sensitivity.append(result)


    return sensitivity


NH3_sensi_w_recon(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship_nh3, d_sea, v_ship, f_ship_nh3, NH3_cargo_ghg,
                  el_it, heat_recon_y, t_it, el_recon_y)

NH3_sensi_w_recon_cargo_emissions = pd.DataFrame(sensitivity, cargo_emissions_range, columns=['NH3 transport emission [g CO2eq/kg H2]'])
NH3_sensi_w_recon_cargo_emissions.index.name = 'H2 Emissions [g CO2eq/kg H2]'
# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'NH3_sensi_w_recon_cargo_emissions.csv')
NH3_sensi_w_recon_cargo_emissions.to_csv(output_file, sep=',')

"""## Plots"""
# Plot transport sensi EF
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#plt.subplot(1,2,1)


plt.plot(lh2_transport_sensi_EF, color='blue', linestyle='solid', label ='LH2')
plt.plot(NH3_sensi_w_recon_EF, color='darkorange', linestyle='-', label ='NH3')
plt.plot(NH3_sensi_wo_recon_EF, color='darkorange', linestyle='--', label ='NH3 w/o cracking')
plt.plot(Pipeline_emissions_sensi_EF, color='cornflowerblue', linestyle='-', label ='Pipeline')

plt.axvline(x=30, color='grey', linestyle = '--')
plt.axvline(x=275, color='grey', linestyle = '--')
plt.axvline(x=118, color='grey', linestyle = '--')
plt.axvline(x=367, color='grey', linestyle = '--')
plt.text(30,7600, 'NOR 2021', horizontalalignment='center', verticalalignment='center')
plt.text(367,7600, 'GER 2021', horizontalalignment='center', verticalalignment='center')
plt.text(275,7600, 'EU 2021', horizontalalignment='center', verticalalignment='center')
plt.text(118,7600, 'EU 2030', horizontalalignment='center', verticalalignment='center')

plt.grid(True, axis='y')
plt.xlim(0,400)
plt.ylim(0,)
#plt.grid(True, axis='x')
ax.set_axisbelow(True)
plt.locator_params(axis='x', nbins=12)
plt.ylabel('[g CO2eq/kg H2]')
plt.xlabel('Electricity emission factor [g CO2eq/kWh]')
#plt.ylim(0,3000)
plt.legend(loc = 'upper left')


title = '\Transport_EF_sensi'
plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()

"""### Plot Transport distance sensi Emissions"""

fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#plt.subplot(1,2,1)


plt.plot(lh2_transport_sensi_distance_LEOT, color='blue', linestyle='-', label = 'LH2')
plt.plot(NH3_sensi_w_recon_distance_LEOT, color='darkorange', linestyle='-', label ='NH3')
plt.plot(NH3_sensi_wo_recon_distance_LEOT, color='darkorange', linestyle='--', label ='NH3 w/o cracking')
plt.plot(Pipeline_sensi_distance_LEOT, color='royalblue', linestyle='-', label ='Pipeline')
plt.grid(True, axis='y')
#plt.grid(True, axis='x')
ax.set_axisbelow(True)

plt.ylabel('[g CO2eq/kg H2]')
plt.xlim(0,10000)
plt.ylim(0,)
plt.xlabel('Transport Distance in km')
plt.legend()



title = '\Transport_distance_sensi_emissions'
plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()

"""### Plot Transport sensi Cargo emissions"""

fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#plt.subplot(1,2,1)
plt.plot(lh2_transport_sensi_cargo_emissions, color='blue', linestyle='-', label = 'LH2')
plt.plot(NH3_sensi_w_recon_cargo_emissions, color='darkorange', linestyle='-', label ='NH3')

plt.grid(True, axis='y')
#plt.grid(True, axis='x')
ax.set_axisbelow(True)

plt.ylabel('[g CO2eq/kg H2]')
#plt.xlim(0,10000)
#plt.ylim(0,)
plt.xlabel('Cargo H2 emissions [g CO2eq/kg H2]')
plt.legend()



title = '\Transport_sensi_cargo_emissions'
plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()


# Plot transport emission comparison

# Import from Policy calculations

sensi_year = 2030
PE_base = (Pipeline_emissions.loc[sensi_year]['Pipeline_emissions'])
PE_pol = (pipe_em_pol.loc[sensi_year]['Pipeline_emissions'])

fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)


#Pipe em
pipe_base = plt.bar(0, PE_base, color = 'dodgerblue')
pipe_pol = plt.bar(1, PE_pol, color = 'royalblue')
#Lh2 em base
CE = (LH2_Conversion_emissions.loc[sensi_year]['LH2_Conversion_emissions'])
EE = (LH2_Export_terminal_emissions.loc[sensi_year]['LH2_Export_terminal_emissions'])
SE = (LH2_Shipping_emissions.loc[sensi_year]['LH2_Shipping_emissions'])
IE = (LH2_Import_terminal_emissions.loc[sensi_year]['LH2_Import_terminal_emissions'])
RE = (LH2_Reconversion_emissions.loc[sensi_year]['LH2_Reconversion_emissions'])
x = 2
RE_plt = plt.bar(x ,RE,  color = 'orange', bottom= CE + EE + SE + IE)
IE_plt = plt.bar(x ,IE,  color = 'wheat', bottom= CE + EE + SE)
SE_plt = plt.bar(x ,SE,  color = 'crimson', bottom= CE + EE)
EE_plt = plt.bar(x ,EE, color = 'cornflowerblue', bottom= CE)
CE_plt = plt.bar(x ,CE, color = 'darkblue')

#lh2 em pol
CE = (lh2_con_pol.loc[sensi_year]['LH2_Conversion_emissions'])
EE = (Lh2_et_pol.loc[sensi_year]['LH2_Export_terminal_emissions'])
SE = (lh2_ship_pol.loc[sensi_year]['LH2_Shipping_emissions'])
IE = (Lh2_it_pol.loc[sensi_year]['LH2_Import_terminal_emissions'])
RE = (lh2_recon_pol.loc[sensi_year]['LH2_Reconversion_emissions'])
x = 3
RE_plt = plt.bar(x ,RE,  color = 'orange', bottom= CE + EE + SE + IE)
IE_plt = plt.bar(x ,IE,  color = 'wheat', bottom= CE + EE + SE)
SE_plt = plt.bar(x ,SE,  color = 'crimson', bottom= CE + EE)
EE_plt = plt.bar(x ,EE, color = 'cornflowerblue', bottom= CE)
CE_plt = plt.bar(x ,CE, color = 'darkblue')
#NH3 em base
CE = (NH3_Conversion_emissions.loc[sensi_year]['NH3_Conversion_emissions'])
EE = (NH3_Export_terminal_emissions.loc[sensi_year]['NH3_Export_terminal_emissions'])
SE = (NH3_Shipping_emissions.loc[sensi_year]['NH3_Shipping_emissions'])
IE = (NH3_Import_terminal_emissions.loc[sensi_year]['NH3_Import_terminal_emissions'])
RE = (NH3_Reconversion_emissions.loc[sensi_year]['NH3_Reconversion_emissions'])

n = 4
#plt.bar(n, RE,  color = 'orange', bottom=CE + EE + SE + IE)
plt.bar(n, IE, color = 'wheat', bottom=CE + EE + SE)
plt.bar(n, SE, color = 'crimson', bottom=CE + EE)
plt.bar(n, EE, color = 'cornflowerblue',  bottom=CE)
plt.bar(n, CE, color = 'darkblue')
#NH3 em pol
CE = (NH3_con_pol.loc[sensi_year]['NH3_Conversion_emissions'])
EE = (NH3_et_pol.loc[sensi_year]['NH3_Export_terminal_emissions'])
SE = (NH3_ship_pol.loc[sensi_year]['NH3_Shipping_emissions'])
IE = (NH3_it_pol.loc[sensi_year]['NH3_Import_terminal_emissions'])
RE = (NH3_recon_pol.loc[sensi_year]['NH3_Reconversion_emissions'])

n = 5
#plt.bar(n, RE,  color = 'orange', label='Reconversion emissions', bottom=CE + EE + SE + IE)
plt.bar(n, IE, color = 'wheat',label='Import terminal emissions', bottom=CE + EE + SE)
plt.bar(n, SE, color = 'crimson',label='Shipping emissions', bottom=CE + EE)
plt.bar(n, EE, color = 'cornflowerblue', label='Export terminal emissions', bottom=CE)
plt.bar(n, CE, color = 'darkblue',label='Conversion emissions')


labels = ['Pipeline\nBase','Pipeline\nPolicy', 'LH2\nBase', 'LH2\nPolicy','NH3\nBase', 'NH3\nPolicy',]
#plt.title('Emission breakdown for LH2 shipping ', fontweight='bold')
plt.legend()
plt.xticks(range(6),labels)

plt.ylabel('[g CO2eq/kg H2]')

title = '\Transport_emission_comparison_wo_recon'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()

##plot transport emissions over time
# Plot emission curves of hydrogen transport
fig, ax = plt.subplots(figsize=(10, 4), layout = 'constrained')
plt.plot(LH2_transport_emissions, color='blue', linestyle='solid', label='LH2')
plt.plot(NH3_transport_emissions, color='darkorange', linestyle='solid', label='NH3')
plt.plot(NH3_transport_emissions_wo_recon, color='darkorange', linestyle='--', label='NH3 w/o cracking')

plt.plot(Pipeline_emissions, color='dodgerblue', linestyle='solid', label='New pipeline')

#ax.yaxis.set_major_locator(mtick.LinearLocator(7))
plt.axvline(x=av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=av_retro, color='royalblue', linestyle = '--')
plt.text(av_new, 3, 'New pipeline\navailable', horizontalalignment='center', verticalalignment='center')
plt.text(av_retro,3,  'Retrofit pipeline\navailable', horizontalalignment='center', verticalalignment='center')
plt.xlim(2025,2050)
plt.ylim(0,)
plt.grid(True, axis='y')
ax.set_axisbelow(True)
#plt.title('Hydrogen transport costs ', fontweight='bold')
ax.legend()
plt.ylabel('Transport emissions [g CO2eq/kg H2]')
title = '\LEOT_time'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

# Plot transport cost comparison
pipe_costs_new_pol = policy.New_Pipeline_costs_off
pipe_costs_retro_pol = policy.Retrofit_pipeline_costs_off

LH2_Liquefaction_costs_pol = policy.LH2_Liquefaction_costs
LH2_Export_terminal_costs_pol = policy.LH2_Export_terminal_costs
LH2_Shipping_costs_pol = policy.LH2_Shipping_costs
LH2_Import_terminal_costs_pol = policy.LH2_Import_terminal_costs
LH2_Reconversion_costs_pol = policy.LH2_Reconversion_costs

NH3_Conversion_costs_pol = policy.NH3_Conversion_costs
NH3_Export_terminal_costs_pol = policy.NH3_Export_terminal_costs
NH3_Shipping_costs_pol = policy.NH3_Shipping_costs
NH3_Import_terminal_costs_pol = policy.NH3_Import_terminal_costs
NH3_Reconversion_costs_pol =policy.NH3_Reconversion_costs


fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)


#Pipe retro base
PE_retro = (Retrofit_pipeline_costs_off.loc[sensi_year]['Retrofit_pipeline_costs_off'])
pipe_retro = plt.bar(0, PE_retro, color = 'royalblue')

#Pipe retro Pol
PE_retro = (pipe_costs_retro_pol.loc[sensi_year]['Retrofit_pipeline_costs_off'])
pipe_retro = plt.bar(1, PE_retro, color = 'royalblue')

#Pipe new base
PE = (New_Pipeline_costs_off.loc[sensi_year]['New_Pipeline_costs_off'])
pipe_new = plt.bar(2, PE, color ='dodgerblue')

#Pipe new Pol
PE = (pipe_costs_new_pol.loc[sensi_year]['New_Pipeline_costs_off'])
pipe_new = plt.bar(3, PE, color ='dodgerblue')

#Lh2 base
LC = (LH2_Liquefaction_costs.loc[sensi_year]['LH2_Liquefaction_costs'])
EC = (LH2_Export_terminal_costs.loc[sensi_year]['LH2_Export_terminal_costs'])
SC = (LH2_Shipping_costs.loc[sensi_year]['LH2_Shipping_costs'])
IC = (LH2_Import_terminal_costs.loc[sensi_year]['LH2_Import_terminal_costs'])
RC = (LH2_Reconversion_costs.loc[sensi_year]['LH2_Reconversion_costs'])

x = 4
RC_LH2 = plt.bar(x,RC , color= 'orange',bottom= LC + EC + SC + IC)
IC_LH2 = plt.bar(x,IC , color= 'wheat',bottom= LC + EC + SC)
SC_LH2 = plt.bar(x,SC , color= 'crimson', bottom= LC + EC)
EC_LH2 = plt.bar(x,EC , color= 'cornflowerblue',bottom= LC)
LC_LH2 = plt.bar(x,LC , color= 'darkblue')
#Lh2 pol
LC = (LH2_Liquefaction_costs_pol.loc[sensi_year]['LH2_Liquefaction_costs'])
EC = (LH2_Export_terminal_costs_pol.loc[sensi_year]['LH2_Export_terminal_costs'])
SC = (LH2_Shipping_costs_pol.loc[sensi_year]['LH2_Shipping_costs'])
IC = (LH2_Import_terminal_costs_pol.loc[sensi_year]['LH2_Import_terminal_costs'])
RC = (LH2_Reconversion_costs_pol.loc[sensi_year]['LH2_Reconversion_costs'])

x = 5
RC_LH2 = plt.bar(x,RC , color= 'orange',bottom= LC + EC + SC + IC)
IC_LH2 = plt.bar(x,IC , color= 'wheat',bottom= LC + EC + SC)
SC_LH2 = plt.bar(x,SC , color= 'crimson', bottom= LC + EC)
EC_LH2 = plt.bar(x,EC , color= 'cornflowerblue',bottom= LC)
LC_LH2 = plt.bar(x,LC , color= 'darkblue')
#NH3 base
CC = (con_costs.loc[sensi_year]['NH3_Conversion_costs'])
EC = (et_costs.loc[sensi_year]['NH3_Export_terminal_costs'])
SC = (ship_costs.loc[sensi_year]['NH3_Shipping_costs'])
IC = (it_costs.loc[sensi_year]['NH3_Import_terminal_costs'])
RC = (recon_costs.loc[sensi_year]['NH3_Reconversion_costs'])

n = 6
RC_NH3 = plt.bar(n,RC , color= 'orange', bottom= CC + EC + SC + IC)
IC_NH3 = plt.bar(n,IC , color = 'wheat',  bottom= CC + EC + SC)
SC_NH3 = plt.bar(n,SC ,color = 'crimson',  bottom= CC + EC)
EC_NH3 = plt.bar(n,EC , color = 'cornflowerblue', bottom= CC)
CC_NH3 = plt.bar(n,CC , color = 'darkblue')
#NH3 pol
CC = (NH3_Conversion_costs_pol.loc[sensi_year]['NH3_Conversion_costs'])
EC = (NH3_Export_terminal_costs_pol.loc[sensi_year]['NH3_Export_terminal_costs'])
SC = (NH3_Shipping_costs_pol.loc[sensi_year]['NH3_Shipping_costs'])
IC = (NH3_Import_terminal_costs_pol.loc[sensi_year]['NH3_Import_terminal_costs'])
RC = (NH3_Reconversion_costs_pol.loc[sensi_year]['NH3_Reconversion_costs'])

n = 7
RC_NH3 = plt.bar(n,RC , color= 'orange', label='Reconversion costs',bottom= CC + EC + SC + IC)
IC_NH3 = plt.bar(n,IC , color = 'wheat', label='Import terminal costs', bottom= CC + EC + SC)
SC_NH3 = plt.bar(n,SC ,color = 'crimson', label='Shipping costs', bottom= CC + EC)
EC_NH3 = plt.bar(n,EC , color = 'cornflowerblue', label='Export terminal costs', bottom= CC)
CC_NH3 = plt.bar(n,CC , color = 'darkblue', label='Conversion costs')

labels = ['Retrofit pipeline\nBase','Retrofit pipeline\nPolicy', 'New pipeline\nBase', 'New pipeline\nPolicy','LH2\nBase', 'LH2\nPolicy','NH3\nBase','NH3\nPolicy']
#plt.title('Emission breakdown for LH2 shipping ', fontweight='bold')
plt.legend()
plt.xticks(range(8),labels)

#plt.yticks(y, ['10', '100', '1000'])
#plt.yscale("log")
plt.ylabel('[€/kg H2]')

title = '\Transport_cost_comparison_2030'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()

"""## Transport inputs"""

Retrofit_pipeline_costs_off

New_Pipeline_costs_off

LH2_transport_costs

NH3_transport_costs

"""## Minimal transport costs"""

def choose_minimal_transport_costs():

    if year < this_year + AV_pipe_new:
        result = min(TC_LH2, TC_NH3)

    elif year < this_year + AV_pipe_retro:
        result = min(TC_pipe_new, TC_LH2, TC_NH3)
    else:
        result = min(TC_pipe_new, TC_pipe_retro, TC_LH2, TC_NH3)


    return result

def choose_minimal_transport_costs_tech():


    if  year < this_year + AV_pipe_new:
        if min(TC_LH2, TC_NH3) == TC_LH2:
            result = 'LH2'
        else: result = 'NH3'

    elif year < this_year + AV_pipe_retro:
        if min(TC_pipe_new, TC_LH2, TC_NH3) == TC_LH2:
            result = 'LH2'
        elif min(TC_pipe_new, TC_LH2, TC_NH3) == TC_NH3:
            result = 'NH3'
        else:
            result = 'New pipeline'
    else:
        if min(TC_pipe_new, TC_pipe_retro, TC_LH2, TC_NH3) == TC_LH2:
            result = 'LH2'
        elif min(TC_pipe_new, TC_pipe_retro, TC_LH2, TC_NH3) == TC_NH3:
            result = 'NH3'
        elif (TC_pipe_new, TC_pipe_retro, TC_LH2, TC_NH3) == TC_pipe_new:
            result = 'New pipeline'
        else:
            result = 'Retrofit pipeline'


    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Minimal_offshore_transport_costs', 'Technology'])
result.index.name = 'Years'

for year in years:

    # get all costs
    TC_pipe_new = float(New_Pipeline_costs_off.loc[year]['New_Pipeline_costs_off'])

    TC_pipe_retro = float(Retrofit_pipeline_costs_off.loc[year]['Retrofit_pipeline_costs_off'])

    TC_LH2 = float(LH2_transport_costs.loc[year]['LH2_transport_costs'])

    TC_NH3 = float(NH3_transport_costs.loc[year]['NH3_transport_costs'])

    result.Minimal_offshore_transport_costs.loc[year] = choose_minimal_transport_costs()
    result.Technology.loc[year] = choose_minimal_transport_costs_tech()

result

output_file = os.path.join(path_csv, 'LCOT_min_tech.csv')
result.to_csv(output_file, sep = ';')

LCOT_min_tech = result


LCOT_min_plt = LCOT_min_tech.drop(columns='Technology')


"""### Plot transport costs"""

# Plot cost curve for seaborne transport
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(LCOT_min_plt, color='cyan', linestyle='dashed')
plt.title('Cost curve for transport')
plt.legend(['Minimal transport costs in €/kg_H2'])
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

# Plot cost curves of hydrogen transport baseline  policy
fig, ax = plt.subplots(figsize=(10, 4), layout = 'constrained')
plt.subplot(1,2,1)
#baseline
plt.plot(LH2_transport_costs, color='blue', linestyle='solid', label='LH2')
plt.plot(NH3_transport_costs, color='darkorange', linestyle='solid', label='NH3')
plt.plot(NH3_transport_costs_wo_recon, color='darkorange', linestyle='--', label='NH3 w/o cracking')

plt.plot(New_Pipeline_costs_off, color='dodgerblue', linestyle='solid', label='New pipeline')
plt.plot(Retrofit_pipeline_costs_off, color='royalblue', linestyle='solid', label='Retrofit pipeline')
#ax.yaxis.set_major_locator(mtick.LinearLocator(7))
plt.axvline(x=av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=av_retro, color='royalblue', linestyle = '--')
plt.text(av_new, 3, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(av_retro,3,  'Retrofit', horizontalalignment='center', verticalalignment='center')
plt.xlim(2025,2050)
plt.ylim(0,2.9)
plt.grid(True, axis='y')
ax.set_axisbelow(True)
plt.legend()
plt.ylabel('[€/kg H2]')
#policy
plt.subplot(1,2,2)
#import policy cost curves.
LH2_transport_costs_pol = policy.LH2_transport_costs
NH3_transport_costs_pol = policy.NH3_transport_costs
NH3_transport_costs_wo_recon_pol = policy.NH3_transport_costs_wo_recon
New_Pipeline_costs_off = policy.New_Pipeline_costs_off
Retrofit_pipeline_costs_off = policy.Retrofit_pipeline_costs_off
#plot
plt.plot(LH2_transport_costs_pol, color='blue', linestyle ='-', label='LH2')
plt.plot(NH3_transport_costs_pol, color='darkorange', linestyle ='-',label='NH3')
plt.plot(NH3_transport_costs_wo_recon_pol, color='darkorange', linestyle='--', label='NH3 w/o cracking')
plt.plot(New_Pipeline_costs_off, color='dodgerblue', linestyle='solid', label='New pipeline')
plt.plot(Retrofit_pipeline_costs_off, color='royalblue', linestyle='solid', label='Retrofit pipeline')

plt.axvline(x=policy.av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=policy.av_retro, color='royalblue', linestyle = '--')
plt.text(policy.av_new, 3, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(policy.av_retro,3,  'Retrofit', horizontalalignment='center', verticalalignment='center')
plt.xlim(2025,2050)
plt.ylim(0,2.9)
plt.grid(True, axis='y')
ax.set_axisbelow(True)
#plt.title('Hydrogen transport costs ', fontweight='bold')
plt.legend()
plt.ylabel('[€/kg H2]')
#save
title = '\LCOT_time_base_pol'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

## Respective transport emissions
years = np.arange(2025, 2051)
emissions = []
def LCOT_append_emissions():
    for year in years:

        if LCOT_min_tech.loc[year]['Technology'] == 'New pipeline':
            result = Pipeline_emissions.loc[year]['Pipeline_emissions']

        elif LCOT_min_tech.loc[year]['Technology'] == 'Retrofit pipeline':
            result = Pipeline_emissions.loc[year]['Pipeline_emissions']

        elif LCOT_min_tech.loc[year]['Technology'] == 'LH2':
            result =  LH2_transport_emissions.loc[year]['LH2_transport_emissions']

        else:
            result = NH3_transport_emissions.loc[year]['NH3_transport_emissions']

# from g Co2eq/kg H2 to kg CO2eq/kg H2
        emissions.append(result)

    return emissions

LCOT_append_emissions()

LCOT_min_tech_em = LCOT_min_tech.assign(Emissions=emissions)
LCOT_min_tech_em.Emissions = LCOT_min_tech_em.Emissions.divide(1000)

output_file = os.path.join(path_csv, 'LCOT_min_tech_em.csv')
LCOT_min_tech_em.to_csv(output_file, sep =';')

## Minimal supply costs
def calculate_supply_costs():
    result = PC + TC

    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Supply_costs'])
result.index.name = 'Years'

for year in years:
    # get all costs

    PC = float(LCOH_min_tech_em.loc[year]['Minimal_production_costs'])
    TC = float(LCOT_min_tech_em.loc[year]['Minimal_offshore_transport_costs'])


    # calculate costs of specific year
    result.Supply_costs.loc[year] = calculate_supply_costs()

result

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Minimal_supply_costs.csv')
result.to_csv(output_file, sep=',')

Supply_costs = result

## Minimal supply emissions
def calculate_supply_emissions():
    result = PE + TE

    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Supply_emissions'])
result.index.name = 'Years'

for year in years:
    # get all emissions

    PE = float(LCOH_min_tech_em.loc[year]['Emissions'])
    TE = float(LCOT_min_tech_em.loc[year]['Emissions'])


    # calculate emissions of specific year
    result.Supply_emissions.loc[year] = calculate_supply_emissions()

result

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Minimal_supply_costs_emissions.csv')
result.to_csv(output_file, sep=',')

Supply_emissions = result

# Plot cost curve for total supply costs
fig, ax = plt.subplots(figsize=(10, 6), layout = 'constrained')
plt.grid(True, axis='y')
ax.set_axisbelow(True)
plt.plot(result, color='cyan', linestyle='dashed')
plt.title('Supply costs costs in €/kg_H2')
plt.xlabel('Year')
plt.ylabel('Cost')
title = '\Min_supply_cost'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

# Cost breakdown for H2 supply
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
fig.patch.set_visible(False)
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)

PC = (LCOH_min_tech_em.loc[::5]['Minimal_production_costs']).apply(pd.to_numeric)
TC = (LCOT_min_tech_em.loc[::5]['Minimal_offshore_transport_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence


TC_plt = plt.bar(x,TC, width, label='Transport', bottom= PC)
PC_plt = plt.bar(x,PC, width, label='Production')

plt.title('Lowest H2 supply costs from Norway to Germany', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('Cost [€/kg H2]')
plt.xlabel('Years')

title = '\Supply_cost_bars'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

# Min. supply cost stackplot base pol
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#baseline
plt.subplot(1,2,1)
x = np.arange(2025, 2051)
plt.grid(True, axis='y')
ax.set_axisbelow(True)

PC = (LCOH_min_tech_em.loc[::,'Minimal_production_costs']).apply(pd.to_numeric)
TC = (LCOT_min_tech_em.loc[::,'Minimal_offshore_transport_costs']).apply(pd.to_numeric)

labels = ['Production', 'Transport']
colors = ['lightgrey', 'lightsteelblue']
plt.stackplot(x, PC, TC, labels = labels, colors = colors )
#orientation
top = 5.1
plt.axvline(x=av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=av_retro, color='royalblue', linestyle = '--')
plt.text(av_new, top, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(av_retro,top,  'Retrofit', horizontalalignment='center', verticalalignment='center')

plt.ylabel('Supply costs [€/kg H2]')
plt.xlabel('Baseline')
plt.xlim(2025,2050)
plt.ylim(0,5)
plt.legend()

#policy
plt.subplot(1,2,2)
plt.grid(True, axis='y')
ax.set_axisbelow(True)

TC = (policy.LCOT_min_tech_em.loc[::,'Minimal_offshore_transport_costs']).apply(pd.to_numeric)
PC = (policy.LCOH_min_tech_em.loc[::,'Minimal_production_costs']).apply(pd.to_numeric)

labels = ['Production', 'Transport']
colors = ['lightgrey', 'lightsteelblue']
plt.stackplot(x, PC,TC, labels = labels, colors = colors )
#orientation
plt.axvline(x=policy.av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=policy.av_retro, color='royalblue', linestyle = '--')
plt.text(policy.av_new, top, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(policy.av_retro, top,  'Retrofit', horizontalalignment='center', verticalalignment='center')
plt.ylabel('Supply costs [€/kg H2]')
plt.xlabel('Policy')
plt.xlim(2025,2050)
plt.ylim(0,5)
plt.legend()

title = '\Supply_costs_stackplot'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

# Supply emissions base pol
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#baseline
plt.subplot(1,2,1)
x = np.arange(2025, 2051)
plt.grid(True, axis='y')
ax.set_axisbelow(True)

PE = (LCOH_min_tech_em.loc[::,'Emissions']).apply(pd.to_numeric)
TE = (LCOT_min_tech_em.loc[::,'Emissions']).apply(pd.to_numeric)

labels = ['Production', 'Transport']
colors = ['lightgrey', 'lightsteelblue']
plt.stackplot(x, PE, TE, labels = labels, colors = colors )

top = 10.6
ylim = top + 0.1
plt.axvline(x=av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=av_retro, color='royalblue', linestyle = '--')
plt.text(av_new, top, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(av_retro,top,  'Retrofit', horizontalalignment='center', verticalalignment='center')

plt.ylabel('Supply emissions [kg CO2eq/kg H2]')
plt.xlabel('Baseline')
plt.xlim(2025,2050)
plt.ylim(0,ylim)
plt.legend()

#policy
plt.subplot(1,2,2)
plt.grid(True, axis='y')
ax.set_axisbelow(True)

TE = (policy.LCOT_min_tech_em.loc[::,'Emissions']).apply(pd.to_numeric)
PE = (policy.LCOH_min_tech_em.loc[::,'Emissions']).apply(pd.to_numeric)

labels = ['Production', 'Transport']
colors = ['lightgrey', 'lightsteelblue']
plt.stackplot(x, PE,TE, labels = labels, colors = colors )

plt.axvline(x=policy.av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=policy.av_retro, color='royalblue', linestyle = '--')
plt.text(policy.av_new, top, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(policy.av_retro, top, 'Retrofit', horizontalalignment='center', verticalalignment='center')
plt.ylabel('Supply emissions [kg CO2eq/kg H2]')
plt.xlabel('Policy')
plt.xlim(2025,2050)
plt.ylim(0,ylim)
plt.legend()

title = '\Supply_emissions_stackplot'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()


### All together now
## Supply costs

# Min. supply cost stackplot base pol
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#baseline
plt.subplot(1,2,1)
x = np.arange(2025, 2051)
plt.grid(True, axis='y')
ax.set_axisbelow(True)

PC = (LCOH_min_tech_em.loc[::,'Minimal_production_costs']).apply(pd.to_numeric)
TC = (LCOT_min_tech_em.loc[::,'Minimal_offshore_transport_costs']).apply(pd.to_numeric)

labels = ['Production', 'Transport']
colors = ['lightgrey', 'lightsteelblue']
##transport
#plt.plot(LH2_transport_costs, color='blue', linestyle='solid', label='LH2')
#plt.plot(NH3_transport_costs, color='darkorange', linestyle='solid', label='NH3')
#plt.plot(NH3_transport_costs_wo_recon, color='darkorange', linestyle='--', label='NH3 w/o cracking')
#plt.plot(New_Pipeline_costs_off, color='dodgerblue', linestyle='solid', label='New pipeline')
#plt.plot(Retrofit_pipeline_costs_off, color='royalblue', linestyle='solid', label='Retrofit pipeline')
##production
plt.plot(LCOH_green_NOR_Base, color ='green', linestyle ='-', label='LCOH Green')
plt.plot(LCOH_blue_NOR_Base, color ='blue', linestyle ='-', label='LCOH Blue ')
##supply
plt.stackplot(x, PC, TC, labels = labels, colors = colors )

####orientation
top = 5.1
plt.axvline(x=av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=av_retro, color='royalblue', linestyle = '--')
plt.text(av_new, top, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(av_retro,top,  'Retrofit', horizontalalignment='center', verticalalignment='center')

plt.ylabel('Supply costs [€/kg H2]')
plt.xlabel('Baseline')
plt.xlim(2025,2050)
plt.ylim(0,)
plt.legend()

#policy
plt.subplot(1,2,2)
plt.grid(True, axis='y')
ax.set_axisbelow(True)
# production
plt.plot(LCOH_green_NOR_Policy, color ='green', linestyle= '-', label='LCOH Green')
plt.plot(LCOH_blue_NOR_Policy, color ='blue', linestyle= '-', label='LCOH Blue')
#transport
#plt.plot(LH2_transport_costs_pol, color='blue', linestyle ='-', label='LH2')
#plt.plot(NH3_transport_costs_pol, color='darkorange', linestyle ='-',label='NH3')
#plt.plot(NH3_transport_costs_wo_recon_pol, color='darkorange', linestyle='--', label='NH3 w/o cracking')
#plt.plot(New_Pipeline_costs_off, color='dodgerblue', linestyle='solid', label='New pipeline')
#plt.plot(Retrofit_pipeline_costs_off, color='royalblue', linestyle='solid', label='Retrofit pipeline')
#supply
TC = (policy.LCOT_min_tech_em.loc[::,'Minimal_offshore_transport_costs']).apply(pd.to_numeric)
PC = (policy.LCOH_min_tech_em.loc[::,'Minimal_production_costs']).apply(pd.to_numeric)

labels = ['Production', 'Transport']
colors = ['lightgrey', 'lightsteelblue']
plt.stackplot(x, PC,TC, labels = labels, colors = colors )
#orientation
plt.axvline(x=policy.av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=policy.av_retro, color='royalblue', linestyle = '--')
plt.text(policy.av_new, top, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(policy.av_retro, top,  'Retrofit', horizontalalignment='center', verticalalignment='center')
plt.ylabel('Supply costs [€/kg H2]')
plt.xlabel('Policy')
plt.xlim(2025,2050)
plt.ylim(0,)
plt.legend()

title = '\Supply_costs_complex'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

## inverted complex graph
# Min. supply cost stackplot base pol
fig, ax = plt.subplots(figsize=(10,4), layout = 'constrained')
#baseline
plt.subplot(1,2,1)
x = np.arange(2025, 2051)
plt.grid(True, axis='y')
ax.set_axisbelow(True)

PC = (LCOH_min_tech_em.loc[::,'Minimal_production_costs']).apply(pd.to_numeric)
TC = (LCOT_min_tech_em.loc[::,'Minimal_offshore_transport_costs']).apply(pd.to_numeric)

labels = ['Transport', 'Production']
colors = ['lightsteelblue', 'lightgrey']
##transport
plt.plot(LH2_transport_costs, color='blue', linestyle='solid', label='LH2')
plt.plot(NH3_transport_costs, color='darkorange', linestyle='solid', label='NH3')
plt.plot(NH3_transport_costs_wo_recon, color='darkorange', linestyle='--', label='NH3 w/o cracking')
plt.plot(New_Pipeline_costs_off, color='dodgerblue', linestyle='solid', label='New pipeline')
plt.plot(Retrofit_pipeline_costs_off, color='royalblue', linestyle='solid', label='Retrofit pipeline')
##production
#plt.plot(LCOH_green_NOR_Base, color ='green', linestyle ='-', label='LCOH Green')
#plt.plot(LCOH_blue_NOR_Base, color ='blue', linestyle ='-', label='LCOH Blue ')
##supply
plt.stackplot(x, TC, PC, labels = labels, colors = colors )

####orientation
top = 5.1
plt.axvline(x=av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=av_retro, color='royalblue', linestyle = '--')
plt.text(av_new, top, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(av_retro,top,  'Retrofit', horizontalalignment='center', verticalalignment='center')

plt.ylabel('Supply costs [€/kg H2]')
plt.xlabel('Baseline')
plt.xlim(2025,2050)
plt.ylim(0,)
plt.legend()

#policy
plt.subplot(1,2,2)
plt.grid(True, axis='y')
ax.set_axisbelow(True)
# production
#plt.plot(LCOH_green_NOR_Policy, color ='green', linestyle= '-', label='LCOH Green')
#plt.plot(LCOH_blue_NOR_Policy, color ='blue', linestyle= '-', label='LCOH Blue')
#transport
plt.plot(LH2_transport_costs_pol, color='blue', linestyle ='-', label='LH2')
plt.plot(NH3_transport_costs_pol, color='darkorange', linestyle ='-',label='NH3')
plt.plot(NH3_transport_costs_wo_recon_pol, color='darkorange', linestyle='--', label='NH3 w/o cracking')
plt.plot(New_Pipeline_costs_off, color='dodgerblue', linestyle='solid', label='New pipeline')
plt.plot(Retrofit_pipeline_costs_off, color='royalblue', linestyle='solid', label='Retrofit pipeline')
#supply
TC = (policy.LCOT_min_tech_em.loc[::,'Minimal_offshore_transport_costs']).apply(pd.to_numeric)
PC = (policy.LCOH_min_tech_em.loc[::,'Minimal_production_costs']).apply(pd.to_numeric)

labels = [ 'Transport','Production']
colors = ['lightsteelblue', 'lightgrey']
plt.stackplot(x, TC, PC, labels = labels, colors = colors )
#orientation
plt.axvline(x=policy.av_new, color='dodgerblue', linestyle = '--')
plt.axvline(x=policy.av_retro, color='royalblue', linestyle = '--')
plt.text(policy.av_new, top, 'New', horizontalalignment='center', verticalalignment='center')
plt.text(policy.av_retro, top,  'Retrofit', horizontalalignment='center', verticalalignment='center')
plt.ylabel('Supply costs [€/kg H2]')
plt.xlabel('Policy')
plt.xlim(2025,2050)
plt.ylim(0,)
plt.legend()

title = '\Supply_costs_complex_inverted'
plt.savefig(path_plt + title + '.png', transparent=True)
plt.show()

# Final supply df
Supply_base = pd.concat([LCOH_min_tech_em, LCOT_min_tech_em, Supply_costs, Supply_emissions], axis=1)
# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Supply_base.csv')
Supply_base.to_csv(output_file, sep=',')

# Final supply df
Supply_pol = pd.concat([policy.LCOH_min_tech_em, policy.LCOT_min_tech_em, policy.Supply_costs, policy.Supply_emissions], axis=1)
# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Supply_pol.csv')
Supply_pol.to_csv(output_file, sep=',')

'''
# export result to excel
path_supply = r'\\dena.de\Daten\Home\Reuter\Desktop\H2_pathways_repo\data\Results\Supply_base.xlsx'
writer = pd.ExcelWriter(path_supply, engine='openpyxl', mode='a', if_sheet_exists='replace')
book = load_workbook(path_supply)
#writer.book = book
#writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=True, startcol=ws.max_column)
book.save(path_supply)
book.close()
'''