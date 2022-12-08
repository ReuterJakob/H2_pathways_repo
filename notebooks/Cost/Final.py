"""# Production cost"""

lowest_green = ((lcoh_green.loc['Norway_Onshore_2_low_temp_optimistic', 2025:2050]).mul(0.89))

lowest_green = pd.DataFrame(lowest_green)
lowest_green.index.name = 'Years'


# Interest rate (WACC) in %
i = float(tea_blue.loc['Discount rate [%]']['NGR with CCS'])
i

# Economic lifetime of the plant in years
l_ngr = float(tea_blue.loc['Lifetime [Years]']['NGR with CCS'])
l_ngr

opex_share = float(tea_blue.loc['Opex [% of Capex]']['NGR with CCS'])
opex_share

# Calculate the amortisation factor alpha
alpha_ngr = (i * (1 + i) ** l_ngr) / (1 + i) ** (l_ngr - 1)
alpha_ngr
#round(alpha, 2)

CF = float(tea_blue.loc['Availability [%]']['NGR with CCS'])
CF

# Plant efficiency in %
n = float(tea_blue.loc['Efficiency [%]']['NGR with CCS'])
n

# P_ccs is the cost for transporting and storing CO2 in $/t_co2
P_ccs = float(tea_blue.loc['CO2 transport and storage cost [€/t CO2]']['NGR with CCS'])
P_ccs

# LHV of hydrogen is 33.33 kWh/kg
LHV_h2 = float(tea_blue.loc['LHV H2 [kWh/kg]']['NGR with CCS'])
LHV_h2

"""## Calc. blue LCOH

Definition of the cost calculation function for LCOH from NGR. Time relevant variables = capex, opex, P_ng, P_co2
LHV H2 [kWh/kg]
capex_y  [€/kW]
opex_y [€/kW/a]
CF [%]
P_ng_y [€/MWh]
Q_ce [kgCO2/kgH2]
Q_ue [kgCO2/kgH2]
P_ccs [€/t CO2]
P_co2_y [€/t CO2]
"""

def calculate_lcoh_ngr(alpha, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs, Q_ue_y, P_co2_y, LHV_h2):

    result = float(LHV_h2 * ((alpha * capex_y + opex_y) / (CF * 8760) + P_ng_y / 1000 * n) + (Q_ce_y * P_ccs + Q_ue_y * P_co2_y) / 1000)

    return result

# Calculation of LCOH from NGR for every year from 2025 to 2050.
years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['LCOH_blue'])
result.index.name = 'Years'


for year in years:

    # get capex, opex, natural gas price and CO2 price of specific year
    capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
    opex_y = capex_y * opex_share
    Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2]'][year])
    Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - Norway'][year])
    P_ng_y = float(prices.loc['Gas prices in Germany [€_2020/MWh]'][year])
    P_co2_y = float(prices.loc['EU ETS [€_2020/t_CO2]'][year])

    # calculate lcoe of specific year
    result.LCOH_blue.loc[year] = calculate_lcoh_ngr(alpha=alpha_ngr, capex_y=capex_y, opex_y=opex_y, CF=CF, P_ng_y=P_ng_y,
                                                    n=n, Q_ce_y=Q_ce_y, P_ccs=P_ccs, Q_ue_y=Q_ue_y, P_co2_y=P_co2_y, LHV_h2=LHV_h2)

result

LCOH_BLUE = result
LCOH_BLUE

# Plot cost curve of hydrogen production from NGR with CCS
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'blue', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Cost curve for blue hydrogen production', fontweight='bold')
#plt.xlabel('Year')
plt.ylabel('[€/kg H2]')
plt.show()

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'LCOH_NGR.csv')
result.to_csv(output_file, sep = ';')

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Results']

result.to_excel(writer, sheet_name='Reference Results', index=True, startcol=ws.max_column)
book.save('/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx')
book.close()



"""## Green LCOH"""

# LCOH from RES in €_2019/kg H2 from EWI 2019 [$_2019/kg H2]
green_off1_opt_lowtemp = (lcoh_green.loc['Norway_Offshore_1_low_temp_optimistic', 2025:2050]).mul(0.89)
green_off1_opt_lowtemp

# Plot cost curve of hydrogen production from RES
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(green_off1_opt_lowtemp, color = 'green', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Cost curve for green hydrogen production', fontweight='bold')
plt.xlabel('Year')
plt.ylabel('Cost [€/kg H2]')
plt.show()

"""## Green and blue

"""

lowest_green = (lcoh_green.loc['Norway_Onshore_1_low_temp_optimistic', 2025:2050]).mul(0.89)

# Plot cost curves of hydrogen production from NGR with CCS and RES
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'blue', linestyle = 'solid', label='Blue hydrogen')
plt.plot(lowest_green, color = 'green', linestyle = 'solid', label='Green hydrogen')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Cost curves for green and blue hydrogen production', fontweight='bold')
ax.legend()
plt.xlabel('Year')
#plt.ylim(0)
plt.ylabel('Cost [€/kg H2]')
plt.show()

"""## Sensitivity analysis"""

year = 2050

capex_y = float(tea_blue.loc['Capex [€/kW]'][year])
opex_y = capex_y * opex_share
Q_ce_y = float(GHG.loc['Captured emissions [kg CO2/kg H2]'][year])
Q_ue_y = float(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - Norway'][year])
P_ng_y = float(prices.loc['Gas prices in Germany [€_2020/MWh]'][year])
P_co2_y = float(prices.loc['EU ETS [€_2020/t_CO2]'][year])

gas_prices = np.arange(10,100,5)

sensitivity = []

def calculate_lcoh_ngr_sensi(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs, Q_ue_y, P_co2_y, LHV_h2):

    for x in range(21):
        factor = x/10

        result = float(LHV_h2 * ((alpha_ngr * capex_y + opex_y) / (CF * 8760)+ P_ng_y / 1000 * n) + (Q_ce_y * P_ccs * factor + Q_ue_y * P_co2_y ) / 1000)

        sensitivity.append(result)


    return sensitivity

calculate_lcoh_ngr_sensi(alpha_ngr, capex_y, opex_y, CF, P_ng_y, n, Q_ce_y, P_ccs, Q_ue_y, P_co2_y, LHV_h2)

lcoh_ngr_sensi_P_CCS= pd.DataFrame(sensitivity, index= range(21), columns=['LCOH_NGR [€/kg H2]'])
lcoh_ngr_sensi_P_CCS.index.name = 'CCS_price_change in %'

lcoh_ngr_sensi_P_CCS

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lcoh_ngr_sensi_P_CCS_.csv')
lcoh_ngr_sensi_P_CCS.to_csv(output_file, sep = ';')

import matplotlib.ticker as mtick

fig, ax = plt.subplots(figsize=(5,4))
#plt.subplot(1,2,1)


plt.plot(lcoh_ngr_sensi, color='blue', linestyle='solid', label = 'Gas price')
plt.plot(lcoh_ngr_sensi_P_Co2, color='dodgerblue',linestyle='-', label = 'CO2 price')
plt.plot(lcoh_ngr_sensi_P_CCS, color='royalblue',linestyle='-', label = 'CCS cost')
plt.grid(True, axis='y')
plt.grid(True, axis='x')
ax.set_axisbelow(True)
ax.xaxis.set_major_formatter(mtick.PercentFormatter(10, decimals=None))
plt.locator_params(axis='x', nbins=5)
plt.ylabel('[€/kg H2]')
plt.xlabel('Change')
plt.legend()



title = 'LCOH_ngr_sensi'
plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()

plt.ylabel('[€/kg H2]')
plt.xlabel('Gas price change')
#plt.legend()

#Capture rates
plt.subplot(1,2,2)
plt.plot(capture_sensi, color='blue', linestyle='solid')
plt.grid(True, axis='y')
ax.set_axisbelow(True)
plt.ylabel('[g CO2eq/MJ H2]')
plt.xlabel('System capture rate'           )
#plt.xticks(np.arange(0.55,0.95, 0.1), ['55%', '60%', '65%', '70%', '75%', '80%', '85%', '90%', '95%'] )
plt.xticks(np.arange(0.55,1, 0.1), ['55%',  '65%',  '75%',  '85%',  '95%'] )



"""# Transport cost"""

#Definition of variables for pipeline transport costs
tea_pipe = pd.read_excel(path, sheet_name='Pipeline Transport', decimal=',', index_col=0)
tea_pipe

tra_d = pd.read_excel(path, sheet_name='Transport Distances', decimal=',')
tra_d

el_price = pd.read_excel(path, sheet_name= 'Commodity Prices', index_col=0, decimal=',')
el_price

# WACC for all transport investments (pipelines, terminals, liquefaction plants)
i_tra = float(tea_pipe.loc['Discount rate [%]']['Parameter'])
i_tra

i = i_tra

"""## Pipeline"""

# Economic lifetime for pipelines
l_pipe = float(tea_pipe.loc['Lifetime [Years]']['Parameter'])
l_pipe
# Amortisation factor for onshore pipelines
alpha = (i_tra * (1 + i_tra) ** l_pipe) / (((1 + i_tra) ** l_pipe) - 1)
alpha
# Utilisation of the pipeline in %
pipe_use = float(tea_pipe.loc['Pipeline load factor [%]']['Parameter'])
pipe_use
# Pipeline Opex [€/a as % of Capex]
pipe_opex_share = float(tea_pipe.loc['Pipeline Opex [€/a as % of Capex]']['Parameter'])
pipe_opex_share
# Compression Opex [€/a as % of Capex]
comp_opex_share = float(tea_pipe.loc['Compressor Opex [€/a as % of Capex]']['Parameter'])
comp_opex_share
# New pipeline

"""### New offshore pipeline"""

# Pipeline capex offshore new [€/kg/1000km]
capex_pipe_new_off_EHB = float(
    tea_pipe.loc['Medium - New Offshore (EHB 2022) Capex Pipeline [€/kg/1000km]']['Parameter'])
capex_pipe_new_off_EHB
# Compression capex offshore new [€/kg/1000km]
capex_comp_new_off_EHB = float(
    tea_pipe.loc['Medium - New Offshore (EHB 2022) Capex Compression [€/kg/1000km]']['Parameter'])
capex_comp_new_off_EHB
capex_new_off = capex_pipe_new_off_EHB + capex_comp_new_off_EHB
capex_new_off
opex_new_off = capex_pipe_new_off_EHB * pipe_opex_share + capex_comp_new_off_EHB * comp_opex_share
opex_new_off
# Onshore pipeline distance between two countries [km]
d_on = float(tra_d[(tra_d['Origin_Country'] == 'Norway') & (tra_d['Destination_Country'] == 'Germany')][
                 'onshore distance (km)'].values)
d_on
#not used here
off_factor = float(tea_pipe.loc['Offshore Capex cost factor ']['Parameter'])
off_factor
# Offshore pipeline distance between two countries [km]
d_sea = float(tra_d[(tra_d['Origin_Country'] == 'Norway') & (tra_d['Destination_Country'] == 'Germany')][
                  'offshore distance (km)'].values)
d_sea

"""### Electricity cost of compression"""

# Compression capacity in [MWel/1000km]
capa_comp = float(tea_pipe.loc['Medium - Compression capacity @ 48inch [MW_el/1000km]']['Parameter'])
capa_comp
# Load factor in hours/day
lf_comp = float(tea_pipe.loc['Assumed load factor for compressor electricity consumption [h/a]']['Parameter'])
lf_comp
#Capacity at full load factor [kg H2/a]
capa_pipe = float(tea_pipe.loc['Capacity at full load factor [kg H2/a]']['Parameter'])
capa_pipe
# Electricity cost for compression [€/kg/1000km]
#comp_el_cost = capa_comp * pipe_use * 8760 * p_el_y / capa_pipe


"""## Calc. offshore new pipeline cost"""

# Costs for new pipeline transport [€/kg]
#alpha, capex_pipe_new_on_EHB, capex_comp_new_on_EHB, d_on, d_off, capex_pipe_new_off_EHB, capex_comp_new_off_EHB, capa_comp, pipe_use, capa_pipe
def calculate_off_pipe_new():

    result = ((alpha * capex_new_off / pipe_use + opex_new_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea / 1000

    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['New_Pipeline_costs_off'])
result.index.name = 'Years'

for year in years:
    p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

    result.New_Pipeline_costs_off.loc[year] = calculate_off_pipe_new()

result

New_Pipeline_costs_off = result

"""## Sensitivity

### Electricity price sensi
"""

year = 2030
p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

P_el_sensi = np.arange(0,121,10)
sensitivity = []

def calculate_off_pipe_retrofit_sensi():
    for p_el_y in P_el_sensi:

        result = (((alpha * capex_retrofit_off / pipe_use + opex_retrofit_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea) / 1000

        sensitivity.append(result)

    return sensitivity

calculate_off_pipe_retrofit_sensi()

retro_pipe_sensi_P_el = pd.DataFrame(sensitivity, P_el_sensi, columns=['Retrofit pipeline transport costs [€/kg H2]'])
retro_pipe_sensi_P_el.index.name = 'Electricity Price [€/MWh]'
retro_pipe_sensi_P_el

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'retro_pipe_sensi_P_el.csv')
retro_pipe_sensi_P_el.to_csv(output_file, sep=';')

"""### Transport distance sensi"""

transport_distance = np.arange(0,10001, 500)
sensitivity = []
def calculate_off_pipe_retrofit_sensi_distance():
    for d_off in transport_distance:

        result = (((alpha * capex_retrofit_off / pipe_use + opex_retrofit_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_off) / 1000

        sensitivity.append(result)

    return sensitivity

calculate_off_pipe_retrofit_sensi_distance()

retro_pipe_sensi_distance = pd.DataFrame(sensitivity, transport_distance, columns=['Retrofit pipeline transport costs [€/kg H2]'])
retro_pipe_sensi_distance.index.name = 'Transport distance in km'
retro_pipe_sensi_distance

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'retro_pipe_sensi_distance.csv')
retro_pipe_sensi_distance.to_csv(output_file, sep=';')

"""## Retrofit offshore pipeline cost"""

#Retrofit  capex in [€/kg/100km]
capex_pipe_retrofit_off_EHB = float(tea_pipe.loc['Medium - retrofit Offshore (EHB 2022) Capex Pipeline [€/kg/1000km]']['Parameter'])
capex_pipe_retrofit_off_EHB

#Compression capex in [€/kg/100km]
capex_comp_retrofit_off_EHB = float(tea_pipe.loc['Medium - retrofit Offshore (EHB 2022) Capex Compression [€/kg/1000km]']['Parameter'])
capex_comp_retrofit_off_EHB

capex_retrofit_off = capex_pipe_retrofit_off_EHB + capex_comp_retrofit_off_EHB
capex_retrofit_off

opex_retrofit_off = capex_pipe_retrofit_off_EHB * pipe_opex_share + capex_comp_retrofit_off_EHB * comp_opex_share
opex_retrofit_off

"""### Calc. offshore retrofit pipeline costs"""

# Costs for new pipeline transport [€/kg]
#alpha, capex_pipe_new_on_EHB, capex_comp_new_on_EHB, d_on, d_off, capex_pipe_new_off_EHB, capex_comp_new_off_EHB, capa_comp, pipe_use, capa_pipe
def calculate_off_pipe_retrofit():
    result = (((alpha * capex_retrofit_off / pipe_use + opex_retrofit_off) + (
                capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea) / 1000
    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Retrofit_pipeline_costs_off'])
result.index.name = 'Years'

for year in years:
    p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

    result.Retrofit_pipeline_costs_off.loc[year] = calculate_off_pipe_retrofit()

result

Retrofit_pipeline_costs_off = result

"""### Sensitivity
#### Electricity price sensi
"""

year = 2030
#p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

P_el_sensi = np.arange(0,121,10)
sensitivity = []

def calculate_off_pipe_new_sensi_P_el():
    for p_el_y in P_el_sensi:


        result = ((alpha * capex_new_off / pipe_use + opex_new_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_sea / 1000

        sensitivity.append(result)

    return sensitivity

calculate_off_pipe_new_sensi_P_el()

new_pipe_sensi_P_el = pd.DataFrame(sensitivity, P_el_sensi, columns=['New pipeline transport costs [€/kg H2]'])
new_pipe_sensi_P_el.index.name = 'Electricity price [€/MWh]'
new_pipe_sensi_P_el

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'new_pipe_sensi_P_el.csv')
new_pipe_sensi_P_el.to_csv(output_file, sep=';')

"""#### Transport distance sensi"""

year = 2030
p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

transport_distance = np.arange(0,10001, 500)
sensitivity = []
def calculate_off_pipe_retrofit_sensi_distance():
    for d_off in transport_distance:

        result = ((alpha * capex_new_off / pipe_use + opex_new_off)
                  + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_off / 1000

        #result = (((alpha * capex_retrofit_off / pipe_use + opex_retrofit_off) + (capa_comp * pipe_use * 8760 * p_el_y / capa_pipe)) * d_off) / 1000
        sensitivity.append(result)

    return sensitivity

calculate_off_pipe_retrofit_sensi_distance()

new_pipe_sensi_distance = pd.DataFrame(sensitivity, transport_distance, columns=['New pipeline transport costs [€/kg H2]'])
new_pipe_sensi_distance.index.name = 'Transport distance in km'
new_pipe_sensi_distance

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'new_pipe_sensi_distance.csv')
new_pipe_sensi_distance.to_csv(output_file, sep=';')

"""## LH2 transport"""

tea_lh2 = pd.read_excel(path, sheet_name='LH2', decimal=',', index_col=0)
tea_lh2

"""### Liquefaction"""

# WACC for liquefaction plant in %
i_liq = float(tea_lh2.loc['Liquefaction - Discount rate [%]']['LH2'])
i_liq

# Economic lifetime for the liquefaction plant [years]
l_liq = float(tea_lh2.loc['Liquefaction - Lifetime [Years]']['LH2'])
l_liq

# Amortisation factor alpha for the liquefaction plant
alpha_liq = (i_liq * (1 + i_liq) ** l_liq) /(((1 + i_liq) ** l_liq) - 1)
alpha_liq

opex_liq_share = float(tea_lh2.loc['Liquefaction - Opex opt. [% of Capex]']['LH2'])
opex_liq_share

"""Calc. liquefaction cost"""

# Definition of the cost calculation function for liquefaction costs. Time relevant variables = capex [€/tpa], opex [€/tpa], electricity use [kwh/kg H2], electricity price [$/MWh]
def calculate_liq_costs(alpha_liq, capex_liq_y, opex_liq_y, el_liq_y, p_el_y):

    result = float((alpha_liq * capex_liq_y/1000 + opex_liq_y/1000) + el_liq_y * p_el_y*0.89/1000)

    return result

# Calculation of liquefaction costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['LH2_Liquefaction_costs'])
result.index.name = 'Years'

for year in years:

    # get capex, opex, natural gas price and CO2 price of specific year
    capex_liq_y = float(tea_lh2.loc['Liquefaction - Capex opt. [€/t/a]'][year])
    opex_liq_y = capex_liq_y * opex_liq_share
    el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
    p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

    # calculate costs of specific year
    result.LH2_Liquefaction_costs.loc[year] = calculate_liq_costs(alpha_liq=alpha_liq, capex_liq_y=capex_liq_y, opex_liq_y=opex_liq_y, el_liq_y=el_liq_y, p_el_y=p_el_y)

result

LH2_Liquefaction_costs = result

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'LH2_Liquefaction_costs.csv')
result.to_csv(output_file, sep = ';')

"""Plot liquefaction costs"""

# Plot cost curve for liquefaction
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'cyan', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Cost curve for liquefaction', fontweight = 'bold')
plt.legend(['Liquefaction costs in €/kg_H2'])
plt.xlabel('Year')
plt.ylabel('Cost [€/kg_H2]')
plt.show()

"""### Export terminal"""

# Lifetime of import and export terminals
l_ter = float(tea_lh2.loc['Export Terminal - Technical lifetime [Years]']['LH2'])

# Amortisation factor for the export terminal
alpha_et = (i_tra * (1 + i_tra) ** l_ter) / (((1 + i_tra) ** l_ter) - 1)

# Average storage time in the export terminal tanks in days
t_et = float(tea_lh2.loc['Export Terminal - Storage length per load [Days]']['LH2'])

# Electricity consumption for the export terminal [kWh/kg_h2]
el_et = float(tea_lh2.loc['Export Terminal - Electricity use [kWh/kgH2]']['LH2'])
el_et

# Boil-off hydrogen in [%/day]
bog_et = float(tea_lh2.loc['Export Terminal - Boil off rate [%/day]']['LH2'])
bog_et

# Average storage time in the export terminal tanks in days
t_et = float(tea_lh2.loc['Export Terminal - Storage length per load [Days]']['LH2'])
t_et

el_reliq = float(tea_lh2.loc['Export Terminal - Energy use to reliquefy opt. [kWh/kg H2/day]']['LH2'])
el_reliq

opex_et_share =  float(tea_lh2.loc['Export Terminal - Annual OPEX [% of CAPEX]']['LH2'])
opex_et_share

"""Calc. export terminal cost

Boil-off and transfer gas are re-liquefied @

Time relevant variables = capex [$/tpa], opex [$/tpa], electricity use [kwh/kg H2], electricity price [$/MWh] storage time in days
"""

# Definition of the cost calculation function for the Export terminal.
def calculate_export_terminal_costs(alpha_et, capex_et_y, opex_et_y, el_et, p_el_y, el_reliq):
    result = (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * t_et) * p_el_y* 0.89/1000

    return result

#Calculation of export terminal costs [$/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Export_terminal_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, xx
    capex_et_y = float(tea_lh2.loc['Export Terminal - CAPEX/tank [€/t/a]'][year])
    opex_et_y = capex_et_y * opex_et_share
    p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])


    # calculate costs of specific year
    result.LH2_Export_terminal_costs.loc[year] = calculate_export_terminal_costs(alpha_et=alpha_et, capex_et_y=capex_et_y,
                                                              opex_et_y=opex_et_y, el_et=el_et, p_el_y=p_el_y, el_reliq=el_reliq)

result

LH2_Export_terminal_costs =result

# export result to excel
from openpyxl import load_workbook
path = '/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx'

writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save('/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx')
book.close()

# Create csv file from results dataframe
path = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path, 'LH2_Export_terminal_costs.csv')
result.to_csv(output_file, sep=';')

"""Plot export terminal costs"""

# Plot cost curve for export terminal costs
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'red', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='red', linestyle='dashed')
plt.title('Export terminal costs over time', fontweight='bold', fontsize='large')
plt.legend(['Export terminal costs in €/kg_H2'])
plt.xlabel('Year', fontsize='large')
plt.ylabel('Cost [€/kg_H2]', fontsize='large')
plt.show()

"""### Shipping"""

# Technical lifetime of ships in years
l_ship = float(tea_lh2.loc['Shipping - Technical Lifetime [Years]']['LH2'])
l_ship

# Amortisation factor for shipping
alpha_ship = (i_tra * (1 + i_tra) ** l_ship) / (((1 + i_tra) ** l_ship) - 1)
alpha_ship

# Distance between countries via ship in [km]
d_sea = float(tra_d[(tra_d['Origin_Country'] == 'Norway') & (tra_d['Destination_Country'] == 'Germany')]['sea distance (km)'].values)
d_sea

# Ship speed in [km/h]
v_ship = float(tea_lh2.loc['Shipping - Ship speed [km/h]']['LH2'])
v_ship

# Berthing time (time for unloading and loading in a harbour) in [h]
h_ship = float(tea_lh2.loc['Shipping - Berthing time [hours]']['LH2'])
h_ship

# Boil-off when shipping in [%/day] to [%/hour]
bog_ship = float(tea_lh2.loc['Shipping - Boil off opt. [%/day]']['LH2'])/24
bog_ship

# Fuel consumption of a ship in [kg_h2/t/km]
f_ship = float(tea_lh2.loc['Shipping - Fuel use [kg H2/t/km]']['LH2'])/1000
f_ship

opex_ship_share = float(tea_lh2.loc['Shipping - Annual OPEX [% of CAPEX]']['LH2'])
opex_ship_share

"""Fuel cost calculation"""

# Cost of transported hydrogen in [$/kg_h2] in year y
LCOH_min = pd.read_csv(r'\\dena.de\Daten\Home\Reuter\Desktop\H2_pathways_repo\data\processed\Minimal_production_costs.csv', delimiter=';',index_col=0, decimal=',')
LCOH_min

# Cost of transported hydrogen carrier in [€/kg_h2] in year y
H2_costs = (LCOH_min.loc[::]['Minimal_production_costs']).apply(pd.to_numeric)
Con_costs = (LH2_Liquefaction_costs.loc[::]['LH2_Liquefaction_costs']).apply(pd.to_numeric)
ET_costs =  (LH2_Export_terminal_costs.loc[::]['LH2_Export_terminal_costs']).apply(pd.to_numeric)

# €/kg_H2
LH2_cargo_cost = (H2_costs + Con_costs + ET_costs).to_frame('LH2_cargo_cost')
LH2_cargo_cost

"""Calc. shipping costs @ 10.000km"""

d_sea = 100

"""Definition of the cost calculation function for the maritime shipping.
 [%]
d_sea [km]
v_ship [km/h]
h_ship [h] berthing time
bog_ship [%/day]
f_ship [kg/kg/km]] Fuel use
Time relevant variables:
capex [€/kg/a]
opex [€/kg/a]
lcoh in [€/kg_h2]

Fuel assumptions: outward journey - boil-off is used as fuel.
Return journey: Residual H2 is used as fuel.
"""

def calculate_ship_costs():

    result = (alpha_ship * capex_ship_y + opex_ship_y) / (8760/(2*(d_sea/v_ship + h_ship)))\
             /(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea))\
             + (bog_ship * d_sea/v_ship + f_ship * d_sea) * lcoh

    return result

# Calculation of shipping costs [€/kg_h2] from 2025 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Shipping_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, xx
    capex_ship_y = float(tea_lh2.loc['Shipping - Capex/Ship opt. [€/t/a]'][year])/1000
    opex_ship_y = capex_ship_y * opex_ship_share
    lcoh = float(LH2_cargo_cost.loc[year]['LH2_cargo_cost'])

    # calculate costs of specific year
    result.LH2_Shipping_costs.loc[year] = calculate_ship_costs()

result

LH2_Shipping_costs = result

# export result to excel
from openpyxl import load_workbook
path = '/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx'

writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save('/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx')
book.close()

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'LH2_Shipping_costs.csv')
result.to_csv(output_file, sep=';')

"""Plot shipping costs"""

# Plot cost curve of hydrogen production from NGR with CCS
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'green', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Shipping costs over time in €/kg H2', fontweight='bold')
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

"""### Import terminal"""

# Definition of variables for the import terminal costs

# Amortisation factor for the import terminal
alpha_it = alpha_et

# Electricity consumption in kWh/kg H2
el_it = float(tea_lh2.loc['Import Terminal - Electricity use [kWh/kg H2]']['LH2'])
el_it

# Import terminal boil-off in [%/day]
bog_it = float(tea_lh2.loc['Import Terminal - Boil-off [%/day]']['LH2'])
bog_it

# Average storage time in the export terminal tanks in days
t_it = float(tea_lh2.loc['Import Terminal - Storage length per load [days]']['LH2'])
t_it

opex_it_share =  float(tea_lh2.loc['Import Terminal - Annual OPEX [% of Opex]']['LH2'])
opex_it_share

# Import terminal costs in [$/kg_h2]:
# Capex for the import terminal in [$/tpa]
# Opex for the import terminal in [$/tpa]
# Electricity consumption for the import terminal in [kWh/kg_h2]
# Price for electricity in importing country [$/MWh]
# Import terminal boil-off in [%/h]
# Berthing time in import terminal in [h]
# Cost of transported hydrogen in [$/kg_h2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Import_terminal_costs'])
result.index.name = 'Years'

def calculate_import_terminal_costs():
    result = (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * t_it) * p_el_y* 0.89/1000

    return result

for year in years:
    # get capex, opex, xx
    capex_it_y = float(tea_lh2.loc['Import Terminal - CAPEX [€/t/a]'][year])
    opex_it_y = capex_it_y * opex_it_share
    p_el_y = float(el_price.loc['Electricity prices in Germany [€_2020/MWh]'][year])


    # calculate costs of specific year
    result.LH2_Import_terminal_costs.loc[year] = calculate_import_terminal_costs()

result

LH2_Import_terminal_costs = result

# export result to excel
from openpyxl import load_workbook
path = '/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx'

writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save('/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx')
book.close()

# Create csv file from results dataframe
path = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path, 'LH2_Import_terminal_costs.csv')
result.to_csv(output_file, sep=';')

"""Plot import terminal costs"""

fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'red', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='red', linestyle='dashed')
plt.title('Import terminal costs over time', fontweight='bold', fontsize='large')
plt.legend(['Import terminal costs in €/kg_H2'])
plt.xlabel('Year', fontsize='large')
plt.ylabel('Cost [€/kg_H2]', fontsize='large')
plt.show()

"""### Reconversion"""

# Economic lifetime for the reconversion plant [years]
l_recon = float(tea_lh2.loc['Reconversion - Lifetime [Years]']['LH2'])
l_recon

# Amortisation factor alpha for the reconversion plant
alpha_recon =  (i_liq * (1 + i_liq) ** l_recon) /(((1 + i_liq) ** l_recon) - 1)

# Electricty consumption for the conversion in [kWh/kg_h2]
el_recon = float(tea_lh2.loc['Import Terminal - Electricity use [kWh/kg H2]']['LH2'])
el_recon

opex_recon_share = float(tea_lh2.loc['Reconversion - Opex opt. [% of Capex]']['LH2'])
opex_recon_share

def calculate_recon_costs(alpha_recon, capex_recon_y, opex_recon_y, el_recon_y, p_el_y):
    result = float(
        (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + el_recon_y * p_el_y / 1000 * 0.89)
    return result

# Calculation of conversion costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Reconversion_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_recon_y = float(tea_lh2.loc['Reconversion - Capex opt. [€/t/a]'][year])
    opex_recon_y = capex_recon_y * opex_recon_share
    el_recon_y = float(tea_lh2.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
    p_el_y = float(el_price.loc['Electricity prices in Germany [€_2020/MWh]'][year])

    # calculate costs of specific year
    result.LH2_Reconversion_costs.loc[year] = calculate_recon_costs(alpha_recon=alpha_recon, capex_recon_y=capex_recon_y,
                                                                opex_recon_y=opex_recon_y, el_recon_y=el_recon_y,
                                                                p_el_y=p_el_y)
result

LH2_Reconversion_costs = result

# export result to excel
from openpyxl import load_workbook
path = '/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx'

writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save('/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx')
book.close()

# Create csv file from results dataframe
path = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path, 'LH2_Reconversion_costs.csv')
result.to_csv(output_file, sep=';')

"""## Total LH2 transport costs"""

liq_costs = LH2_Liquefaction_costs
et_costs = LH2_Export_terminal_costs
ship_costs = LH2_Shipping_costs
it_costs = LH2_Import_terminal_costs
recon_costs = LH2_Reconversion_costs

it_costs

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_transport_costs'])
result.index.name = 'Years'

# Calculate total transport costs from liquefaction costs (LC), export terminal costs (EC), Shipping costs (SC), import terminal costs (IC)
def calculate_seaborne_transport_costs():
    result = LC + EC + SC + IC + RC

    return result

for year in years:
    # get all costs
    LC = float(liq_costs.loc[year]['LH2_Liquefaction_costs'])
    EC = float(et_costs.loc[ year]['LH2_Export_terminal_costs'])
    SC = float(ship_costs.loc[year]['LH2_Shipping_costs'])
    IC = float(it_costs.loc[year]['LH2_Import_terminal_costs'])
    RC = float(recon_costs.loc[year]['LH2_Reconversion_costs'])
    # calculate costs of specific year
    result.LH2_transport_costs.loc[year] = calculate_seaborne_transport_costs()

result

LH2_transport_costs =result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'LH2_transport_costs.csv')
result.to_csv(output_file, sep=';')

# Plot cost curve for seaborne transport
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'green', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Cost curve for LH2 shipping [€/kg H2]', fontweight='bold')
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

# Cost breakdown for LH2 shipping
fig, ax = plt.subplots(figsize=(10,6))
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)
LC = (liq_costs.loc[::5]['LH2_Liquefaction_costs']).apply(pd.to_numeric)
EC = (et_costs.loc[::5]['LH2_Export_terminal_costs']).apply(pd.to_numeric)
SC = (ship_costs.loc[::5]['LH2_Shipping_costs']).apply(pd.to_numeric)
IC = (it_costs.loc[::5]['LH2_Import_terminal_costs']).apply(pd.to_numeric)
RC = (recon_costs.loc[::5]['LH2_Reconversion_costs']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

RC_plt = plt.bar(x,RC, width, label='Regasification costs', bottom= LC + EC + SC + IC)
IC_plt = plt.bar(x,IC, width, label='Import terminal costs', bottom= LC + EC + SC)
SC_plt = plt.bar(x,SC, width, label='Shipping costs', bottom= LC + EC)
EC_plt = plt.bar(x,EC, width, label='Export terminal costs', bottom= LC)
LC_plt = plt.bar(x,LC, width, label='Liquefaction costs')

plt.title('Cost breakdown for LH2 shipping', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('Transport cost [€/kg H2]')
plt.xlabel('Years')
plt.show()

"""### Sensitivity"""

year = 2030
el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
capex_liq_y = float(tea_lh2.loc['Liquefaction - Capex opt. [€/t/a]'][year])
opex_liq_y = capex_liq_y * opex_liq_share
el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
capex_et_y = float(tea_lh2.loc['Export Terminal - CAPEX/tank [€/t/a]'][year])
opex_et_y = capex_et_y * opex_et_share
capex_it_y = float(tea_lh2.loc['Import Terminal - CAPEX [€/t/a]'][year])
opex_it_y = capex_it_y * opex_it_share
capex_ship_y = float(tea_lh2.loc['Shipping - Capex/Ship opt. [€/t/a]'][year])/1000
opex_ship_y = capex_ship_y * opex_ship_share
lcoh = float(LH2_cargo_cost.loc[year]['LH2_cargo_cost'])
capex_recon_y = float(tea_lh2.loc['Reconversion - Capex opt. [€/t/a]'][year])
opex_recon_y = capex_recon_y * opex_recon_share
el_recon_y = float(tea_lh2.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])

p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

"""#### Electricity price sensi"""

P_el_sensi = np.arange(0,121,10)
sensitivity = []

def lh2_transport_sensi_P_el(p_el_y, alpha_liq, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_liq_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y):

    for p_el_y in P_el_sensi:

        result = \
        ((alpha_liq * capex_liq_y/1000 + opex_liq_y/1000) + el_liq_y * p_el_y/1000)\
        + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * t_et) * p_el_y/1000 \
        + (alpha_ship * capex_ship_y + opex_ship_y) / (8760/(2*(d_sea/v_ship + h_ship)))/(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea)) + (bog_ship * d_sea/v_ship + f_ship * d_sea) * lcoh\
        + (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * t_it) * p_el_y/1000 \
        + (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + el_recon_y * p_el_y / 1000                 # Reconversion

        sensitivity.append(result)

    return sensitivity

lh2_transport_sensi_P_el(p_el_y, alpha_liq, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_liq_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y)

lh2_transport_sensi_P_el= pd.DataFrame(sensitivity, P_el_sensi, columns=['LH2 transport costs [€/kg H2]'])
lh2_transport_sensi_P_el.index.name = 'Electricity price [€/MWh]'
lh2_transport_sensi_P_el

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lh2_transport_sensi_P_el.csv')
lh2_transport_sensi_P_el.to_csv(output_file, sep = ';')

"""#### Transport distance sensi"""

transport_distance = np.arange(0,10001, 500)
sensitivity = []

def lh2_transport_d(p_el_y, alpha_liq, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_liq_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y):

    for d_sea in transport_distance:

        result = \
        ((alpha_liq * capex_liq_y/1000 + opex_liq_y/1000) + el_liq_y * p_el_y/1000)\
        + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * t_et) * p_el_y/1000\
        + (alpha_ship * capex_ship_y + opex_ship_y) / (8760/(2*(d_sea/v_ship + h_ship)))/(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea)) + (bog_ship * d_sea/v_ship + f_ship * d_sea) * lcoh\
        + (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * t_it) * p_el_y/1000\
        + (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + el_recon_y * p_el_y / 1000\

        sensitivity.append(result)

    return sensitivity

lh2_transport_d(p_el_y, alpha_liq, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_liq_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y)

lh2_transport_sensi_distance= pd.DataFrame(sensitivity, transport_distance, columns=['LH2 transport costs [€/kg H2]'])
lh2_transport_sensi_distance.index.name = 'Transport distance in km'
lh2_transport_sensi_distance

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lh2_transport_sensi_distance.csv')
lh2_transport_sensi_distance.to_csv(output_file, sep = ';')

"""## NH3 Transport"""

tea_lnh3 = pd.read_excel(path, sheet_name='LNH3', decimal=',', index_col=0)
tea_lnh3

"""### Conversion"""

# Economic lifetime for the conversion plant [years]
l_con = float(tea_lnh3.loc['Conversion - Lifetime [Years]']['NH3'])
l_con

# Amortisation factor alpha for the conversion plant
alpha_con = (i * (1 + i) ** l_con) /(((1 + i) ** l_con) - 1)
alpha_con

'not used'# Conversion efficiency [% of LHV]
#eff_con = float(tea_lnh3.loc['Conversion - Efficiency opt. [% of LHV]'][2025])

opex_con_share = (tea_lnh3.loc['Conversion - Opex opt. [% of Capex]']['NH3'])
opex_con_share

"""Calc. conversion cost

Definition of the cost calculation function for conversion costs. Time relevant variables =
capex [€/tpa]
opex [€/tpa]
electricity use [kwh/kg H2]
electricity price [$/MWh]
"""

def calculate_con_costs():

    result = float((alpha_con * capex_con_y/1000 + opex_con_y/1000) + el_con_y * p_el_y/1000)

    return result

# Calculation of conversion costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['LNH3_Conversion_costs'])
result.index.name = 'Years'

for year in years:

    # get capex, opex, natural gas price and CO2 price of specific year
    capex_con_y = float(tea_lnh3.loc['Conversion - Capex opt. [€/t/a]'][year])
    opex_con_y = capex_con_y * opex_con_share
    el_con_y = float(tea_lnh3.loc['Conversion - Electricity consumption opt. [kWh/kgH2]'][year])
    p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

    # calculate costs of specific year
    result.LNH3_Conversion_costs.loc[year] = calculate_con_costs()

result

LNH3_Conversion_costs = result

# export result to excel
path = '/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx'
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

# Create csv file from results dataframe
path=r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv,'LNH3_Conversion_costs.csv')
result.to_csv(output_file, sep = ';')

"""Plot conversion costs"""

# Plot cost curve for conversion of H2 to NH3
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'cyan', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Cost curve for conversion of H2 to NH3 [€/kg]', fontweight = 'bold')
plt.legend(['Conversion costs'])
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

"""### Export terminal"""

# Lifetime of import and export terminals
l_ter = float(tea_lnh3.loc['Export Terminal - Technical lifetime [Years]']['NH3'])
# Amortisation factor for the export terminal
alpha_et = (i * (1 + i) ** l_ter) / (((1 + i) ** l_ter) - 1)

# Electricity consumption for the export terminal [kWh/kg_h2]
el_et = float(tea_lnh3.loc['Export Terminal - Electricity use [kWh/kgH2]']['NH3'])
el_et

# Boil-off hydrogen in [%/day]
bog_et = float(tea_lnh3.loc['Export Terminal - Boil off rate [%/day]']['NH3'])
bog_et

# Average storage time in the export terminal tanks in days
t_et = float(tea_lnh3.loc['Export Terminal - Storage length per load [Days]']['NH3'])
t_et

#Lower heating value of Nh3 [kWh/kg]
NH3_lhv =  float(tea_lnh3.loc['Properties - LHV [kWh/kg]']['NH3'])
NH3_lhv

#Lower heating value of H2 [kWh/kg]
H2_lhv = 33.33

# transformed to kWh/kg h2/day
el_reliq = float(tea_lnh3.loc['Export Terminal - Energy use to reliquefy opt. [kWh/kg NH3/day]']['NH3'])
el_reliq

opex_et_share

"""Calc. export terminal cost

Time relevant variables =
capex [€/tpa]
opex [€/tpa]
el_et electricity use [kwh/kg H2]
electricity price [$/MWh]
constant =
boil off in [%/d]
storage time in days
Energy use to reliquefy opt. [kWh/kg NH3/day]
"""

# Definition of the cost calculation function for the Export terminal.

def calculate_export_terminal_costs():
    result = (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * (NH3_lhv/H2_lhv) * t_et) * p_el_y/1000

    return result

#Calculation of export terminal costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LNH3_Export_terminal_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, xx
    capex_et_y = float(tea_lnh3.loc['Export Terminal - CAPEX for storage tanks [€/t/a]'][year])
    opex_et_y = float(tea_lnh3.loc['Export Terminal - Annual OPEX [€/t/a]'][year])
    p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

    # calculate costs of specific year
    result.LNH3_Export_terminal_costs.loc[year] = calculate_export_terminal_costs()

result

LNH3_Export_terminal_costs = result

# export result to excel
path = '/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx'
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

# Create csv file from results dataframe
path_csv= r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'LNH3_Export_terminal_costs.csv')
result.to_csv(output_file, sep=';')

"""Plot export terminal costs"""

# Plot cost curve for export terminal costs
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'red', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='red', linestyle='dashed')
plt.title('Export terminal costs over time', fontweight='bold', fontsize='large')
plt.legend(['Export terminal costs in €/kg_H2'])
plt.xlabel('Year', fontsize='large')
plt.ylabel('Cost [€/kg_H2]', fontsize='large')
plt.show()

"""### Shipping"""

# Technical lifetime of ships in years
l_ship = float(tea_lnh3.loc['Shipping - Technical Lifetime [Years]']['NH3'])
l_ship

# Amortisation factor for shipping
alpha_ship = (i * (1 + i) ** l_ship) / (((1 + i) ** l_ship) - 1)
alpha_ship

# Distance between countries via ship in [km]
d_sea = float(tra_d[(tra_d['Origin_Country'] == 'Norway') & (tra_d['Destination_Country'] == 'Germany')]['sea distance (km)'].values)
d_sea

# Ship speed in [km/h]
v_ship = float(tea_lnh3.loc['Shipping - Ship speed [km/h]']['NH3'])
v_ship

# Berthing time (time for unloading and loading in a harbour) in [h]
h_ship = float(tea_lnh3.loc['Shipping - Berthing time [hours]']['NH3'])
h_ship

#Lower heating value of Nh3[ kWh / kg]
NH3_lhv = float(tea_lnh3.loc['Properties - LHV [kWh/kg]']['NH3'])
NH3_lhv
#Lower heating value of H2 [kWh/kg]
H2_lhv = 33.33

# Boil-off when shipping in [%_NH3/day] to [%/hour]
bog_ship = float(tea_lnh3.loc['Shipping - Boil off opt. [%/day]']['NH3'])/24 * (NH3_lhv/H2_lhv)
bog_ship

# Fuel consumption of a ship in [kg_NH3/t_NH3/km]
#f_ship = float(tea_lnh3.loc['Shipping - Fuel use [kg NH3/t/km]']['NH3'])/1000 * (NH3_lhv/H2_lhv)

opex_ship_share = (tea_lnh3.loc['Shipping - Annual OPEX [% of CAPEX]']['NH3'])
opex_ship_share

# Cost of transported hydrogen in [$/kg_h2] in year y

# Cost of transported hydrogen carrier in [€/kg_h2] in year y
H2_costs = (LCOH_min.loc[::]['Minimal_production_costs']).apply(pd.to_numeric)
Con_costs = (LNH3_Conversion_costs.loc[::]['LNH3_Conversion_costs']).apply(pd.to_numeric)
ET_costs = (LNH3_Export_terminal_costs.loc[::]['LNH3_Export_terminal_costs']).apply(pd.to_numeric)

# €/kg_H2 transported in the form of NH3
LNH3_cargo_cost = (H2_costs + Con_costs + ET_costs).to_frame('LNH3_cargo_cost')
LNH3_cargo_cost

#d_sea = 10000

"""Definition of the cost calculation function for the maritime shipping.
Time relevant variables:
capex [€/t/a]
opex [€/t/a]
lcoh in [€/kg_h2]
static:
alpha [%]
d_sea [km]
v_ship [km/h]
h_ship [h] berthing time
bog_ship [%_H2/day]
f_ship [kg_H2/kg_NH3/km]] Fuel use

"""

def calculate_ship_costs(alpha_ship, capex_ship_y, opex_ship_y, d_sea, v_ship, h_ship, bog_ship, f_ship, LNH3_costs_y):
    result = (alpha_ship * capex_ship_y + opex_ship_y)/(8760/(2*(d_sea/v_ship + h_ship)))\
             /(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea))\
             + (bog_ship * d_sea/v_ship + f_ship * d_sea) * LNH3_costs_y

    return result

# Calculation of shipping costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LNH3_Shipping_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, xx
    capex_ship_y = float(tea_lnh3.loc['Shipping - Capex/Ship opt. [€/t/a]'][year])/1000
    opex_ship_y = capex_ship_y * opex_ship_share
    f_ship = float(tea_lnh3.loc['Shipping - Fuel use [kg NH3/t/km]'][year])/1000 * (NH3_lhv/H2_lhv)
    LNH3_costs_y = float(LNH3_cargo_cost.loc[year])

    # calculate costs of specific year
    result.LNH3_Shipping_costs.loc[year] = calculate_ship_costs(alpha_ship=alpha_ship, capex_ship_y=capex_ship_y, opex_ship_y=opex_ship_y, d_sea=d_sea, v_ship=v_ship, h_ship=h_ship, bog_ship=bog_ship, f_ship=f_ship, LNH3_costs_y=LNH3_costs_y)

result

LNH3_Shipping_costs = result

# export result to excel
path = '/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx'
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

# Create csv file from results dataframe
path_csv= r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'LNH3_Shipping_costs.csv')
result.to_csv(output_file, sep=';')

"""Plot shipping costs"""

# Plot cost curve of hydrogen production from NGR with CCS
fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'green', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.title('Shipping costs over time in €/kg H2', fontweight='bold')
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

"""### Import terminal"""

# Definition of variables for the import terminal costs

# Amortisation factor for the import terminal
alpha_it = alpha_et

# Electricity consumption in kWh/kg H2
el_it = float(tea_lnh3.loc['Import Terminal - Electricity use [kWh/kg H2]']['NH3'])
el_it

# Import terminal boil-off in [%/day]
bog_it = float(tea_lnh3.loc['Import Terminal - Boil-off [%/day]']['NH3'])
bog_it

# Average storage time in the export terminal tanks in days
t_it = float(tea_lnh3.loc['Import Terminal - Storage length per load [days]']['NH3'])
t_it

opex_it_share = opex_et_share

"""Time relevant variables =
capex [€/tpa]
opex [€/tpa]
electricity use [kwh/kg H2]
electricity price [$/MWh] in Germany
lcoh in [€/kg_h2]
constant =
boil off in [%/d]
storage time in [days]


"""

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LNH3_Import_terminal_costs'])
result.index.name = 'Years'

def calculate_import_terminal_costs(alpha_it, capex_it_y, opex_it_y, el_it, p_el_y, t_it):
    result = (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * p_el_y /1000
    return result

for year in years:
    # get capex, opex, xx
    capex_it_y = float(tea_lnh3.loc['Import Terminal - CAPEX for storage tanks [€/t/a]'][year])
    opex_it_y = float(tea_lnh3.loc['Import Terminal - Annual OPEX [€/t/a]'][year])
    p_el_y = float(el_price.loc['Electricity prices in Germany [€_2020/MWh]'][year])


    # calculate costs of specific year
    result.LNH3_Import_terminal_costs.loc[year] = calculate_import_terminal_costs(alpha_it=alpha_it, capex_it_y=capex_it_y,
                                                              opex_it_y=opex_it_y, el_it=el_it, p_el_y=p_el_y, t_it=t_it)

result

LNH3_Import_terminal_costs = result

# export result to excel
path = '/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx'
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

# Create csv file from results dataframe
path_csv= r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'LNH3_Import_terminal_costs.csv')
result.to_csv(output_file, sep=';')

"""
Plot import terminal costs"""

fig, ax = plt.subplots(figsize=(10,6))
plt.plot(result, color = 'red', linestyle = 'solid')
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='red', linestyle='dashed')
plt.title('Import terminal costs over time', fontweight='bold', fontsize='large')
plt.legend(['Import terminal costs in €/kg_H2'])
plt.xlabel('Year', fontsize='large')
plt.ylabel('Cost [€/kg_H2]', fontsize='large')
plt.show()

"""


### Reconversion"""

# Economic lifetime for the reconversion plant [years]
l_recon = float(tea_lnh3.loc['Reconversion - Lifetime [Years]']['NH3'])
l_recon

# Amortisation factor alpha for the conversion plant
alpha_recon = (i * (1 + i) ** l_recon) / (((1 + i) ** l_recon) - 1)
alpha_recon

opex_recon_share = float(tea_lnh3.loc['Reconversion - Annual OPEX opt. [% of CAPEX]']['NH3'])

"""Calc. reconversion cost
Definition of the cost calculation function for conversion costs.Time relevant variables =
capex[€/tpa]
opex[€/tpa]
electricity use[kwh/kg H2]
electricity price[$/ MWh]

Energy demand for heat in cracking is provided locally
"""

def calculate_recon_costs(alpha_recon, capex_recon_y, opex_recon_y, el_recon_y, p_el_y):
    result = float((alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + (heat_recon_y + el_recon_y) * p_el_y / 1000)
    return result

# Calculation of conversion costs [€/kg_h2] from 2030 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LNH3_Reconversion_costs'])
result.index.name = 'Years'

for year in years:
    # get capex, opex, natural gas price and CO2 price of specific year
    capex_recon_y = float(tea_lnh3.loc['Reconversion - Capex opt. [€/t/a]'][year])
    opex_recon_y = capex_recon_y * opex_recon_share
    el_recon_y = float(tea_lnh3.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
    heat_recon_y = float(tea_lnh3.loc['Reconversion - Heat consumption opt. [kWh/kg H2]'][year])
    p_el_y = float(el_price.loc['Electricity prices in Germany [€_2020/MWh]'][year])

    # calculate costs of specific year
    result.LNH3_Reconversion_costs.loc[year] = calculate_recon_costs(alpha_recon=alpha_recon, capex_recon_y=capex_recon_y,
                                                            opex_recon_y=opex_recon_y, el_recon_y=el_recon_y, p_el_y=p_el_y)
result

LNH3_Reconversion_costs = result

"""### Sensitivity"""

year = 2030

capex_con_y = float(tea_lnh3.loc['Conversion - Capex opt. [€/t/a]'][year])
opex_con_y = capex_con_y * opex_con_share
el_con_y = float(tea_lnh3.loc['Conversion - Electricity consumption opt. [kWh/kgH2]'][year])

capex_et_y = float(tea_lnh3.loc['Export Terminal - CAPEX for storage tanks [€/t/a]'][year])
opex_et_y = capex_et_y * opex_et_share

capex_it_y = float(tea_lnh3.loc['Import Terminal - CAPEX for storage tanks [€/t/a]'][year])
opex_it_y = capex_it_y * opex_it_share

capex_ship_y = float(tea_lnh3.loc['Shipping - Capex/Ship opt. [€/t/a]'][year])/1000
opex_ship_y = capex_ship_y * opex_ship_share
f_ship = float(tea_lnh3.loc['Shipping - Fuel use [kg NH3/t/km]'][year])/1000 * (NH3_lhv/H2_lhv)
LNH3_costs_y = float(LNH3_cargo_cost.loc[year])

capex_recon_y = float(tea_lnh3.loc['Reconversion - Capex opt. [€/t/a]'][year])
opex_recon_y = capex_recon_y * opex_recon_share
el_recon_y = float(tea_lnh3.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
heat_recon_y = float(tea_lnh3.loc['Reconversion - Heat consumption opt. [kWh/kg H2]'][year])

p_el_y = float(el_price.loc['Electricity prices in Norway [€_2021/MWh]'][year])

"""#### Electricity price sensi"""

P_el_sensi = np.arange(0,121,10)
sensitivity = []

def nh3_transport_sensi_P_el(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y):

    for p_el_y in P_el_sensi:

        result = \
        (alpha_con * capex_con_y/1000 + opex_con_y/1000) + el_con_y * p_el_y/1000\
        + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * (NH3_lhv/H2_lhv) * t_et) * p_el_y /1000 \
        + (alpha_ship * capex_ship_y + opex_ship_y)/(8760/(2*(d_sea/v_ship + h_ship)))/(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea))+ (bog_ship * d_sea/v_ship + f_ship * d_sea) * LNH3_costs_y\
        + (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * p_el_y /1000 \
        + (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + (heat_recon_y + el_recon_y) * p_el_y / 1000

        sensitivity.append(result)

    return sensitivity

nh3_transport_sensi_P_el(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y)

lnh3_transport_sensi_P_el= pd.DataFrame(sensitivity, P_el_sensi, columns=['LNH3 transport costs [€/kg H2]'])
lnh3_transport_sensi_P_el.index.name = 'Electricity price [€/MWh]'
lnh3_transport_sensi_P_el

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lnh3_transport_sensi_P_el.csv')
lnh3_transport_sensi_P_el.to_csv(output_file, sep = ';')

"""#### Transport distance sensi"""

transport_distance = np.arange(0,10001, 500)
sensitivity = []
def nh3_transport_sensi_distance(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y):

    for d_sea in transport_distance:

        result = \
        (alpha_con * capex_con_y/1000 + opex_con_y/1000) + el_con_y * p_el_y/1000 \
        + (alpha_et * capex_et_y/1000 + opex_et_y/1000) + (el_et + el_reliq * (NH3_lhv/H2_lhv) * t_et) * p_el_y /1000 \
        + (alpha_ship * capex_ship_y + opex_ship_y)/(8760/(2*(d_sea/v_ship + h_ship)))/(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea))+ (bog_ship * d_sea/v_ship + f_ship * d_sea) * LNH3_costs_y\
        + (alpha_it * capex_it_y/1000 + opex_it_y/1000) + (el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * p_el_y /1000\
        + (alpha_recon * capex_recon_y / 1000 + opex_recon_y / 1000) + (heat_recon_y + el_recon_y) * p_el_y / 1000

        sensitivity.append(result)

    return sensitivity

nh3_transport_sensi_distance(p_el_y, alpha_con, alpha_et, alpha_ship, alpha_it, alpha_recon,capex_con_y, capex_et_y,capex_ship_y, capex_it_y,  bog_ship, capex_recon_y, d_sea ,v_ship, f_ship, el_et ,el_it ,t_it ,el_recon_y)

lnh3_transport_sensi_distance= pd.DataFrame(sensitivity, transport_distance, columns=['LNH3 transport costs [€/kg H2]'])
lnh3_transport_sensi_distance.index.name = 'Transport distance in km'
lnh3_transport_sensi_distance

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'lnh3_transport_sensi_distance.csv')
lnh3_transport_sensi_distance.to_csv(output_file, sep = ';')

"""## Sensitivity Plots

### Electricity price sensi
"""

lh2_transport_sensi_P_el

fig, ax = plt.subplots(figsize=(10,4))
#plt.subplot(1,2,1)

plt.plot(lh2_transport_sensi_P_el, color='blue', linestyle='solid', label = 'LH2')
plt.plot(lnh3_transport_sensi_P_el, color='darkorange', linestyle='solid', label = 'NH3')
plt.plot(new_pipe_sensi_P_el, color='dodgerblue',linestyle='-', label = 'New pipeline')
plt.plot(retro_pipe_sensi_P_el, color='royalblue',linestyle='-', label = 'Retrofit pipeline')
plt.grid(True, axis='y')
#plt.grid(True, axis='x')
ax.set_axisbelow(True)
#ax.xaxis.set_major_formatter(mtick.PercentFormatter(10, decimals=None))
plt.locator_params(axis='x', nbins=8)
plt.locator_params(axis='y', nbins=7)

plt.axvline(x=57, color='grey', linestyle = '--')
plt.axvline(x=103, color='grey', linestyle = '--')
plt.text(57,2.8, 'Norway 2030', horizontalalignment='center', verticalalignment='center')
plt.text(103,2.8, 'Germany 2023', horizontalalignment='center', verticalalignment='center')


plt.ylabel('[€/kg H2]')
#plt.ylim(top = 1.5)
plt.xlabel('Electricity price [€/MWh]')
plt.legend(loc = 'upper left')



title = 'LH2_P_el_sensi'
#plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()

"""### Transport distance sensi"""

fig, ax = plt.subplots(figsize=(10,4))
#plt.subplot(1,2,1)


plt.plot(lh2_transport_sensi_distance, color='blue', linestyle='-', label = 'LH2')
plt.plot(lnh3_transport_sensi_distance, color='darkorange', linestyle='-', label = 'NH3')
plt.plot(new_pipe_sensi_distance, color='red',linestyle='-', label = 'New pipeline')
plt.plot(retro_pipe_sensi_distance, color='royalblue',linestyle='-', label = 'Retrofit pipeline')
plt.grid(True, axis='y')
#plt.grid(True, axis='x')
ax.set_axisbelow(True)
plt.locator_params(axis='x', nbins=12)
plt.ylabel('[€/kg H2]')
plt.ylim(top = 3)
plt.xlabel('Transport Distance in km')
plt.legend()



title = 'LH2_EF_sensi'
#plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()

"""# Transport emissions

## Pipeline transport

Compression capacity in [MWel/1000km]
Utilisation of the pipeline in %
8760 h/a
Capacity at full load factor [kg H2/a]
Transport distance in [km]
"""

# Calc. energy use for compression [kWh/kg_H2]
en_comp = capa_comp * 1000 * pipe_use * 8760 / capa_pipe / 1000 * (d_on + d_sea)
en_comp

"""Energy use per kg H2 delivered times emission factor for energy use.
en_comp = Total energy use for compression [kWh/kg H2]
EF_y_n = Emission factor of used energy in year y and country n [g CO2eq/kWh]
"""

def Pipeline_emissions():

    result = en_comp * EF_y_n
    return result

"""Calc. pipeline emissions over time

Using grid electricity for compression.
Compressor station for subsea pipelines in exporting country.
Emission factor for grid electricity in year y and country n [g CO2eq/kWh]
Energy use for compression in [kWh/kg H2]
"""

years = np.arange(2025,2051)
result = pd.DataFrame(index=years, columns=['Pipeline_emissions'])
result.index.name = 'Years'

for year in years:

    EF_y_n = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Norway'][year])

    result.Pipeline_emissions.loc[year] = Pipeline_emissions()

result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

Pipeline_emissions = result

"""### Sensitivity"""

#EF_y_n = 118

EF_sensi = np.arange(0,301,20)
sensitivity = []
def Pipeline_emissions_sensi_EF():

    for EF_y_n in EF_sensi:

        result = en_comp * EF_y_n

        sensitivity.append(result)

    return result

Pipeline_emissions_sensi_EF()

Pipeline_emissions_sensi_EF = pd.DataFrame(sensitivity, EF_sensi, columns=['Pipeline transport emission [g CO2eq/kg H2]'])
Pipeline_emissions_sensi_EF.index.name = 'Electricity emission factor [g CO2eq/kWh]'
Pipeline_emissions_sensi_EF

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'pipe_EF_sensi.csv')
Pipeline_emissions_sensi_EF.to_csv(output_file, sep=';')

"""Plot emission breakdown for pipeline transport

## LH2 transport

Liquefaction energy use

Electricity use for liquefaction in [kWh/kg H2]
Emission factor for grid electricity in year y and country n [g CO2eq/kWh]

### Liquefaction
"""

def Conversion_emissions():
    result = el_liq_y * EF_y_n
    return result

# Calc. liquefaction emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Conversion_emissions'])
for year in years:
    el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
    EF_y_n = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Norway'][year])

    result.LH2_Conversion_emissions.loc[year] = Conversion_emissions()

result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

LH2_Conversion_emissions = result

output_file = os.path.join(path_csv,'LH2_Conversion_emissions.csv')
result.to_csv(output_file, sep = ';')

"""### Export terminal"""

# Electricity consumption for the export terminal [kWh/kg_h2]
el_et = float(tea_lh2.loc['Export Terminal - Electricity use [kWh/kgH2]']['LH2'])
el_et

el_reliq = float(tea_lh2.loc['Export Terminal - Energy use to reliquefy opt. [kWh/kg H2/day]']['LH2'])
el_reliq

# Average storage time in the export terminal tanks in days
t_et = float(tea_lh2.loc['Export Terminal - Storage length per load [Days]']['LH2'])
t_et

def ET_emissions():
    result = (el_et + el_reliq * t_et) * EF_y_n
    return result

# Calc. liquefaction emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Export_terminal_emissions'])
result.index.name = 'Years'

for year in years:
    EF_y_n = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Norway'][year])

    result.LH2_Export_terminal_emissions.loc[year] = ET_emissions()

result

LH2_Export_terminal_emissions = result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

output_file = os.path.join(path_csv,'LH2_Export_terminal_emissions.csv')
result.to_csv(output_file, sep = ';')



"""### Shipping

Calc. emission factor for cargo/fuel

Cost of transported hydrogen in [$/kg_h2] in year y
"""

# Commented out IPython magic to ensure Python compatibility.
# %store -r min_prod_cost
min_prod_cost

# Emissions of transported hydrogen carrier in [g CO2eq/kg_H2] in year y
H2_emissions = (min_prod_cost.loc[::]['Emissions']).apply(pd.to_numeric) * 1000
Con_emissions = (LH2_Conversion_emissions.loc[::]['LH2_Conversion_emissions']).apply(pd.to_numeric)
ET_emissions = (LH2_Export_terminal_emissions.loc[::]['LH2_Export_terminal_emissions']).apply(pd.to_numeric)

# Emission of cargo in [g CO2eq/kg_H2]
LH2_cargo_emissions = (H2_emissions + Con_emissions + ET_emissions).to_frame('LH2_cargo_emissions')
LH2_cargo_emissions

output_file = os.path.join(path_csv,'LH2_cargo_emissions.csv')
result.to_csv(output_file, sep = ';')

# Boil-off when shipping in [%/day] to [%/hour]
bog_ship = float(tea_lh2.loc['Shipping - Boil off opt. [%/day]']['LH2']) / 24
bog_ship

# Fuel consumption of a ship in [kg_h2/t/km]
f_ship = float(tea_lh2.loc['Shipping - Fuel use [kg H2/t/km]']['LH2']) / 1000
f_ship

# Distance between countries via ship in [km]
d_sea = float(tra_d[(tra_d['Origin_Country'] == 'Norway') & (tra_d['Destination_Country'] == 'Germany')][
                  'sea distance (km)'].values)
d_sea

# Ship speed in [km/h]
v_ship = float(tea_lh2.loc['Shipping - Ship speed [km/h]']['LH2'])
v_ship

#d_sea = 10000

"""Calc. LH2 shipping emissions"""

def calculate_ship_emissions():
    result = 1/(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea)) \
             + (bog_ship * d_sea/v_ship + f_ship * d_sea) * LH2_cargo_ghg

    return result

# Calculation of shipping emissions [g CO2eq/kg_h2] from 2025 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Shipping_emissions'])
result.index.name = 'Years'

for year in years:

    LH2_cargo_ghg = float(LH2_cargo_emissions.loc[year]['LH2_cargo_emissions'])

    # calculate emissions of specific year
    result.LH2_Shipping_emissions.loc[year] = calculate_ship_emissions()

result

LH2_Shipping_emissions = result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

output_file = os.path.join(path_csv,'LH2_Shipping_emissions.csv')
result.to_csv(output_file, sep = ';')



"""### Import terminal

"""

# Electricity consumption for the import terminal [kWh/kg_h2]
el_it = float(tea_lh2.loc['Import Terminal - Electricity use [kWh/kg H2]']['LH2'])
el_it

# Electricity consumption to reliquefy BOG [kWh/kg_h2]
el_reliq = float(tea_lh2.loc['Import Terminal - Energy use to reliquefy opt. [kWh/kg H2/day]']['LH2'])
el_reliq

# Average storage time in the import terminal tanks in days
t_it = float(tea_lh2.loc['Import Terminal - Storage length per load [days]']['LH2'])
t_it

def IT_emissions():
    result = (el_it + el_reliq * t_it) * EF_y_n
    return result


# Calc. liquefaction emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Import_terminal_emissions'])
result.index.name = 'Years'

for year in years:
    EF_y_n = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Germany'][year])

    result.LH2_Import_terminal_emissions.loc[year] = IT_emissions()

result

LH2_Import_terminal_emissions = result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

output_file = os.path.join(path_csv, 'LH2_Import_terminal_emissions.csv')
result.to_csv(output_file, sep=';')



"""### Reconversion

"""

def calculate_recon_emissions():
    result = el_recon_y * EF_y_n
    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_Reconversion_emissions'])
result.index.name = 'Years'

for year in years:

    el_recon_y = float(tea_lh2.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
    EF_y_n = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Germany'][year])

    # calculate emissions of specific year
    result.LH2_Reconversion_emissions.loc[year] = calculate_recon_emissions()
result

LH2_Reconversion_emissions = result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

path_csv = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'LH2_Reconversion_emissions.csv')
result.to_csv(output_file, sep=';')



"""### Total LH2 transport emissions

con_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Conversion_emissions.csv', delimiter=';', decimal=',', index_col= 0)
et_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Export_terminal_emissions.csv', delimiter=';', decimal=',', index_col= 0)
ship_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Shipping_emissions.csv', delimiter=';', decimal=',', index_col= 0)
it_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Import_terminal_emissions.csv', delimiter=';', decimal=',', index_col= 0)
recon_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LH2_Reconversion_emissions.csv', delimiter=';', decimal=',', index_col= 0)
"""

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['LH2_transport_emissions'])
result.index.name = 'Years'

# [g CO2eq/kg H2]
def calculate_LH2_transport_emissions():
    result = CE + EE + SE + IE + RE

    return result

for year in years:
    # get all emissions
    CE = float(LH2_Conversion_emissions.loc[year]['LH2_Conversion_emissions'])
    EE = float(LH2_Export_terminal_emissions.loc[year]['LH2_Export_terminal_emissions'])
    SE = float(LH2_Shipping_emissions.loc[year]['LH2_Shipping_emissions'])
    IE = float(LH2_Import_terminal_emissions.loc[year]['LH2_Import_terminal_emissions'])
    RE = float(LH2_Reconversion_emissions.loc[year]['LH2_Reconversion_emissions'])
    # calculate costs of specific year
    result.LH2_transport_emissions.loc[year] = calculate_LH2_transport_emissions()

result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

# Emission breakdown for LH2 seaborne transport
fig, ax = plt.subplots(figsize=(10,6))
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)
CE = (LH2_Conversion_emissions.loc[::5]['LH2_Conversion_emissions']).apply(pd.to_numeric)
EE = (LH2_Export_terminal_emissions.loc[::5]['LH2_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (LH2_Shipping_emissions.loc[::5]['LH2_Shipping_emissions']).apply(pd.to_numeric)
IE = (LH2_Import_terminal_emissions.loc[::5]['LH2_Import_terminal_emissions']).apply(pd.to_numeric)
RE = (LH2_Reconversion_emissions.loc[::5]['LH2_Reconversion_emissions']).apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence

RE_plt = plt.bar(x,RE, width, label='Regasification emissions', bottom= CE + EE + SE + IE)
IE_plt = plt.bar(x,IE, width, label='Import terminal emissions', bottom= CE + EE + SE)
SE_plt = plt.bar(x,SE, width, label='Shipping emissions', bottom= CE + EE)
EE_plt = plt.bar(x,EE, width, label='Export terminal emissions', bottom= CE)
CE_plt = plt.bar(x,CE, width, label='Liquefaction emissions')

#plt.title('Emission breakdown for LH2 shipping ', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('[g CO2eq/kg H2]')
plt.xlabel('Years')

title = 'LH2_emissions_2'
plt.savefig(path_plt + title + '.png', transparent = True)

plt.show()

"""@ 10.000 km shipping distance. Large contributers to total emissions are emissions factors for grid electricity in import/exporting countries (assumed zero from 2045).

### Sensitivity
"""

year = 2030
el_liq_y = float(tea_lh2.loc['Liquefaction - Electricity consumption opt. [kWh/kgH2]'][year])
EF_y_n = 118# float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Norway'][year])
LH2_cargo_ghg = float(LH2_cargo_emissions.loc[year]['LH2_cargo_emissions'])
el_recon_y = float(tea_lh2.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
#EF_y_G = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Germany'][year])

EF_sensi = np.arange(0,301,20)
sensitivity = []
def lh2_transport_sensi_EF(el_liq_y, EF_y_n, el_et, el_reliq, t_et, bog_ship, d_sea ,v_ship, f_ship, LH2_cargo_ghg ,el_it ,t_it ,el_recon_y):

    for EF_y_n in EF_sensi:

        result = \
        (el_liq_y * EF_y_n )\
        + ((el_et + el_reliq * t_et) * EF_y_n )\
        + (1/(1-(bog_ship * d_sea/v_ship) - (f_ship  * d_sea)) + (bog_ship * d_sea/v_ship + f_ship * d_sea) * LH2_cargo_ghg)\
        + ((el_it + el_reliq * t_it) * EF_y_n )\
        + (el_recon_y * EF_y_n )

        sensitivity.append(result)

    return sensitivity

lh2_transport_sensi_EF(el_liq_y, EF_y_n, el_et, el_reliq, t_et, bog_ship, d_sea ,v_ship, f_ship, LH2_cargo_ghg ,el_it ,t_it ,el_recon_y)

lh2_transport_sensi_EF= pd.DataFrame(sensitivity, EF_sensi, columns=['LH2 transport emission [g CO2eq/kg H2]'])
lh2_transport_sensi_EF.index.name = 'Electricity emission [g CO2eq/kWh]'
lh2_transport_sensi_EF

# Create csv file from results dataframe
output_file = os.path.join(path_csv,'LH2_EF_sensi.csv')
lh2_transport_sensi_EF.to_csv(output_file, sep = ';')

import matplotlib.ticker as mtick

fig, ax = plt.subplots(figsize=(5,4))
#plt.subplot(1,2,1)


plt.plot(lh2_transport_sensi, color='blue', linestyle='solid', label = 'Emission factor')
#plt.plot(lcoh_ngr_sensi_P_Co2, color='dodgerblue',linestyle='-', label = 'CO2 price')
#plt.plot(lcoh_ngr_sensi_P_CCS, color='royalblue',linestyle='-', label = 'CCS cost')
plt.grid(True, axis='y')
#plt.grid(True, axis='x')
ax.set_axisbelow(True)
ax.xaxis.set_major_formatter(mtick.PercentFormatter(10, decimals=None))
plt.locator_params(axis='x', nbins=5)
plt.ylabel('[g CO2eq/kg H2]')
plt.xlabel('Change')
plt.legend()



title = 'LH2_EF_sensi'
#plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()

"""## NH3 transport

### Conversion emissions
"""

def Conversion_emissions():
    result = el_con_y * EF_y_n
    return result

# Calc. liquefaction emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Conversion_emissions'])
for year in years:
    el_con_y = float(tea_lnh3.loc['Conversion - Electricity consumption opt. [kWh/kgH2]'][year])
    EF_y_n = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Norway'][year])

    result.NH3_Conversion_emissions.loc[year] = Conversion_emissions()

result

# export result to excel
path = '/Users/jakob/PycharmProjects/H2_pathways_repo/data/raw/H2_supply_route_assessment.xlsx'
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

path_csv = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'NH3_Conversion_emissions.csv')
result.to_csv(output_file, sep=';')

NH3_Conversion_emissions = result

"""### Export terminal"""

# Electricity consumption for the export terminal [kWh/kg_h2]
el_et_nh3 = float(tea_lnh3.loc['Export Terminal - Electricity use [kWh/kgH2]']['NH3'])
el_et_nh3

el_reliq_nh3 = float(tea_lnh3.loc['Export Terminal - Energy use to reliquefy opt. [kWh/kg NH3/day]']['NH3'])
el_reliq_nh3

#Lower heating value of Nh3[ kWh / kg]
NH3_lhv = float(tea_lnh3.loc['Properties - LHV [kWh/kg]']['NH3'])
NH3_lhv
#Lower heating value of H2 [kWh/kg]
H2_lhv = 33.33

"""same storage time as for LH2"""

# Average storage time in the export terminal tanks in days
t_et

def ET_emissions():
    result = (el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_y_n
    return result

# Calc. liquefaction emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Export_terminal_emissions'])
result.index.name = 'Years'

for year in years:
    EF_y_n = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Norway'][year])

    result.NH3_Export_terminal_emissions.loc[year] = ET_emissions()

result

writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

path_csv = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'NH3_Export_terminal_emissions.csv')
result.to_csv(output_file, sep=';')

NH3_Export_terminal_emissions = result

"""### Shipping

Emissions of transported hydrogen carrier in [g CO2/kg_h2] in year y
NH3_Conversion_emissions = pd.read_csv(
    "/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Conversion_emissions.csv", delimiter=';',
    decimal=',', index_col=0)
NH3_Export_terminal_emissions = pd.read_csv(
    "/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Export_terminal_emissions.csv", delimiter=';',
    decimal=',', index_col=0)
"""

# Boil-off when shipping in [%/day] to [%/hour]
bog_ship = float(tea_lnh3.loc['Shipping - Boil off opt. [%/day]']['NH3']) / 24 * (NH3_lhv/H2_lhv)
bog_ship

# Fuel consumption of a ship in [kg_h2/t/km]
f_ship = float(tea_lnh3.loc['Shipping - Fuel use [kg NH3/t/km]']['NH3']) / 1000 * (NH3_lhv/H2_lhv)
f_ship

"""d_sea and v_ship are equivalent to LH2 transport

Calc. emission factor for cargo/fuel
"""

# Emissions of transported hydrogen carrier in [g CO2eq/kg_H2] in year y
H2_emissions = (min_prod_cost.loc[::]['Emissions']).apply(pd.to_numeric) * 1000
Con_emissions = (NH3_Conversion_emissions.loc[::]['NH3_Conversion_emissions']).apply(pd.to_numeric)
ET_emissions = (NH3_Export_terminal_emissions.loc[::]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
# Emission of cargo in [g CO2eq/kg_H2]
NH3_cargo_emissions = (H2_emissions + Con_emissions + ET_emissions).to_frame('NH3_cargo_emissions')
NH3_cargo_emissions

output_file = os.path.join(path_csv, 'NH3_cargo_emissions.csv')
result.to_csv(output_file, sep=';')

#d_sea = 10000

# Calc.NH3 shipping emissions
def calculate_ship_emissions():
    result = 1 / (1 - (bog_ship * d_sea / v_ship) - (f_ship * d_sea))\
             + (bog_ship * d_sea / v_ship + f_ship * d_sea) * NH3_cargo_ghg
    return result

# Calculation of shipping emissions [g CO2eq/kg_h2] from 2025 to 2050.
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Shipping_emissions'])
result.index.name = 'Years'

for year in years:
    NH3_cargo_ghg = float(LH2_cargo_emissions.loc[year]['LH2_cargo_emissions'])
    f_ship = float(tea_lnh3.loc['Shipping - Fuel use [kg NH3/t/km]'][year])/1000 * (NH3_lhv/H2_lhv)
    # calculate emissions of specific year
    result.NH3_Shipping_emissions.loc[year] = calculate_ship_emissions()

result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

path_csv = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'NH3_Shipping_emissions.csv')
result.to_csv(output_file, sep=';')

NH3_Shipping_emissions = result

"""### Import terminal"""

# Electricity consumption for the import terminal [kWh/kg_h2]
el_it = float(tea_lnh3.loc['Import Terminal - Electricity use [kWh/kg H2]']['NH3'])
el_it

# Electricity consumption to reliquefy BOG [kWh/kg_h2]
el_reliq = float(tea_lnh3.loc['Import Terminal - Energy use to reliquefy opt. [kWh/kg NH3/day]']['NH3'])
el_reliq

# Average storage time in the import terminal tanks in days
t_it = float(tea_lnh3.loc['Import Terminal - Storage length per load [days]']['NH3'])
t_it

def IT_emissions():
    result = (el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * EF_y_n
    return result

# Calc. import terimal emissions over time [g CO2eq/kg H2]
years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Import_terminal_emissions'])
result.index.name = 'Years'

for year in years:
    EF_y_n = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Germany'][year])

    result.NH3_Import_terminal_emissions.loc[year] = IT_emissions()

result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

path_csv = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'NH3_Import_terminal_emissions.csv')
result.to_csv(output_file, sep=';')

NH3_Import_terminal_emissions = result

"""### Reconversion"""

def calculate_recon_emissions():
    result = (heat_recon_y + el_recon_y) * EF_y_n
    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_Reconversion_emissions'])
result.index.name = 'Years'

for year in years:
    el_recon_y = float(tea_lnh3.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
    heat_recon_y = float(tea_lnh3.loc['Reconversion - Heat consumption opt. [kWh/kg H2]'][year])
    EF_y_n = float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Germany'][year])

     # calculate emissions of specific year
    result.NH3_Reconversion_emissions.loc[year] = calculate_recon_emissions()
result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

path_csv = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'NH3_Reconversion_emissions.csv')
result.to_csv(output_file, sep=';')

NH3_Reconversion_emissions = result

"""### Total NH3 transport emissions

con_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Conversion_emissions.csv',
                     delimiter=';', decimal=',', index_col=0)
et_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Export_terminal_emissions.csv',
                    delimiter=';', decimal=',', index_col=0)
ship_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Shipping_emissions.csv',
                      delimiter=';', decimal=',', index_col=0)
it_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Import_terminal_emissions.csv',
                    delimiter=';', decimal=',', index_col=0)
recon_em = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/NH3_Reconversion_emissions.csv',
                       delimiter=';', decimal=',', index_col=0)
"""

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['NH3_transport_emissions'])
result.index.name = 'Years'

# [g CO2eq/kg H2]
def calculate_NH3_transport_emissions():
    result = CE + EE + SE + IE + RE

    return result

for year in years:
    # get all emissions
    CE = float(NH3_Conversion_emissions.loc[year]['NH3_Conversion_emissions'])
    EE = float(NH3_Export_terminal_emissions.loc[year]['NH3_Export_terminal_emissions'])
    SE = float(NH3_Shipping_emissions.loc[year]['NH3_Shipping_emissions'])
    IE = float(NH3_Import_terminal_emissions.loc[year]['NH3_Import_terminal_emissions'])
    RE = float(NH3_Reconversion_emissions.loc[year]['NH3_Reconversion_emissions'])
    # calculate costs of specific year
    result.NH3_transport_emissions.loc[year] = calculate_NH3_transport_emissions()

result

NH3_transport_emissions = result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

# Emission breakdown for NH3 seaborne transport
fig, ax = plt.subplots(figsize=(10, 6))
plt.grid(True, axis='y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)
CE = (NH3_Conversion_emissions.loc[::5]['NH3_Conversion_emissions']).apply(pd.to_numeric)
EE = (NH3_Export_terminal_emissions.loc[::5]['NH3_Export_terminal_emissions']).apply(pd.to_numeric)
SE = (NH3_Shipping_emissions.loc[::5]['NH3_Shipping_emissions']).apply(pd.to_numeric)
IE = (NH3_Import_terminal_emissions.loc[::5]['NH3_Import_terminal_emissions']).apply(pd.to_numeric)
RE = (NH3_Reconversion_emissions.loc[::5]['NH3_Reconversion_emissions']).apply(pd.to_numeric)
width = 2  # the width of the bars: can also be len(x) sequence

RE_plt = plt.bar(x, RE, width, label='Reconversion (Ammonia cracking) emissions', bottom=CE + EE + SE + IE)
IE_plt = plt.bar(x, IE, width, label='Import terminal emissions', bottom=CE + EE + SE)
SE_plt = plt.bar(x, SE, width, label='Shipping emissions', bottom=CE + EE)
EE_plt = plt.bar(x, EE, width, label='Export terminal emissions', bottom=CE)
CE_plt = plt.bar(x, CE, width, label='Conversion (Ammonia synthesis) emissions')

#plt.title('Emission breakdown for NH3 shipping [g CO2eq/kg H2]', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('[g CO2eq/kg H2]')

title = 'NH3_emissions_w_reconversion'
plt.savefig(path_plt+title+'.png', transparent = True)

plt.show()

"""Largest contributer is reconversion at around 15 kWh/kg H2 are required to crack ammonia and it is assumed that electricity and heat are sourced from grid electricity, which has a high emission factor of around 300 g CO2eq/kWh in 2025 (in Germany).

### Sensitivity
"""

year = 2030

el_con_y = float(tea_lnh3.loc['Conversion - Electricity consumption opt. [kWh/kgH2]'][year])
NH3_cargo_ghg = float(LH2_cargo_emissions.loc[year]['LH2_cargo_emissions'])
f_ship = float(tea_lnh3.loc['Shipping - Fuel use [kg NH3/t/km]'][year])/1000 * (NH3_lhv/H2_lhv)
el_recon_y = float(tea_lnh3.loc['Reconversion - Electricity consumption opt. [kWh/kg H2]'][year])
heat_recon_y = float(tea_lnh3.loc['Reconversion - Heat consumption opt. [kWh/kg H2]'][year])
EF_y_n = 118  # float(GHG.loc['GHG intensity of electricity generation [g CO2eq/kWh] - Norway'][year])

EF_sensi = np.arange(0,301,20)
sensitivity = []

def nh3_transport_sensi_EF(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship, d_sea, v_ship, f_ship, NH3_cargo_ghg, el_it, heat_recon_y, t_it, el_recon_y):

    for EF_y_n in EF_sensi:

        result =\
            (el_con_y * EF_y_n )\
        + ((el_et_nh3 + el_reliq_nh3 * (NH3_lhv/H2_lhv) * t_et) * EF_y_n )\
        + (1 / (1 - (bog_ship * d_sea / v_ship) - (f_ship * d_sea)) + (bog_ship * d_sea / v_ship + f_ship * d_sea) * NH3_cargo_ghg)\
        + ((el_it + el_reliq * (NH3_lhv/H2_lhv) * t_it) * EF_y_n )\
        + ((heat_recon_y + el_recon_y) * EF_y_n )

        sensitivity.append(result)


    return sensitivity

nh3_transport_sensi_EF(el_con_y, EF_y_n, el_et_nh3, el_reliq_nh3, t_et, bog_ship, d_sea, v_ship, f_ship, NH3_cargo_ghg, el_it, heat_recon_y, t_it, el_recon_y)

nh3_transport_sensi_EF = pd.DataFrame(sensitivity, EF_sensi, columns=['NH3 transport emission [g CO2eq/kg H2]'])
nh3_transport_sensi_EF.index.name = 'Electricity emission [g CO2eq/KWh]]'
nh3_transport_sensi_EF

#nh3_transport_sensi_wo_recon = pd.DataFrame(sensitivity, EF_sensi, columns=['NH3 transport emission w/o recon [g CO2eq/kg H2]'])
nh3_transport_sensi_wo_recon.index.name = 'Electricity emission factor [g CO2eq/kWh]'
nh3_transport_sensi_wo_recon

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'NH3_EF_sensi_wo_recon.csv')
nh3_transport_sensi_wo_recon.to_csv(output_file, sep=';')

"""## Plots"""

fig, ax = plt.subplots(figsize=(10,4))
#plt.subplot(1,2,1)


plt.plot(lh2_transport_sensi_EF, color='navy', linestyle='solid', label = 'LH2')
plt.plot(nh3_transport_sensi_EF, color='darkorange',linestyle='-', label = 'NH3')
plt.plot(nh3_transport_sensi_wo_recon, color='darkorange',linestyle='--', label = 'NH3 w/o cracking')
plt.plot(Pipeline_emissions_sensi_EF, color='slateblue',linestyle='-', label = 'Pipeline')

plt.axvline(x=30, color='grey', linestyle = '-')
plt.axvline(x=275, color='grey', linestyle = '-')
plt.axvline(x=118, color='grey', linestyle = '-')
plt.text(30,6000, 'Norway 2022', horizontalalignment='center', verticalalignment='center')
plt.text(275,6000, 'EU 2021', horizontalalignment='center', verticalalignment='center')
plt.text(118,6000, 'EU 2030', horizontalalignment='center', verticalalignment='center')

plt.grid(True, axis='y')
#plt.grid(True, axis='x')
ax.set_axisbelow(True)
plt.locator_params(axis='x', nbins=12)
plt.ylabel('[g CO2eq/kg H2]')
plt.xlabel('Electricity emission factor [g CO2eq/kWh]')
#plt.ylim(0,3000)
plt.legend(loc = 'upper left')



title = 'Transport_EF_sensi'
plt.savefig(path_plt + title + '.png', transparent=True)

plt.show()

"""# Supply cost & emissions

## Min. production cost
"""

def choose_minimal_production_costs():
    result = min(LCOH_green, LCOH_blue)

    return result

def choose_minimal_production_cost_technology():
    if min(LCOH_green, LCOH_blue) == LCOH_green:
        result = 'Green'
    else:
        result = 'Blue'

    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Minimal_production_costs', 'Technology'])
result.index.name = 'Years'
for year in years:
    # get all costs

    LCOH_green = float(lowest_green.loc[year]['Norway_Onshore_2_low_temp_optimistic'])
    LCOH_blue = float(lcoh_blue.loc[year]['LCOH_blue'])

    # calculate costs of specific year
    result.Minimal_production_costs.loc[year] = choose_minimal_production_costs()
    result.Technology.loc[year] = choose_minimal_production_cost_technology()

result

min_prod_cost = result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
#writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Minimal_production_costs_technology.csv')
result.to_csv(output_file, sep=';')

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Minimal_production_costs'])
result.index.name = 'Years'

for year in years:
    # get all costs

    LCOH_green = float(lowest_green.loc[year]['Norway_Onshore_2_low_temp_optimistic'])
    LCOH_blue = float(lcoh_blue.loc[year]['LCOH_blue'])

    # calculate costs of specific year
    result.Minimal_production_costs.loc[year] = choose_minimal_production_costs()



result

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'Minimal_production_costs.csv')
result.to_csv(output_file, sep=';')

"""### Plot production costs"""



# Plot cost curves of hydrogen production from NGR with CCS and RES
fig, ax = plt.subplots(figsize=(10, 6))

plt.plot(lowest_green, color='green', linestyle='solid', label='Green hydrogen')
plt.plot(lcoh_blue, color='blue', linestyle='solid', label='Blue hydrogen')
plt.grid(True, axis='y')
ax.set_axisbelow(True)
#plt.title('Cost curves for green and blue hydrogen production', fontweight='bold')
ax.legend()
plt.ylabel('[€/kg H2]', fontweight='bold')
plt.xticks(ticks=None, labels=np.arange(2025,2051,5))

title = 'Green_and_blue_costs'
plt.savefig(path_plt+title+'.png', transparent = True)


plt.show()

# Plot cost curve for production costst
fig, ax = plt.subplots(figsize=(10,6))
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='black', linestyle='solid')
plt.title('Minimal production costs in €/kg_H2')
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

"""## Respective production emissions"""

blue_emissions = pd.DataFrame(GHG.loc['Blue hydrogen emissions [kg CO2/kg H2] - Norway', 2025:2050])
blue_emissions.index.name = 'Years'
blue_emissions.columns= ['Emissions']
blue_emissions

"""
Emissions from blue hydrogen including h2 leakage
result = x.add(1*blue_prod_leakage*GWP20_H2)
result"""

# Create csv file from results dataframe
path_csv = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path_csv, 'Blue_production_emissions_H2_leakage.csv')
result.to_csv(output_file, sep=';')

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

#Green_production_emissions = pd.read_csv('/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/Green_production_emissions.csv', delimiter = ';', index_col=0)
Green_production_emissions

min_prod_cost = pd.read_csv(r"\\dena.de\Daten\Home\Reuter\Desktop\H2_pathways_repo\data\processed\Minimal_production_costs_technology.csv", delimiter = ';', index_col=0)
min_prod_cost

years = np.arange(2025, 2051)
emissions = []
def append_emissions():
    for year in years:

        if min_prod_cost.loc[year]['Technology'] == 'Green':
            result = 0
        else:
            result = blue_emissions.loc[year]['Emissions']

        emissions.append(result)

    return emissions

append_emissions()

# Commented out IPython magic to ensure Python compatibility.
min_prod_cost_emissions = min_prod_cost['Emissions'] = emissions
min_prod_cost
# %store min_prod_cost

# Create csv file from results dataframe
output_file = os.path.join(path_csv, 'min_prod_cost_emissions.csv')
min_prod_cost.to_csv(output_file, sep=';')

"""## Transport inputs"""

availabilities = pd.read_excel(path, sheet_name='General Assumptions', decimal=',', index_col=0)
availabilities

this_year = int(availabilities.loc['Technology availability']['Value'])
this_year

AV_pipe_new = int(availabilities.loc['New H2 pipeline - time from FID to commissioning [years]']['Value'])
AV_pipe_new

AV_pipe_retro = int(availabilities.loc['Retrofit H2 pipeline - availability from now [years]']['Value'])
AV_pipe_retro

AV_pipe_co2 = int(availabilities.loc['New CO2 pipeline - time from FID to commissioning [years]']['Value'])
AV_pipe_co2

AV_LH2 = int(availabilities.loc['LH2 shipping - availability']['Value'])
AV_LH2

AV_NH3 = int(availabilities.loc['Ammonia shipping - availability']['Value'])
AV_NH3

pipe_retro_off = pd.read_csv("/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/Retrofit_pipeline_costs_off.csv", delimiter=';', index_col= 0)
pipe_retro_off

pipe_new_off = pd.read_csv("/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/New_pipeline_costs_off.csv", delimiter=';', index_col=0)
pipe_new_off

LH2 = pd.read_csv("/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/Lh2_transport_costs.csv", delimiter=';', index_col=0)
LH2

NH3 = pd.read_csv("/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/LNH3_transport_costs.csv", delimiter=';', index_col=0)
NH3

"""## Minimal transport costs"""

def choose_minimal_transport_costs():

    if year < this_year + AV_pipe_new:
        result = min(TC_LH2, TC_NH3)

    elif year < this_year + AV_pipe_retro:
        result = min(TC_pipe_new, TC_LH2, TC_NH3)
    else:
        result = min(TC_pipe_new, TC_pipe_retro, TC_LH2, TC_NH3)


    return result

def choose_minimal_transport_costs_tech():


    if  year < this_year + AV_pipe_new:
        if min(TC_LH2, TC_NH3) == TC_LH2:
            result = 'LH2'
        else: result = 'NH3'

    elif year < this_year + AV_pipe_retro:
        if min(TC_pipe_new, TC_LH2, TC_NH3) == TC_LH2:
            result = 'LH2'
        elif min(TC_pipe_new, TC_LH2, TC_NH3) == TC_NH3:
            result = 'NH3'
        else:
            result = 'New pipeline'
    else:
        if min(TC_pipe_new, TC_pipe_retro, TC_LH2, TC_NH3) == TC_LH2:
            result = 'LH2'
        elif min(TC_pipe_new, TC_pipe_retro, TC_LH2, TC_NH3) == TC_NH3:
            result = 'NH3'
        elif (TC_pipe_new, TC_pipe_retro, TC_LH2, TC_NH3) == TC_pipe_new:
            result = 'New pipeline'
        else:
            result = 'Retrofit pipeline'


    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Minimal_offshore_transport_costs', 'Technology'])
result.index.name = 'Years'

for year in years:

    # get all costs
    TC_pipe_new = float(pipe_new_off.loc[year]['New_Pipeline_costs_off'])

    TC_pipe_retro = float(pipe_retro_off.loc[year]['Retrofit_pipeline_costs_off'])

    TC_LH2 = float(LH2.loc[year]['LH2_transport_costs'])

    TC_NH3 = float(NH3.loc[year]['LNH3_transport_costs'])

    result.Minimal_offshore_transport_costs.loc[year] = choose_minimal_transport_costs()
    result.Technology.loc[year] = choose_minimal_transport_costs_tech()

result

# export result to excel
writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
book = load_workbook(path)
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

ws = book['Reference Results']

result.to_excel(writer, sheet_name='Reference Results', index=False, startcol=ws.max_column)
book.save(path)
book.close()

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Minimal_offshore_transport_costs'])
result.index.name = 'Years'

for year in years:

    # get all costs
    TC_pipe_new = float(pipe_new_off.loc[year]['New_Pipeline_costs_off'])

    TC_pipe_retro = float(pipe_retro_off.loc[year]['Retrofit_pipeline_costs_off'])

    TC_LH2 = float(LH2.loc[year]['LH2_transport_costs'])

    TC_NH3 = float(NH3.loc[year]['LNH3_transport_costs'])

    result.Minimal_offshore_transport_costs.loc[year] = choose_minimal_transport_costs()

result

# Create csv file from results dataframe
path = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/Results'
output_file = os.path.join(path, 'Results.xlsx')
result.to_excel(output_file, sheet_name='Minimal_offshore_transport_costs')

"""### Plot transport costs"""

# Plot cost curve for seaborne transport
fig, ax = plt.subplots(figsize=(10,6))
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
plt.plot(result, color='cyan', linestyle='dashed')
plt.title('Cost curve for transport')
plt.legend(['Minimal transport costs in €/kg_H2'])
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

# Plot cost curves of hydrogen transport
fig, ax = plt.subplots(figsize=(10, 6))
plt.plot(LH2, color='blue', linestyle='solid', label='LH2')
plt.plot(NH3, color='yellow', linestyle='solid', label='NH3')
plt.plot(pipe_new_off, color='violet', linestyle='solid', label='New pipeline')
plt.plot(pipe_retro_off, color='pink', linestyle='solid', label='Retrofit pipeline')

plt.axvline(x=this_year + AV_pipe_new, color='violet', linestyle = '--', label= 'New pipeline available')
plt.axvline(x=this_year + AV_pipe_retro, color='pink', linestyle = '--', label= 'Retrofit pipeline available')


plt.grid(True, axis='y')
ax.set_axisbelow(True)
plt.title('Hydrogen transport costs [€/kg H2]', fontweight='bold')
ax.legend()
plt.ylabel('Cost')

plt.show()



"""## Minimal supply costs

"""

mpc = pd.read_csv("/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/Minimal_production_costs.csv", delimiter=';',
    decimal=',', index_col=0)
mpc

mtc = pd.read_csv("/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed/Minimal_offshore_transport_costs.csv", delimiter=';',
    decimal=',', index_col=0)
mtc

def calculate_total_supply_costs():
    result = PC + TC

    return result

years = np.arange(2025, 2051)
result = pd.DataFrame(index=years, columns=['Total_supply_costs'])
result.index.name = 'Years'

for year in years:
    # get all costs

    PC = float(mpc.loc[year]['Minimal_production_costs'])
    TC = float(mtc.loc[year]['Minimal_offshore_transport_costs'])


    # calculate costs of specific year
    result.Total_supply_costs.loc[year] = calculate_total_supply_costs()

result

# Create csv file from results dataframe
path = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/processed'
output_file = os.path.join(path, 'Minimal_supply_costs.csv')
result.to_csv(output_file, sep=';')

path = r'/Users/jakob/PycharmProjects/H2_pathways_repo/data/Results'
output_file = os.path.join(path, 'Results.xlsx')
result.to_excel(output_file, sheet_name='Minimal_supply_costs')

# Plot cost curve for total supply costs
fig, ax = plt.subplots(figsize=(10, 6))
plt.grid(True, axis='y')
ax.set_axisbelow(True)
plt.plot(result, color='cyan', linestyle='dashed')
plt.title('Supply costs costs in €/kg_H2')
plt.xlabel('Year')
plt.ylabel('Cost')
plt.show()

# Cost breakdown for H2 supply
fig, ax = plt.subplots(figsize=(10,6))
plt.grid(True, axis = 'y')
ax.set_axisbelow(True)
x = np.arange(2025, 2051, step=5)

PC = (mpc.loc[::5]['Minimal_production_costs']).apply(pd.to_numeric)
TC = mtc.loc[::5]['Minimal_offshore_transport_costs'].apply(pd.to_numeric)
width = 2       # the width of the bars: can also be len(x) sequence


TC_plt = plt.bar(x,TC, width, label='Transport', bottom= PC)
PC_plt = plt.bar(x,PC, width, label='Production')

plt.title('Lowest H2 supply costs from Norway to Germany', fontweight='bold')
plt.legend(loc='upper right')
plt.ylabel('Cost [€/kg H2]')
plt.xlabel('Years')
plt.show()

mtc